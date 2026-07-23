"""
USt-Firmenwagen Buchungshistorie 06/2026
Erstellt eine Excel-Datei aus der Orchestrator-Datenbank.
Ausführen: python ust_fw_historie_06_2026.py
"""
import sqlite3
import json
import os
import sys
from datetime import datetime

# ── Datenbank-Pfad ──────────────────────────────────────────────────────────
def get_db_path():
    cfg_path = os.path.join(os.path.dirname(__file__), 'orchestrator', 'db_config.json')
    if os.path.exists(cfg_path):
        with open(cfg_path, encoding='utf-8') as f:
            cfg = json.load(f)
        return cfg.get('db_path', '')
    return r'Z:\DB\orchestrator.db'

DB_PATH = get_db_path()

# ── Daten laden ─────────────────────────────────────────────────────────────
def load_data():
    if not os.path.exists(DB_PATH):
        print(f"FEHLER: Datenbank nicht gefunden: {DB_PATH}")
        sys.exit(1)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Alle Läufe für 06/2026
    runs = conn.execute("""
        SELECT id, periode, post_date, bukrs, fis_period, fisc_year,
               ref_doc, header_txt, sap_system, total_amount, row_count,
               status, created_by, created_at
        FROM ust_firmenwagen_runs
        WHERE fis_period IN ('6','06') AND fisc_year='2026'
           OR periode LIKE '06/2026%'
           OR periode LIKE '%06.2026%'
        ORDER BY id DESC
    """).fetchall()

    result = []
    for r in runs:
        r_dict = dict(r)
        lines = conn.execute("""
            SELECT persnr, nachname, kostenstelle, gvw, fahrten, amount,
                   status, belnr
            FROM ust_firmenwagen_lines
            WHERE run_id = ?
            ORDER BY id
        """, (r['id'],)).fetchall()
        r_dict['lines'] = [dict(l) for l in lines]
        result.append(r_dict)

    conn.close()
    return result

# ── Excel erstellen ─────────────────────────────────────────────────────────
def create_excel(runs):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                     Border, Side, numbers)
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("FEHLER: openpyxl nicht installiert.")
        print("Bitte ausführen: pip install openpyxl")
        sys.exit(1)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Farben ──
    HDR_BG   = "1E293B"
    HDR_FG   = "FFFFFF"
    SUB_BG   = "334155"
    BOOKED   = "D1FAE5"
    ERROR_C  = "FEE2E2"
    PENDING  = "FEF9C3"
    ALT_ROW  = "F8FAFC"
    BORDER_C = "CBD5E1"

    thin = Border(
        left=Side(style='thin', color=BORDER_C),
        right=Side(style='thin', color=BORDER_C),
        top=Side(style='thin', color=BORDER_C),
        bottom=Side(style='thin', color=BORDER_C),
    )

    def hdr_style(cell, text, bg=HDR_BG, fg=HDR_FG, bold=True, size=11):
        cell.value = text
        cell.font = Font(name='Calibri', bold=bold, color=fg, size=size)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin

    def data_style(cell, val, align='left', bold=False, bg=None, fmt=None):
        cell.value = val
        cell.font = Font(name='Calibri', bold=bold, size=10)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = thin
        if bg:
            cell.fill = PatternFill('solid', start_color=bg)
        if fmt:
            cell.number_format = fmt

    # ── Übersichtsblatt ──────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Übersicht 06-2026")
    ws_sum.sheet_view.showGridLines = False
    ws_sum.row_dimensions[1].height = 35
    ws_sum.row_dimensions[2].height = 20

    # Titel
    ws_sum.merge_cells('A1:I1')
    tc = ws_sum['A1']
    tc.value = "USt-Firmenwagen Buchungshistorie — Periode 06/2026"
    tc.font = Font(name='Calibri', bold=True, size=14, color=HDR_FG)
    tc.fill = PatternFill('solid', start_color=HDR_BG)
    tc.alignment = Alignment(horizontal='center', vertical='center')

    ws_sum.merge_cells('A2:I2')
    dc = ws_sum['A2']
    dc.value = f"Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}  |  DB: {DB_PATH}"
    dc.font = Font(name='Calibri', italic=True, size=9, color="64748B")
    dc.alignment = Alignment(horizontal='center', vertical='center')

    # Spaltenköpfe
    cols = ['Lauf-ID', 'Periode', 'Buchungsdatum', 'BuKr', 'System',
            'Ref.-Beleg', 'Gesamtbetrag (€)', 'Zeilen', 'Status']
    widths = [10, 12, 16, 8, 10, 18, 18, 8, 12]
    for ci, (col, w) in enumerate(zip(cols, widths), 1):
        hdr_style(ws_sum.cell(row=3, column=ci), col)
        ws_sum.column_dimensions[get_column_letter(ci)].width = w

    # Datenzeilen
    total_amount = 0
    total_lines  = 0
    for ri, r in enumerate(runs, 4):
        ws_sum.row_dimensions[ri].height = 18
        bg = ALT_ROW if ri % 2 == 0 else "FFFFFF"
        st = r.get('status', '')
        if st == 'booked':
            bg = BOOKED
        elif st == 'error':
            bg = ERROR_C

        amt = r.get('total_amount') or 0
        total_amount += amt
        total_lines  += r.get('row_count') or 0

        data_style(ws_sum.cell(ri, 1), f"#{r['id']}", 'center', True, bg)
        data_style(ws_sum.cell(ri, 2), r.get('periode', ''), 'center', bg=bg)
        data_style(ws_sum.cell(ri, 3), r.get('post_date', ''), 'center', bg=bg)
        data_style(ws_sum.cell(ri, 4), r.get('bukrs', ''), 'center', bg=bg)
        data_style(ws_sum.cell(ri, 5), r.get('sap_system', ''), 'center', bg=bg)
        data_style(ws_sum.cell(ri, 6), r.get('ref_doc', '') or r.get('header_txt', ''), bg=bg)
        c = ws_sum.cell(ri, 7)
        data_style(c, amt, 'right', bg=bg, fmt='#,##0.00 €')
        data_style(ws_sum.cell(ri, 8), r.get('row_count', 0), 'center', bg=bg)
        data_style(ws_sum.cell(ri, 9), st, 'center', bg=bg)

    # Summenzeile
    sr = len(runs) + 4
    ws_sum.row_dimensions[sr].height = 20
    for ci in range(1, 10):
        ws_sum.cell(sr, ci).fill = PatternFill('solid', start_color=SUB_BG)
        ws_sum.cell(sr, ci).border = thin
    ws_sum.cell(sr, 1).value = "GESAMT"
    ws_sum.cell(sr, 1).font = Font(name='Calibri', bold=True, color=HDR_FG, size=11)
    ws_sum.cell(sr, 1).alignment = Alignment(horizontal='center', vertical='center')
    c_total = ws_sum.cell(sr, 7)
    c_total.value = total_amount
    c_total.font = Font(name='Calibri', bold=True, color=HDR_FG, size=11)
    c_total.alignment = Alignment(horizontal='right', vertical='center')
    c_total.number_format = '#,##0.00 €'
    c_total.border = thin
    ws_sum.cell(sr, 8).value = total_lines
    ws_sum.cell(sr, 8).font = Font(name='Calibri', bold=True, color=HDR_FG, size=11)
    ws_sum.cell(sr, 8).alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.cell(sr, 8).border = thin

    # ── Detailblatt je Lauf ─────────────────────────────────────────────────
    for r in runs:
        sheet_name = f"Lauf #{r['id']}"
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False

        # Titel
        ws.merge_cells('A1:J1')
        t = ws['A1']
        t.value = (f"USt-Firmenwagen Lauf #{r['id']}  |  {r.get('periode','')}  |  "
                   f"{r.get('sap_system','')}  |  Status: {r.get('status','')}")
        t.font = Font(name='Calibri', bold=True, size=12, color=HDR_FG)
        t.fill = PatternFill('solid', start_color=HDR_BG)
        t.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[1].height = 28

        # Meta
        meta = [
            ('Buchungsdatum', r.get('post_date','')),
            ('Buchungskreis', r.get('bukrs','')),
            ('Periode/Jahr', f"{r.get('fis_period','')}/{r.get('fisc_year','')}"),
            ('Ref.-Beleg', r.get('ref_doc','')),
            ('Buchungstext', r.get('header_txt','')),
            ('Erstellt von', r.get('created_by','')),
            ('Erstellt am', (r.get('created_at','') or '').replace('T',' ')[:16]),
        ]
        for mi, (k, v) in enumerate(meta, 2):
            ws.row_dimensions[mi].height = 16
            lc = ws.cell(mi, 1, k)
            lc.font = Font(name='Calibri', bold=True, size=10, color="475569")
            lc.alignment = Alignment(vertical='center')
            vc = ws.cell(mi, 2, v)
            vc.font = Font(name='Calibri', size=10)
            vc.alignment = Alignment(vertical='center')

        # Spalten
        detail_row = len(meta) + 3
        ws.row_dimensions[detail_row].height = 22
        detail_cols = ['#', 'Pers-Nr.', 'Nachname', 'Kostenstelle',
                       'GWV (€)', 'Fahrten (€)', 'Brutto (€)', 'Status', 'Belegnr.']
        detail_w   = [5, 12, 20, 14, 12, 12, 14, 10, 15]
        for ci, (col, w) in enumerate(zip(detail_cols, detail_w), 1):
            hdr_style(ws.cell(detail_row, ci), col, size=10)
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Detailzeilen
        lauf_total = 0
        for li, ln in enumerate(r.get('lines', []), 1):
            dr = detail_row + li
            ws.row_dimensions[dr].height = 16
            st = ln.get('status', '')
            bg = BOOKED if st == 'booked' else (ERROR_C if st == 'error' else
                 (PENDING if st == 'pending' else ALT_ROW))
            amt = ln.get('amount') or 0
            lauf_total += amt

            data_style(ws.cell(dr, 1), li, 'center', bg=bg)
            data_style(ws.cell(dr, 2), ln.get('persnr', ''), 'center', bg=bg)
            data_style(ws.cell(dr, 3), ln.get('nachname', ''), bg=bg)
            data_style(ws.cell(dr, 4), ln.get('kostenstelle', ''), 'center', bg=bg)
            data_style(ws.cell(dr, 5), ln.get('gvw') or 0, 'right', bg=bg, fmt='#,##0.00')
            data_style(ws.cell(dr, 6), ln.get('fahrten') or 0, 'right', bg=bg, fmt='#,##0.00')
            data_style(ws.cell(dr, 7), amt, 'right', bold=True, bg=bg, fmt='#,##0.00 €')
            data_style(ws.cell(dr, 8), st, 'center', bg=bg)
            data_style(ws.cell(dr, 9), ln.get('belnr', ''), 'center', bg=bg)

        # Summe
        sr2 = detail_row + len(r.get('lines', [])) + 1
        ws.row_dimensions[sr2].height = 20
        ws.merge_cells(f'A{sr2}:F{sr2}')
        sc = ws.cell(sr2, 1)
        sc.value = f"Gesamt: {len(r.get('lines',[]))} Positionen"
        sc.font = Font(name='Calibri', bold=True, size=10, color=HDR_FG)
        sc.fill = PatternFill('solid', start_color=SUB_BG)
        sc.alignment = Alignment(horizontal='right', vertical='center')
        sc.border = thin
        tc2 = ws.cell(sr2, 7)
        tc2.value = lauf_total
        tc2.font = Font(name='Calibri', bold=True, size=11, color=HDR_FG)
        tc2.fill = PatternFill('solid', start_color=SUB_BG)
        tc2.alignment = Alignment(horizontal='right', vertical='center')
        tc2.number_format = '#,##0.00 €'
        tc2.border = thin
        for ci in range(8, 10):
            ws.cell(sr2, ci).fill = PatternFill('solid', start_color=SUB_BG)
            ws.cell(sr2, ci).border = thin

    # ── Speichern ────────────────────────────────────────────────────────────
    ts = datetime.now().strftime('%Y%m%d_%H%M')
    out_dir = os.path.dirname(os.path.abspath(__file__))
    out_path = os.path.join(out_dir, f"UstFirmenwagen_Historie_06_2026_{ts}.xlsx")
    wb.save(out_path)
    print(f"\n✅ Excel-Datei erstellt:\n   {out_path}")
    print(f"   {len(runs)} Läufe  |  {total_lines} Positionen  |  {total_amount:,.2f} €")
    return out_path

# ── Main ─────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print(f"DB: {DB_PATH}")
    print("Lade USt-Firmenwagen Daten 06/2026 ...")
    runs = load_data()
    if not runs:
        print("Keine Läufe für 06/2026 gefunden.")
        sys.exit(0)
    print(f"{len(runs)} Lauf/Läufe gefunden.")
    create_excel(runs)
