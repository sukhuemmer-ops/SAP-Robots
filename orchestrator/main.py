"""
SAP Robot Orchestrator
======================
FastAPI-Backend, das das Cockpit (Web-UI) mit den Workern (Windows-VMs mit SAP GUI)
verbindet. Verantwortlichkeiten:

* Persistente Speicherung von Robots, Tasks, Zeitplaenen, Runs (SQLite via SQLAlchemy)
* REST-API fuer das Cockpit (Tasks anlegen, starten, Status abfragen)
* Job-Queue: Worker pollen offene Runs und melden Ergebnisse zurueck
* Scheduler (APScheduler) fuer wiederkehrende Laeufe

Starten (Entwicklung):
    pip install -r requirements.txt
    uvicorn main:app --reload --host 0.0.0.0 --port 8000

API-Doku (automatisch):
    http://localhost:8000/docs
"""
from __future__ import annotations

import logging
import os
import pathlib
from contextlib import asynccontextmanager
from datetime import datetime
from typing import List, Optional

from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from apscheduler.triggers.date import DateTrigger
import io
from fastapi import Depends, FastAPI, File, HTTPException, UploadFile, status
from fastapi.responses import Response

# Verschlüsselung – Modul liegt im Projektstamm (ein Verzeichnis über orchestrator/)
import sys as _sys
import pathlib as _pathlib
_PROJ_ROOT = _pathlib.Path(__file__).parent.parent
if str(_PROJ_ROOT) not in _sys.path:
    _sys.path.insert(0, str(_PROJ_ROOT))
try:
    from crypto import encrypt as _enc, decrypt as _dec  # type: ignore
    _CRYPTO_OK = True
except ImportError:
    _CRYPTO_OK = False
    def _enc(v): return v   # type: ignore
    def _dec(v): return v   # type: ignore
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from sqlalchemy import (Boolean, Column, DateTime, Float, ForeignKey, Integer, String, Text,
                        create_engine, text)
from sqlalchemy.orm import (DeclarativeBase, Mapped, Session, mapped_column,
                            relationship, sessionmaker)

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger("orchestrator")

# ----------------------------------------------------------------------------
# Datenbank — Pfad aus db_config.json oder Umgebungsvariable
# ----------------------------------------------------------------------------
_APP_ENV    = os.getenv("APP_ENV", "prod")
_ORC_DIR    = pathlib.Path(__file__).parent
_CONFIG_FILE = _ORC_DIR / "db_config.json"

import json as _json

def _load_db_path() -> str:
    """Liest den DB-Pfad aus db_config.json oder ORCHESTRATOR_DB-Env-Variable."""
    env_override = os.getenv("ORCHESTRATOR_DB", "").strip()
    if env_override:
        return env_override
    if _CONFIG_FILE.exists():
        try:
            cfg = _json.loads(_CONFIG_FILE.read_text(encoding="utf-8"))
            p = cfg.get("db_path", "").strip()
            if p:
                log.info("DB-Pfad aus db_config.json: %s", p)
                return p
        except Exception as exc:
            log.warning("db_config.json konnte nicht gelesen werden: %s", exc)
    default = str(_ORC_DIR / "orchestrator.db")
    return default

_DB_PATH    = _load_db_path()

# ── Buchungskreis-Registry (SEP & SEQ) ──────────────────────────────────────
BUKRS_LIST = [
    {"code": "VV9",  "name": "Catensys Holding",  "full": "Catensys Holding GmbH"},
    {"code": "0334", "name": "Catensys France",    "full": "Catensys France S.A.S."},
    {"code": "0435", "name": "Catensys Germany",   "full": "Catensys Germany GmbH"},
    {"code": "0436", "name": "Catensys China",     "full": "Catensys China Co. Ltd."},
    {"code": "0437", "name": "Catensys Slovakia",  "full": "Catensys Slovakia s.r.o."},
    {"code": "0438", "name": "Catensys Korea",     "full": "Catensys Korea Co. Ltd."},
    {"code": "0439", "name": "Catensys India",     "full": "Catensys India Pvt. Ltd."},
    {"code": "0440", "name": "Catensys US",        "full": "Catensys US Inc."},
    {"code": "0441", "name": "Catensys Japan",     "full": "Catensys Japan K.K."},
]
BUKRS_CODES = [b["code"] for b in BUKRS_LIST]

DATABASE_URL = f"sqlite:///{_DB_PATH}"
engine = create_engine(
    DATABASE_URL,
    connect_args={"check_same_thread": False, "timeout": 30},
)

# Rechnungen-PDF-Ablage (relativ zum Orchestrator-Verzeichnis → ../rechnungen/)
RECHNUNGEN_DIR = pathlib.Path(__file__).parent.parent / "rechnungen"
RECHNUNGEN_DIR.mkdir(parents=True, exist_ok=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False)


class Base(DeclarativeBase):
    pass


class Robot(Base):
    __tablename__ = "robots"
    id: Mapped[str] = mapped_column(String(8), primary_key=True)  # AP/AR/GL/AA
    name: Mapped[str] = mapped_column(String(64))
    domain: Mapped[str] = mapped_column(String(64))
    tasks: Mapped[List["Task"]] = relationship(back_populates="robot")


class Task(Base):
    __tablename__ = "tasks"
    id: Mapped[int] = mapped_column(primary_key=True)
    robot_id: Mapped[str] = mapped_column(ForeignKey("robots.id"))
    name: Mapped[str] = mapped_column(String(128))
    tcode: Mapped[str] = mapped_column(String(64))           # z. B. "F110" oder "BAPI_ACC_DOCUMENT_POST"
    method: Mapped[str] = mapped_column(String(16))          # GUI | BAPI | OData | Batch
    description: Mapped[str] = mapped_column(Text, default="")
    parameters: Mapped[str] = mapped_column(Text, default="{}")  # JSON-String mit Default-Parametern
    robot: Mapped[Robot] = relationship(back_populates="tasks")
    runs: Mapped[List["Run"]] = relationship(back_populates="task", cascade="all, delete-orphan")


class Schedule(Base):
    __tablename__ = "schedules"
    id: Mapped[int] = mapped_column(primary_key=True)
    task_id: Mapped[int] = mapped_column(ForeignKey("tasks.id"))
    cron: Mapped[str] = mapped_column(String(64))            # z. B. "0 6 * * *"
    active: Mapped[int] = mapped_column(Integer, default=1)


class Run(Base):
    __tablename__ = "runs"
    id: Mapped[int] = mapped_column(primary_key=True)
    task_id: Mapped[int] = mapped_column(ForeignKey("tasks.id"))
    status: Mapped[str] = mapped_column(String(16), default="queued")   # queued|running|ok|error
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    started_at: Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    finished_at: Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    worker_id: Mapped[Optional[str]] = mapped_column(String(64), nullable=True)
    log: Mapped[str] = mapped_column(Text, default="")
    payload: Mapped[str] = mapped_column(Text, default="{}")  # JSON-Parameter fuer diesen Lauf
    task: Mapped[Task] = relationship(back_populates="runs")


class ZinsenSchedule(Base):
    """Monatlicher Buchungsplan fuer Zinsen (genau ein Eintrag, id=1)."""
    __tablename__ = "zinsen_schedule"
    id: Mapped[int]        = mapped_column(primary_key=True)
    cron: Mapped[str]      = mapped_column(String(64), default="30 7 30 * *")
    sap_system: Mapped[str]= mapped_column(String(8),  default="SEQ")
    active: Mapped[int]    = mapped_column(Integer,    default=1)


class ZinsenRunOnce(Base):
    """Einmalige Zinsbuchung zu einem bestimmten Zeitpunkt."""
    __tablename__ = "zinsen_run_once"
    id: Mapped[int]          = mapped_column(primary_key=True)
    fire_at: Mapped[datetime]= mapped_column(DateTime)
    sap_system: Mapped[str]  = mapped_column(String(8),  default="SEQ")
    periode: Mapped[str]     = mapped_column(String(8),  default="")   # z. B. "06/2026"
    status: Mapped[str]      = mapped_column(String(16), default="pending")  # pending|running|ok|error|cancelled
    bapi_jobs: Mapped[str]   = mapped_column(Text, default="[]")   # JSON-Array der BAPI-Aufrufe
    result: Mapped[str]      = mapped_column(Text, default="")     # JSON-Ergebnis nach Ausführung
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class PayrollImport(Base):
    """PayRoll-Import (FIBU LOGA CSV) – gespeichert nach dem Hochladen."""
    __tablename__ = "payroll_imports"
    id: Mapped[int]             = mapped_column(primary_key=True)
    doc_id: Mapped[str]         = mapped_column(String(64))          # LOGA-ID
    run_date: Mapped[str]       = mapped_column(String(32))          # TT.MM.JJJJ
    comp_code: Mapped[str]      = mapped_column(String(8))           # Buchungskreis
    currency: Mapped[str]       = mapped_column(String(4))
    row_count: Mapped[int]      = mapped_column(Integer, default=0)
    file_name: Mapped[str]      = mapped_column(String(256), default="")
    imported_by: Mapped[str]    = mapped_column(String(128), default="")  # SAP-Username
    sap_system: Mapped[str]     = mapped_column(String(16), default="")
    status: Mapped[str]         = mapped_column(String(16), default="imported")
    # imported | validated | booked | reversed | error
    booking_doc_nr: Mapped[str] = mapped_column(String(32), default="")
    storno_doc_nr:  Mapped[str] = mapped_column(String(32), default="")
    created_at: Mapped[datetime]= mapped_column(DateTime, default=datetime.utcnow)
    lines: Mapped[List["PayrollImportLine"]] = relationship(
        back_populates="imp", cascade="all, delete-orphan")


class PayrollImportLine(Base):
    """Einzelne Buchungszeile eines PayRoll-Imports."""
    __tablename__ = "payroll_import_lines"
    id: Mapped[int]           = mapped_column(primary_key=True)
    import_id: Mapped[int]    = mapped_column(ForeignKey("payroll_imports.id"))
    nr: Mapped[int]           = mapped_column(Integer)
    comp_code: Mapped[str]    = mapped_column(String(8))
    currency: Mapped[str]     = mapped_column(String(4))
    pk: Mapped[int]           = mapped_column(Integer)       # 40=Soll, 50=Haben
    account: Mapped[str]      = mapped_column(String(16))    # Sachkonto
    amount: Mapped[str]       = mapped_column(String(32))    # Betrag als String
    costcenter: Mapped[str]   = mapped_column(String(32), default="")
    assignment: Mapped[str]   = mapped_column(String(64), default="")
    item_text: Mapped[str]    = mapped_column(String(256), default="")
    tax_code: Mapped[str]     = mapped_column(String(8), default="")   # SAP-Vorsteuerkennzeichen (P-10)
    imp: Mapped["PayrollImport"] = relationship(back_populates="lines")


class InvoiceRecord(Base):
    """Rechnungsgruppe aus dem Excel-Import (Rechnungserstellungs-Modul)."""
    __tablename__ = "invoice_records"
    id:           Mapped[int]      = mapped_column(primary_key=True)
    group_key:    Mapped[str]      = mapped_column(String(128), unique=True)  # "103054|06/2026"
    kunden_nr:    Mapped[str]      = mapped_column(String(32))
    name:         Mapped[str]      = mapped_column(String(256), default="")
    periode:      Mapped[str]      = mapped_column(String(16))               # "06/2026"
    sap_date:     Mapped[str]      = mapped_column(String(8),  default="")
    leistungsart: Mapped[str]      = mapped_column(String(64), default="Management-Service")
    mwst_code:    Mapped[str]      = mapped_column(String(8),  default="Z4")
    waehrung:     Mapped[str]      = mapped_column(String(4),  default="EUR")
    total:        Mapped[float]    = mapped_column(default=0.0)
    status:       Mapped[str]      = mapped_column(String(32), default="offen")  # offen|erstellt|storniert
    invoice_nr:   Mapped[str]      = mapped_column(String(64), default="")
    order_nr:     Mapped[str]      = mapped_column(String(64), default="")
    positions_json: Mapped[str]    = mapped_column(Text,       default="[]")  # JSON-Array der Positionen
    is_storno:    Mapped[int]      = mapped_column(Integer,    default=0)
    bukrs:        Mapped[str]      = mapped_column(String(4),  default="0435")  # Buchungskreis
    imported_at:  Mapped[datetime] = mapped_column(DateTime,   default=datetime.utcnow)
    updated_at:   Mapped[datetime] = mapped_column(DateTime,   default=datetime.utcnow)


class InvoiceEmailConfig(Base):
    """E-Mail-Konfiguration pro Kundennummer für automatischen Rechnungsversand."""
    __tablename__ = "invoice_email_configs"
    id:               Mapped[int]      = mapped_column(primary_key=True)
    kunden_nr:        Mapped[str]      = mapped_column(String(32), unique=True)
    name:             Mapped[str]      = mapped_column(String(256), default="")
    recipient_email:  Mapped[str]      = mapped_column(String(256))
    sender_email:     Mapped[str]      = mapped_column(String(256))
    smtp_host:        Mapped[str]      = mapped_column(String(256), default="")
    smtp_port:        Mapped[int]      = mapped_column(Integer,     default=587)
    smtp_user:        Mapped[str]      = mapped_column(String(256), default="")
    smtp_pass:        Mapped[str]      = mapped_column(String(256), default="")
    smtp_tls:         Mapped[int]      = mapped_column(Integer,     default=1)   # 1=STARTTLS
    subject_template: Mapped[str]      = mapped_column(Text, default="Rechnung {invoice_nr} – {name} – {periode}")
    body_template:    Mapped[str]      = mapped_column(Text, default="Sehr geehrte Damen und Herren,\n\nim Anhang finden Sie unsere Rechnung {invoice_nr} für den Zeitraum {periode}.\n\nMit freundlichen Grüßen\nCatensys Germany GmbH")
    status:           Mapped[str]      = mapped_column(String(16),  default="aktiv")   # aktiv|inaktiv
    auto_send:        Mapped[int]      = mapped_column(Integer,     default=0)    # 1=auto nach Erstellung
    created_at:       Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)
    updated_at:       Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)


class InvoiceEmailHistory(Base):
    """Versandprotokoll für Rechnungs-E-Mails."""
    __tablename__ = "invoice_email_history"
    id:               Mapped[int]      = mapped_column(primary_key=True)
    invoice_record_id: Mapped[int]     = mapped_column(Integer, default=0)
    kunden_nr:        Mapped[str]      = mapped_column(String(32), default="")
    invoice_nr:       Mapped[str]      = mapped_column(String(64), default="")
    periode:          Mapped[str]      = mapped_column(String(16), default="")
    recipient_email:  Mapped[str]      = mapped_column(String(256))
    sender_email:     Mapped[str]      = mapped_column(String(256))
    subject:          Mapped[str]      = mapped_column(String(512), default="")
    status:           Mapped[str]      = mapped_column(String(16),  default="ausstehend")  # gesendet|fehler|ausstehend
    error_msg:        Mapped[str]      = mapped_column(Text,        default="")
    sent_at:          Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)


class AiConfig(Base):
    """KI-Provider-Konfiguration (Claude / OpenAI) – wird von voice_bot.py geladen."""
    __tablename__ = "ai_config"
    id:         Mapped[int] = mapped_column(primary_key=True)
    provider:   Mapped[str] = mapped_column(String(32),  default="claude")   # claude | openai
    api_key:    Mapped[str] = mapped_column(Text,        default="")
    model:      Mapped[str] = mapped_column(String(128), default="")
    updated_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class SapUser(Base):
    """SAP-Benutzer fuer die Benutzerverwaltung im Cockpit.
    Passwort wird NICHT gespeichert – Eingabe erfolgt zur Laufzeit."""
    __tablename__ = "sap_users"
    id: Mapped[int] = mapped_column(primary_key=True)
    display_name: Mapped[str] = mapped_column(String(128))          # Anzeigename z. B. "Max Mustermann"
    sap_username: Mapped[str] = mapped_column(String(64))           # SAP-Login z. B. "MMUSTERMANN"
    sap_system: Mapped[str] = mapped_column(String(32), default="") # System-ID z. B. "SEP600"
    ashost: Mapped[str] = mapped_column(String(128), default="")    # Application-Server-Host
    sysnr: Mapped[str] = mapped_column(String(4), default="00")     # System-Nummer
    client: Mapped[str] = mapped_column(String(4), default="")      # Mandant z. B. "100"
    lang: Mapped[str] = mapped_column(String(4), default="DE")      # Anmeldesprache
    role: Mapped[str] = mapped_column(String(32), default="Buchung") # Admin | Buchung | Readonly
    notes: Mapped[str] = mapped_column(Text, default="")
    active: Mapped[int] = mapped_column(Integer, default=1)
    created_at: Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


# ============================================================================
# 3W SMART DATABASE ENGINE
# ============================================================================
# Layer 2: 3W-Schicht — WHO did WHAT and WHY (Buchungs-Audit-Trail)
# ============================================================================
class ThreeWAnswer(Base):
    """10W-Schicht: Vollständiges Wissensmodell — Wer/Was/Warum/Wie/Wann/Wo/Womit/Risiken/Alternativen/Erfahrung/Ergebnis."""
    __tablename__ = "three_w_answers"
    id:             Mapped[int]      = mapped_column(primary_key=True)
    # W1 – WER (WHO)
    who:            Mapped[str]      = mapped_column(String(64))
    # W2 – WAS (WHAT)
    action_type:    Mapped[str]      = mapped_column(String(64))
    action_desc:    Mapped[str]      = mapped_column(String(512), default="")
    object_ref:     Mapped[str]      = mapped_column(String(128), default="")
    module:         Mapped[str]      = mapped_column(String(32),  default="")
    # W3 – WARUM (WHY)
    business_rule:  Mapped[str]      = mapped_column(String(64),  default="")
    why_desc:       Mapped[str]      = mapped_column(String(512), default="")
    # W4 – WIE (HOW) — Methode, Prozess, Werkzeug
    wie:            Mapped[str]      = mapped_column(String(512), default="")
    # W5 – WANN (WHEN) — geplant vs. tatsächlich
    when_ts:        Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)
    wann_geplant:   Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    # W6 – WO (WHERE) — System, Modul, Pfad, Buchungskreis
    wo:             Mapped[str]      = mapped_column(String(256), default="")
    sap_system:     Mapped[str]      = mapped_column(String(8),   default="")
    comp_code:      Mapped[str]      = mapped_column(String(8),   default="")
    periode:        Mapped[str]      = mapped_column(String(16),  default="")
    # W7 – WOMIT (WITH WHAT) — BAPI, Werkzeug, Ressourcen
    womit:          Mapped[str]      = mapped_column(String(512), default="")
    # W8 – WELCHE RISIKEN (WHICH RISKS)
    risiken:        Mapped[str]      = mapped_column(Text,        default="")
    # W9 – WELCHE ALTERNATIVEN (WHICH ALTERNATIVES)
    alternativen:   Mapped[str]      = mapped_column(Text,        default="")
    # W10 – WELCHE ERFAHRUNG / WELCHES ERGEBNIS (EXPERIENCE + RESULT)
    erfahrung:      Mapped[str]      = mapped_column(Text,        default="")
    ergebnis:       Mapped[str]      = mapped_column(Text,        default="")
    # META
    status:         Mapped[str]      = mapped_column(String(16),  default="ok")
    metadata_json:  Mapped[str]      = mapped_column(Text,        default="{}")


# ============================================================================
# KNOWLEDGE UNIVERSE — 13-Schichten-Wissensmodell
# ============================================================================
class KnowledgeUniverse(Base):
    """Knowledge Universe: Strukturiertes Wissen in 13 Schichten.
    Jedes Objekt ist kein Datensatz — es ist Wissen mit Kontext, Beziehung und Herkunft."""
    __tablename__ = "knowledge_universe"
    id:            Mapped[int]      = mapped_column(primary_key=True)
    # Klassifikation
    layer:         Mapped[str]      = mapped_column(String(32))
    # business|finance|sap|ai|human|learning|prediction|experience|rule|process|risk|audit|innovation
    object_type:   Mapped[str]      = mapped_column(String(32), default="concept")
    # concept|entity|rule|pattern|decision|risk|process|experience|prediction|fact
    key:           Mapped[str]      = mapped_column(String(256), unique=True)  # eindeutiger Schlüssel
    title:         Mapped[str]      = mapped_column(String(512))
    summary:       Mapped[str]      = mapped_column(String(1024), default="")
    body_json:     Mapped[str]      = mapped_column(Text, default="{}")  # strukturierter Inhalt
    # Wissensqualität
    confidence:    Mapped[int]      = mapped_column(Integer, default=3)   # 1–5
    importance:    Mapped[int]      = mapped_column(Integer, default=3)   # 1–5
    # Herkunft
    source_module: Mapped[str]      = mapped_column(String(32),  default="")
    source_ref:    Mapped[str]      = mapped_column(String(256), default="")
    source_type:   Mapped[str]      = mapped_column(String(32),  default="manual")
    # manual|automatic|sap|document|experience|ai
    created_by:    Mapped[str]      = mapped_column(String(64),  default="system")
    # Kontext
    tags_json:     Mapped[str]      = mapped_column(Text, default="[]")
    valid_from:    Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    valid_until:   Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    active:        Mapped[int]      = mapped_column(Integer, default=1)
    created_at:    Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at:    Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    relations:     Mapped[List["KnowledgeRelation"]] = relationship(
        "KnowledgeRelation", foreign_keys="KnowledgeRelation.from_id",
        back_populates="from_node", cascade="all, delete-orphan")


class KnowledgeRelation(Base):
    """Wissensbeziehungen: Wie Wissen miteinander verbunden ist."""
    __tablename__ = "knowledge_relations"
    id:            Mapped[int]  = mapped_column(primary_key=True)
    from_id:       Mapped[int]  = mapped_column(ForeignKey("knowledge_universe.id"))
    to_id:         Mapped[int]  = mapped_column(ForeignKey("knowledge_universe.id"))
    relation_type: Mapped[str]  = mapped_column(String(32))
    # depends_on|contradicts|supports|derives_from|replaces|extends|causes|prevents
    strength:      Mapped[int]  = mapped_column(Integer, default=3)  # 1–5
    note:          Mapped[str]  = mapped_column(String(512), default="")
    created_at:    Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    from_node:     Mapped["KnowledgeUniverse"] = relationship(
        "KnowledgeUniverse", foreign_keys=[from_id], back_populates="relations")


# ============================================================================
# KNOWLEDGE DECISION ENGINE — Argumentations- und Begründungslogik
# ============================================================================
class DecisionRequest(Base):
    """Decision Engine: Jede Entscheidung wird vollständig begründet.
    Das System argumentiert — es entscheidet nicht blind."""
    __tablename__ = "decision_requests"
    id:               Mapped[int]      = mapped_column(primary_key=True)
    # Frage / Kontext
    question:         Mapped[str]      = mapped_column(Text)
    context_json:     Mapped[str]      = mapped_column(Text, default="{}")
    # Wer fragte / welches Modul
    asked_by:         Mapped[str]      = mapped_column(String(64), default="")
    module:           Mapped[str]      = mapped_column(String(32), default="")
    # Entscheidungsergebnis
    recommendation:   Mapped[str]      = mapped_column(Text, default="")
    # Pro-Argumente
    arguments_pro:    Mapped[str]      = mapped_column(Text, default="[]")    # JSON
    # Contra-Argumente
    arguments_contra: Mapped[str]      = mapped_column(Text, default="[]")    # JSON
    # Risiken
    risks_json:       Mapped[str]      = mapped_column(Text, default="[]")    # JSON
    # Alternativen
    alternatives_json:Mapped[str]      = mapped_column(Text, default="[]")    # JSON
    # Referenziertes Wissen
    knowledge_refs:   Mapped[str]      = mapped_column(Text, default="[]")    # JSON: [knowledge_universe.id]
    # Qualität
    confidence:       Mapped[int]      = mapped_column(Integer, default=3)    # 1–5
    decided_by:       Mapped[str]      = mapped_column(String(16), default="engine")  # engine|human
    # Status
    status:           Mapped[str]      = mapped_column(String(16), default="pending")
    # pending|decided|overridden|archived
    human_override:   Mapped[str]      = mapped_column(Text, default="")
    rationale:        Mapped[str]      = mapped_column(Text, default="")
    # Zeitstempel
    created_at:       Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    decided_at:       Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    # Lerneffekt: wurde die Entscheidung korrekt? (Feedback-Loop)
    outcome_correct:  Mapped[Optional[int]] = mapped_column(Integer, nullable=True)  # 1=korrekt, 0=falsch
    outcome_note:     Mapped[str]      = mapped_column(Text, default="")


# ============================================================================
# Layer 3: Wissensschicht — Buchungsregeln, BAPI-Muster, Fehlerlösungen
# ============================================================================
class KnowledgeItem(Base):
    """Wissensschicht: Strukturiertes Domänenwissen aus dem Buchungsbetrieb."""
    __tablename__ = "knowledge_items"
    id:            Mapped[int]      = mapped_column(primary_key=True)
    category:      Mapped[str]      = mapped_column(String(32))   # rule | bapi | error | note | pattern
    key:           Mapped[str]      = mapped_column(String(128))  # Eindeutiger Schlüssel
    title:         Mapped[str]      = mapped_column(String(256))
    body:          Mapped[str]      = mapped_column(Text, default="")  # Markdown-Inhalt
    source_module: Mapped[str]      = mapped_column(String(32),  default="")  # zinsen | payroll | rechnung
    source_ref:    Mapped[str]      = mapped_column(String(128), default="")  # z.B. "three_w_answers:42"
    tags_json:     Mapped[str]      = mapped_column(Text,        default="[]")  # JSON-Array
    confidence:    Mapped[int]      = mapped_column(Integer,     default=3)   # 1–5
    active:        Mapped[int]      = mapped_column(Integer,     default=1)
    created_at:    Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)
    updated_at:    Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)


# ============================================================================
# Layer 4: Lernschicht — Aggregierte Muster und Statistiken
# ============================================================================
class LearningPattern(Base):
    """Lernschicht: Automatisch erkannte Buchungsmuster aus dem Betrieb."""
    __tablename__ = "learning_patterns"
    id:               Mapped[int]      = mapped_column(primary_key=True)
    pattern_type:     Mapped[str]      = mapped_column(String(64))   # booking_frequency | error_rate | amount_drift
    pattern_key:      Mapped[str]      = mapped_column(String(128))  # z.B. "zinsen:044:monthly_amount"
    description:      Mapped[str]      = mapped_column(String(512), default="")
    occurrence_count: Mapped[int]      = mapped_column(Integer,     default=1)
    last_seen:        Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)
    data_json:        Mapped[str]      = mapped_column(Text,        default="{}")  # Statistik-Rohdaten
    created_at:       Mapped[datetime] = mapped_column(DateTime,    default=datetime.utcnow)


# ============================================================================
# Layer 5: Reorganisationsschicht — Verbesserungsvorschläge (nie automatisch)
# ============================================================================
class StructureSuggestion(Base):
    """Reorganisationsschicht: Strukturvorschläge — werden NIEMALS automatisch umgesetzt."""
    __tablename__ = "structure_suggestions"
    id:           Mapped[int]              = mapped_column(primary_key=True)
    stype:        Mapped[str]              = mapped_column(String(32))   # db_schema | workflow | ui | data_quality
    title:        Mapped[str]              = mapped_column(String(256))
    description:  Mapped[str]             = mapped_column(Text, default="")
    priority:     Mapped[int]             = mapped_column(Integer,  default=3)  # 1=niedrig … 5=kritisch
    status:       Mapped[str]             = mapped_column(String(16), default="open")  # open | accepted | rejected | done
    proposed_by:  Mapped[str]             = mapped_column(String(64), default="system")
    proposed_at:  Mapped[datetime]        = mapped_column(DateTime,  default=datetime.utcnow)
    decided_at:   Mapped[Optional[datetime]] = mapped_column(DateTime, nullable=True)
    notes:        Mapped[str]             = mapped_column(Text, default="")


# ============================================================================
# Layer 6: Governance-Schicht — App-Regeln persistent in der DB
# ============================================================================
class ZtermChangeLog(Base):
    """Protokoll aller Zahlungsbedingungsänderungen (XD02 / VA42)."""
    __tablename__ = "zterm_change_log"
    id:          Mapped[int]      = mapped_column(primary_key=True)
    kunnr:       Mapped[str]      = mapped_column(String(10))
    kname:       Mapped[str]      = mapped_column(String(100), default="")
    bukrs:       Mapped[str]      = mapped_column(String(4), default="")
    tcode:       Mapped[str]      = mapped_column(String(10))
    vbeln:       Mapped[str]      = mapped_column(String(10), default="")
    zterm_old:   Mapped[str]      = mapped_column(String(10), default="")
    zterm_new:   Mapped[str]      = mapped_column(String(10))
    sap_system:  Mapped[str]      = mapped_column(String(10), default="")
    changed_by:  Mapped[str]      = mapped_column(String(50))
    changed_at:  Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    status:      Mapped[str]      = mapped_column(String(10), default="ok")
    note:        Mapped[str]      = mapped_column(Text, default="")


class GovernanceRule(Base):
    """Governance-Schicht: App-Regeln dauerhaft in der DB (Single Source of Truth)."""
    __tablename__ = "governance_rules"
    id:          Mapped[int]      = mapped_column(primary_key=True)
    rule_id:     Mapped[str]      = mapped_column(String(16), unique=True)  # G-5, G-7, Z-5-R, P-10 …
    module:      Mapped[str]      = mapped_column(String(32), default="global")  # global | zinsen | payroll | rechnung
    title:       Mapped[str]      = mapped_column(String(256))
    description: Mapped[str]      = mapped_column(Text, default="")
    priority:    Mapped[int]      = mapped_column(Integer, default=3)  # 1=info … 5=kritisch/sicherheitsrelevant
    active:      Mapped[int]      = mapped_column(Integer, default=1)
    created_at:  Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class ModuleVersion(Base):
    """Versions-Snapshot ganzer Dateien — automatisch nach erfolgreichem Modul-Lauf."""
    __tablename__ = "module_versions"
    id:             Mapped[int]      = mapped_column(primary_key=True)
    module_name:    Mapped[str]      = mapped_column(String(128))   # z.B. "handlers.py"
    version_number: Mapped[int]      = mapped_column(Integer)       # Auto-Zähler pro Modul
    file_path:      Mapped[str]      = mapped_column(String(512))   # absoluter Pfad
    file_content:   Mapped[str]      = mapped_column(Text)          # kompletter Dateiinhalt
    file_hash:      Mapped[str]      = mapped_column(String(64))    # SHA-256 — Duplikat-Schutz
    file_size:      Mapped[int]      = mapped_column(Integer, default=0)
    triggered_by:   Mapped[str]      = mapped_column(String(128), default="")  # Handler-Name
    description:    Mapped[str]      = mapped_column(Text, default="")
    created_at:     Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


class RueckstellungRequest(Base):
    """Rückstellungs-/Abgrenzungsantrag (direkt oder indirekt/Konzern)."""
    __tablename__ = "rueckstellung_requests"
    id:               Mapped[int]      = mapped_column(primary_key=True)
    workflow:         Mapped[str]      = mapped_column(String(16), default="direkt")
    kategorie:        Mapped[str]      = mapped_column(String(64), default="")
    beschreibung:     Mapped[str]      = mapped_column(Text, default="")
    betrag:           Mapped[float]    = mapped_column(default=0.0)
    waehrung:         Mapped[str]      = mapped_column(String(4), default="EUR")
    konto_soll:       Mapped[str]      = mapped_column(String(16), default="")
    konto_haben:      Mapped[str]      = mapped_column(String(16), default="")
    kostenstelle:     Mapped[str]      = mapped_column(String(16), default="")
    bukrs_json:       Mapped[str]      = mapped_column(Text, default="[]")
    periode:          Mapped[str]      = mapped_column(String(16), default="")
    buchungsdatum:    Mapped[str]      = mapped_column(String(10), default="")
    referenz:         Mapped[str]      = mapped_column(String(128), default="")
    sap_system:       Mapped[str]      = mapped_column(String(16), default="SEQ")
    blart:            Mapped[str]      = mapped_column(String(4), default="SA")
    status:           Mapped[str]      = mapped_column(String(32), default="offen")
    requires_approval: Mapped[int]    = mapped_column(Integer, default=0)
    submitted_by:     Mapped[str]      = mapped_column(String(128), default="")
    approved_by:      Mapped[str]      = mapped_column(String(128), default="")
    approval_comment: Mapped[str]      = mapped_column(Text, default="")
    sap_doc_nrs:      Mapped[str]      = mapped_column(Text, default="")
    created_at:       Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at:       Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    approvals: Mapped[List["RueckstellungApproval"]] = relationship(
        back_populates="request", cascade="all, delete-orphan")


class RueckstellungApproval(Base):
    """Freigabe-/Ablehnungs-Protokoll für Rückstellungsanträge."""
    __tablename__ = "rueckstellung_approvals"
    id:         Mapped[int]      = mapped_column(primary_key=True)
    request_id: Mapped[int]      = mapped_column(ForeignKey("rueckstellung_requests.id"))
    action:     Mapped[str]      = mapped_column(String(32))
    actor:      Mapped[str]      = mapped_column(String(128), default="")
    comment:    Mapped[str]      = mapped_column(Text, default="")
    ts:         Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    request:    Mapped["RueckstellungRequest"] = relationship(back_populates="approvals")


class RueckstellungKontenplan(Base):
    """Admin-vordefinierte Sachkonten + Kostenstellen pro Rückstellungs-Kategorie."""
    __tablename__ = "rueckstellung_kontenplan"
    id:          Mapped[int] = mapped_column(primary_key=True)
    kategorie:   Mapped[str] = mapped_column(String(64), unique=True)
    konto_soll:  Mapped[str] = mapped_column(String(16), default="")
    konto_haben: Mapped[str] = mapped_column(String(16), default="")
    kostenstelle: Mapped[str] = mapped_column(String(16), default="")
    blart:       Mapped[str] = mapped_column(String(4), default="SA")
    aktiv:       Mapped[int] = mapped_column(Integer, default=1)
    sort_order:  Mapped[int] = mapped_column(Integer, default=0)


class RueckstellungTemplate(Base):
    """Import-Vorlagen je Buchungskreis + Kategorie (indirekter Workflow)."""
    __tablename__ = "rueckstellung_templates"
    id:           Mapped[int] = mapped_column(primary_key=True)
    bukrs:        Mapped[str] = mapped_column(String(16), default="")   # "" = alle
    kategorie:    Mapped[str] = mapped_column(String(64), default="")   # "" = alle
    name:         Mapped[str] = mapped_column(String(128), default="")
    beschreibung: Mapped[str] = mapped_column(Text, default="")
    spalten_json: Mapped[str] = mapped_column(Text, default="[]")
    aktiv:        Mapped[int] = mapped_column(Integer, default=1)
    created_at:   Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    updated_at:   Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)




class AfaRun(Base):
    """Einmaliger AfA-Lauf (BAPI_AFAPOSTING_POST) zu einem bestimmten Zeitpunkt."""
    __tablename__ = "afa_runs"
    id:          Mapped[int]      = mapped_column(primary_key=True)
    fire_at:     Mapped[datetime] = mapped_column(DateTime)               # geplanter Startzeitpunkt (UTC)
    bukrs:       Mapped[str]      = mapped_column(String(8))              # Buchungskreis
    fiscal_year: Mapped[str]      = mapped_column(String(4))             # Geschäftsjahr
    period:      Mapped[str]      = mapped_column(String(2))             # Periode 01-12
    reason:      Mapped[str]      = mapped_column(String(1), default="1") # 1=Erst, 2=Wdh, 3=Neustart
    testrun:     Mapped[int]      = mapped_column(Integer, default=1)    # 1=Testlauf, 0=Echtlauf
    stop_on_warn:Mapped[int]      = mapped_column(Integer, default=0)
    posting_date:Mapped[str]      = mapped_column(String(8), default="") # YYYYMMDD (leer=auto)
    sap_system:  Mapped[str]      = mapped_column(String(8), default="SEQ")
    status:      Mapped[str]      = mapped_column(String(16), default="pending")
    # pending | running | ok | error | cancelled
    result:      Mapped[str]      = mapped_column(Text, default="")      # JSON-Ergebnis
    created_by:  Mapped[str]      = mapped_column(String(64), default="")
    created_at:  Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)


# ----------------------------------------------------------------------------
# Pydantic-Schemas (API)
# ----------------------------------------------------------------------------
# ── AfaRun Schemas ──────────────────────────────────────────────────────────
class AfaRunIn(BaseModel):
    fire_at:      str               # ISO-Datetime lokale Zeit, z. B. "2026-07-31T23:00:00"
    bukrs:        str               # Buchungskreis
    fiscal_year:  str               # Geschäftsjahr
    period:       str               # Periode 01–12
    reason:       str  = "1"        # 1=Erstbuchung, 2=Wiederholen, 3=Neustart
    testrun:      bool = True
    stop_on_warn: bool = False
    posting_date: str  = ""         # YYYYMMDD; leer = letzter Tag der Periode
    sap_system:   str  = "SEQ"
    created_by:   str  = ""

class AfaRunOut(BaseModel):
    model_config = {"from_attributes": True}
    id:           int
    fire_at:      datetime
    bukrs:        str
    fiscal_year:  str
    period:       str
    reason:       str
    testrun:      int
    stop_on_warn: int
    posting_date: str
    sap_system:   str
    status:       str
    result:       str
    created_by:   str
    created_at:   datetime

# ── PayRoll Schemas ──────────────────────────────────────────────────────────
class PayrollLineIn(BaseModel):
    nr: int
    comp_code: str
    currency: str
    pk: int
    account: str
    amount: float
    costcenter: str = ""
    assignment: str = ""
    item_text: str = ""
    tax_code: str = ""   # SAP-Vorsteuerkennzeichen (App-Regel P-10, nur RESY-Zeilen)


class PayrollImportIn(BaseModel):
    doc_id: str
    run_date: str
    comp_code: str
    currency: str
    row_count: int
    file_name: str = ""
    imported_by: str = ""
    sap_system: str = ""
    lines: List[PayrollLineIn] = []


class PayrollLineUpdate(BaseModel):
    """Felder, die nach dem Import manuell korrigiert werden dürfen."""
    costcenter: Optional[str] = None
    item_text:  Optional[str] = None
    assignment: Optional[str] = None


class PayrollStatusIn(BaseModel):
    status: str          # validated | booked | error | reversed
    booking_doc_nr: str = ""


class PayrollStornoIn(BaseModel):
    storno_doc_nr: str = ""


class InvoicePositionIn(BaseModel):
    buchungstext: str
    betrag: float


class InvoiceRecordIn(BaseModel):
    group_key:    str
    kunden_nr:    str
    name:         str  = ""
    periode:      str
    sap_date:     str  = ""
    leistungsart: str  = "Management-Service"
    mwst_code:    str  = "Z4"
    waehrung:     str  = "EUR"
    total:        float = 0.0
    status:       str  = "offen"
    invoice_nr:   str  = ""
    order_nr:     str  = ""
    positions:    List[InvoicePositionIn] = []
    is_storno:    int  = 0
    bukrs:        str  = "0435"


class InvoiceStatusIn(BaseModel):
    status:     str
    invoice_nr: str = ""
    order_nr:   str = ""


class InvoiceEmailConfigIn(BaseModel):
    kunden_nr:        str
    name:             str  = ""
    recipient_email:  str
    sender_email:     str
    smtp_host:        str  = ""
    smtp_port:        int  = 587
    smtp_user:        str  = ""
    smtp_pass:        str  = ""
    smtp_tls:         int  = 1
    subject_template: str  = "Rechnung {invoice_nr} – {name} – {periode}"
    body_template:    str  = "Sehr geehrte Damen und Herren,\n\nim Anhang finden Sie unsere Rechnung {invoice_nr} für den Zeitraum {periode}.\n\nMit freundlichen Grüßen\nCatensys Germany GmbH"
    status:           str  = "aktiv"
    auto_send:        int  = 0


class SapUserIn(BaseModel):
    display_name: str
    sap_username: str
    sap_system: str = ""
    ashost: str = ""
    sysnr: str = "00"
    client: str = ""
    lang: str = "DE"
    role: str = "Buchung"
    notes: str = ""
    active: int = 1


class SapUserOut(SapUserIn):
    id: int
    created_at: datetime

    class Config:
        from_attributes = True


class TaskIn(BaseModel):
    robot_id: str
    name: str
    tcode: str
    method: str
    description: str = ""
    parameters: str = "{}"


class TaskOut(TaskIn):
    id: int

    class Config:
        from_attributes = True


class RunOut(BaseModel):
    id: int
    task_id: int
    status: str
    created_at: datetime
    started_at: Optional[datetime]
    finished_at: Optional[datetime]
    worker_id: Optional[str]
    log: str

    class Config:
        from_attributes = True


class RunUpdate(BaseModel):
    status: str               # running | ok | error
    log: Optional[str] = None
    worker_id: Optional[str] = None


# ── 10W-Modell Schemas ────────────────────────────────────────────────────────
class ThreeWAnswerIn(BaseModel):
    # W1 WER
    who:           str
    # W2 WAS
    action_type:   str
    action_desc:   str  = ""
    object_ref:    str  = ""
    module:        str  = ""
    business_rule: str  = ""
    why_desc:      str  = ""
    sap_system:    str  = ""
    comp_code:     str  = ""
    periode:       str  = ""
    # W4 WIE
    wie:           str  = ""
    # W5 WANN
    wann_geplant:  Optional[datetime] = None
    # W6 WO
    wo:            str  = ""
    sap_system:    str  = ""
    comp_code:     str  = ""
    periode:       str  = ""
    # W7 WOMIT
    womit:         str  = ""
    # W8 WELCHE RISIKEN
    risiken:       str  = ""
    # W9 WELCHE ALTERNATIVEN
    alternativen:  str  = ""
    # W10 ERFAHRUNG + ERGEBNIS
    erfahrung:     str  = ""
    ergebnis:      str  = ""
    # META
    status:        str  = "ok"
    metadata_json: str  = "{}"


class ThreeWAnswerOut(ThreeWAnswerIn):
    id:      int
    when_ts: datetime
    class Config:
        from_attributes = True


# ── Knowledge Universe Schemas ────────────────────────────────────────────────
class KnowledgeUniverseIn(BaseModel):
    layer:         str
    object_type:   str  = "concept"
    key:           str
    title:         str
    summary:       str  = ""
    body_json:     str  = "{}"
    confidence:    int  = 3
    importance:    int  = 3
    source_module: str  = ""
    source_ref:    str  = ""
    source_type:   str  = "manual"
    created_by:    str  = "system"
    tags_json:     str  = "[]"
    valid_from:    Optional[datetime] = None
    valid_until:   Optional[datetime] = None
    active:        int  = 1


class KnowledgeUniverseOut(KnowledgeUniverseIn):
    id:         int
    created_at: datetime
    updated_at: datetime
    class Config:
        from_attributes = True


class KnowledgeRelationIn(BaseModel):
    from_id:       int
    to_id:         int
    relation_type: str
    strength:      int  = 3
    note:          str  = ""


class KnowledgeRelationOut(KnowledgeRelationIn):
    id:         int
    created_at: datetime
    class Config:
        from_attributes = True


# ── Decision Engine Schemas ────────────────────────────────────────────────────
class DecisionRequestIn(BaseModel):
    question:     str
    context_json: str  = "{}"
    asked_by:     str  = ""
    module:       str  = ""


class DecisionOut(BaseModel):
    id:                int
    question:          str
    recommendation:    str
    arguments_pro:     str
    arguments_contra:  str
    risks_json:        str
    alternatives_json: str
    knowledge_refs:    str
    confidence:        int
    decided_by:        str
    status:            str
    rationale:         str
    created_at:        datetime
    decided_at:        Optional[datetime]
    class Config:
        from_attributes = True


class KnowledgeItemIn(BaseModel):
    category:      str
    key:           str
    title:         str
    body:          str  = ""
    source_module: str  = ""
    source_ref:    str  = ""
    tags_json:     str  = "[]"
    confidence:    int  = 3
    active:        int  = 1


class KnowledgeItemOut(KnowledgeItemIn):
    id:         int
    created_at: datetime
    updated_at: datetime
    class Config:
        from_attributes = True


class StructureSuggestionIn(BaseModel):
    stype:       str
    title:       str
    description: str  = ""
    priority:    int  = 3
    status:      str  = "open"
    proposed_by: str  = "system"
    notes:       str  = ""


class StructureSuggestionOut(StructureSuggestionIn):
    id:          int
    proposed_at: datetime
    decided_at:  Optional[datetime]
    class Config:
        from_attributes = True


class GovernanceRuleIn(BaseModel):
    rule_id:     str
    module:      str  = "global"
    title:       str
    description: str  = ""
    priority:    int  = 3
    active:      int  = 1


class GovernanceRuleOut(GovernanceRuleIn):
    id:         int
    created_at: datetime
    class Config:
        from_attributes = True


# ----------------------------------------------------------------------------
# Lifecycle + Seed
# ----------------------------------------------------------------------------
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def seed_if_empty() -> None:
    """Legt die vier Robots an, falls die DB noch leer ist."""
    with SessionLocal() as db:
        if db.query(Robot).count() == 0:
            defaults = [
                ("AP", "Robot-AP", "Accounts Payable"),
                ("AR", "Robot-AR", "Accounts Receivable"),
                ("GL", "Robot-GL", "General Ledger"),
                ("AA", "Robot-AA", "Asset Accounting"),
            ]
            for rid, name, domain in defaults:
                db.add(Robot(id=rid, name=name, domain=domain))
            db.commit()
            log.info("Datenbank initialisiert mit 4 Robots.")
        # Zinsen-Task sicherstellen
        if not db.query(Task).filter(Task.tcode == "ZINSEN_BAPI").first():
            db.add(Task(robot_id="GL", name="Zinsbuchung IC-Darlehen",
                        tcode="ZINSEN_BAPI", method="BAPI",
                        description="Monatliche IC-Zinsbuchung via BAPI_ACC_DOCUMENT_POST"))
            db.commit()
            log.info("Zinsen-Task angelegt.")
        # ZinsenSchedule sicherstellen (id=1)
        if not db.get(ZinsenSchedule, 1):
            db.add(ZinsenSchedule(id=1, cron="30 7 30 * *", sap_system="SEQ", active=1))
            db.commit()
            log.info("ZinsenSchedule angelegt (Standard: Tag 30, 07:30, SEQ).")
        # ── 3W Governance-Seed: App-Regeln (nur falls noch keine vorhanden) ─────
        if db.query(GovernanceRule).count() == 0:
            _seed_governance_rules(db)
        # ── Rückstellungs-Kontenplan: immer idempotent sicherstellen ────────────
        _ensure_kontenplan(db)
        # ── Import-Vorlagen: Standard-Templates ───────────────────────────────
        if db.query(RueckstellungTemplate).count() == 0:
            _seed_rueck_templates(db)


def _seed_rueck_templates(db: Session) -> None:
    """Standard Import-Vorlagen für Rückstellungsanträge."""
    import json as _json
    templates = [
        ("", "", "Standard-Vorlage",
         "Universelle Vorlage für alle Buchungskreise und Kategorien",
         '[{"label": "Buchungstext", "field": "beschreibung", "required": true, "type": "text"}, {"label": "Betrag (EUR)", "field": "betrag", "required": true, "type": "number"}, {"label": "Kostenstelle", "field": "kostenstelle", "required": false, "type": "text"}, {"label": "Buchungsdatum", "field": "buchungsdatum", "required": true, "type": "date"}, {"label": "Periode (MM/JJJJ)", "field": "periode", "required": false, "type": "text"}, {"label": "Referenz", "field": "referenz", "required": false, "type": "text"}]'),
        ("", "Personalrückstellung", "Personalrückstellung-Vorlage",
         "Vorlage für Personalrückstellungen inkl. Kostenstelle und Mitarbeiteranzahl",
         '[{"label": "Buchungstext", "field": "beschreibung", "required": true, "type": "text"}, {"label": "Betrag (EUR)", "field": "betrag", "required": true, "type": "number"}, {"label": "Kostenstelle", "field": "kostenstelle", "required": true, "type": "text"}, {"label": "Periode (MM/JJJJ)", "field": "periode", "required": true, "type": "text"}, {"label": "Buchungsdatum", "field": "buchungsdatum", "required": true, "type": "date"}, {"label": "Mitarbeiteranzahl", "field": "referenz", "required": false, "type": "text"}]'),
    ]
    for bukrs, kat, name, beschr, cols in templates:
        db.add(RueckstellungTemplate(
            bukrs=bukrs, kategorie=kat, name=name,
            beschreibung=beschr, spalten_json=cols, aktiv=1))
    db.commit()
    log.info("Rückstellungs-Import-Vorlagen initialisiert (%d Einträge).", len(templates))


# Kanonische Kontenplan-Einträge — jederzeit idempotent erweiterbar
_KONTENPLAN_DEFAULTS = [
    ("Urlaubsrückstellung",         "6310000", "3750000", "", "SA", 1),
    ("Drohverlustrückstellung",     "6550000", "3750000", "", "SA", 2),
    ("Gewährleistungsrückstellung", "6501000", "3750000", "", "SA", 3),
    ("Steuerrückstellung",          "7400000", "3750000", "", "SA", 4),
    ("Pensionsrückstellung",        "6200000", "3750000", "", "SA", 5),
    ("Aktive Rechnungsabgrenzung",  "1500000", "6000000", "", "AB", 6),
    ("Passive Rechnungsabgrenzung", "6000000", "3900000", "", "AB", 7),
    ("Sonstige Rückstellung",       "6990000", "3750000", "", "SA", 8),
    ("Personalrückstellung",        "6320000", "3750000", "", "SA", 9),
]

def _seed_kontenplan(db: Session) -> None:
    """Erst-Befüllung (leere Tabelle)."""
    _ensure_kontenplan(db)

def _ensure_kontenplan(db: Session) -> None:
    """Idempotent: fügt fehlende Standard-Einträge hinzu, ohne bestehende zu überschreiben."""
    added = 0
    for kat, soll, haben, kst, blart, sort in _KONTENPLAN_DEFAULTS:
        exists = db.query(RueckstellungKontenplan).filter(
            RueckstellungKontenplan.kategorie == kat).first()
        if not exists:
            db.add(RueckstellungKontenplan(
                kategorie=kat, konto_soll=soll, konto_haben=haben,
                kostenstelle=kst, blart=blart, aktiv=1, sort_order=sort))
            added += 1
    if added:
        db.commit()
        log.info("Kontenplan: %d neue Einträge ergänzt.", added)


def _seed_governance_rules(db: Session) -> None:
    """Legt die App-Regeln als Governance-Einträge an (einmalig beim ersten Start)."""
    rules = [
        # Global
        ("G-1",  "global",   5, "Passwörter niemals speichern",
         "Passwörter werden NICHT in der DB gespeichert. Eingabe nur zur Laufzeit / sessionStorage."),
        ("G-2",  "global",   5, "Kein .env-Commit",
         ".env-Dateien dürfen niemals ins Git-Repository committed werden."),
        ("G-5",  "zinsen",   5, "SAP-System aus session.sap_system (Priorität)",
         "SAP_CONN.system wird IMMER aus session.sap_system gesetzt, BEVOR session.ashost geprüft wird."),
        ("G-6",  "global",   4, "SEP/SEQ-Systemdefinition",
         "SEP = 172.28.189.8 (Produktion), SEQ = 172.28.189.11 (Qualitätssystem)."),
        ("G-7",  "global",   5, "RFC-Username immer aus session.sap_username",
         "RFC-Username kommt ausschließlich aus session.sap_username (HUEMMKMA). Nie aus manuellem Input."),
        ("G-9",  "global",   3, "Globaler Benutzer/SAP-System-Header",
         "Alle Cockpit-Seiten zeigen angemeldeten Benutzer + SAP-System (SEP/SEQ) via AUTH.renderHeader()."),
        # Zinsen
        ("Z-5",  "zinsen",   4, "Monatlicher Zinsbetrag: Tageszins × Monatstage",
         "Zinsen = Darlehensbetrag × Zinssatz / 360 × Tage. Z-5-R: Restschuld nach Rückzahlungen."),
        ("Z-5-R","zinsen",   4, "Restschuld-Zinsen bei Tilgungen",
         "Z-5 verwendet getLoanRestschuld(): Darlehensbetrag abzgl. aller Rückzahlungen bis Periodenende."),
        ("Z-SAP","zinsen",   5, "Buchungsnummer: 10-stellige BELNR aus OBJ_KEY links",
         "SAP-Belegnummer (BELNR) = linke 10 Zeichen von OBJ_KEY. Niemals aus BAPI_RETURN.MESSAGE."),
        # Payroll
        ("P-8",  "payroll",  3, "Einmal pro Monat – Duplikat-Sperre",
         "Pro Buchungskreis und Monat darf nur eine Lohnbuchhaltungsdatei importiert werden."),
        ("P-9",  "payroll",  3, "Booking Doc Nr: 18-stelliges OBJ_KEY-Format",
         "Format: BELNR(10) + BUKRS(4) + GJAHR(4) = 18 Stellen. Migration für 10-stellige Altdaten."),
        ("P-10", "payroll",  3, "Steuercode nur für RESY-Zeilen",
         "SAP-Vorsteuerkennzeichen (z.B. P1) wird nur bei RESY-Zeilen in payroll_import_lines gesetzt."),
        # Rechnung
        ("R-9",  "rechnung", 3, "1 Rechnung pro Tochter/Monat",
         "Pro Kunden-Nr und Periode darf nur eine Rechnung erstellt werden (Duplikat-Sperre)."),
    ]
    for rule_id, module, priority, title, description in rules:
        db.add(GovernanceRule(
            rule_id=rule_id, module=module, priority=priority,
            title=title, description=description, active=1,
        ))
    db.commit()
    log.info("Governance-Regeln angelegt (%d App-Regeln).", len(rules))


def _migrate_db() -> None:
    """Ergaenzt fehlende Spalten in bestehenden Tabellen (SQLite ALTER TABLE)."""
    import sqlite3
    db_path = DATABASE_URL.replace("sqlite:///", "")
    try:
        con = sqlite3.connect(db_path)
        cur = con.cursor()

        # zinsen_run_once: bapi_jobs, result (neu ab v2)
        cur.execute("PRAGMA table_info(zinsen_run_once)")
        existing = {row[1] for row in cur.fetchall()}
        for col, typedef in [
            ("bapi_jobs", "TEXT NOT NULL DEFAULT '[]'"),
            ("result",    "TEXT NOT NULL DEFAULT ''"),
        ]:
            if col not in existing:
                cur.execute(f"ALTER TABLE zinsen_run_once ADD COLUMN {col} {typedef}")
                log.info("Migration: Spalte '%s' zu zinsen_run_once hinzugefuegt.", col)

        # payroll_imports: created_at, sap_system (neu ab v3)
        cur.execute("PRAGMA table_info(payroll_imports)")
        pi_cols = {row[1] for row in cur.fetchall()}
        for col, typedef in [
            ("created_at",     "DATETIME"),
            ("sap_system",     "VARCHAR(8) NOT NULL DEFAULT 'SEQ'"),
            ("booking_doc_nr", "VARCHAR(32) NOT NULL DEFAULT ''"),
            ("storno_doc_nr",  "VARCHAR(32) NOT NULL DEFAULT ''"),
        ]:
            if col not in pi_cols:
                cur.execute(f"ALTER TABLE payroll_imports ADD COLUMN {col} {typedef}")
                log.info("Migration: Spalte '%s' zu payroll_imports hinzugefuegt.", col)

        # payroll_import_lines: assignment, item_text (neu ab v3)
        cur.execute("PRAGMA table_info(payroll_import_lines)")
        pl_cols = {row[1] for row in cur.fetchall()}
        for col, typedef in [
            ("assignment", "VARCHAR(64) NOT NULL DEFAULT ''"),
            ("item_text",  "VARCHAR(255) NOT NULL DEFAULT ''"),
            ("costcenter", "VARCHAR(32) NOT NULL DEFAULT ''"),
            ("tax_code",   "VARCHAR(8) NOT NULL DEFAULT ''"),   # P-10: SAP-Vorsteuerkennzeichen
        ]:
            if col not in pl_cols:
                cur.execute(f"ALTER TABLE payroll_import_lines ADD COLUMN {col} {typedef}")
                log.info("Migration: Spalte '%s' zu payroll_import_lines hinzugefuegt.", col)

        # App-Regel P-9: booking_doc_nr auf 18-stelliges OBJ_KEY-Format hochrechnen
        # Format: BELNR(10) + BUKRS(4) + GJAHR(4) = 18 Stellen
        # Bestehende Eintraege mit 10-stelliger Nummer werden automatisch migriert.
        cur.execute("""
            SELECT id, booking_doc_nr, comp_code, run_date
            FROM payroll_imports
            WHERE status = 'booked'
              AND length(booking_doc_nr) = 10
              AND booking_doc_nr GLOB '[0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9][0-9]'
        """)
        for row in cur.fetchall():
            imp_id, belnr, bukrs, run_date = row
            # Geschaeftsjahr aus run_date (Format: DD.MM.YYYY oder YYYY-MM-DD)
            gjahr = ''
            if run_date and len(run_date) >= 4:
                import re as _re
                m = _re.search(r'(\d{4})$', run_date) or _re.search(r'^(\d{4})', run_date)
                if m:
                    gjahr = m.group(1)
            if gjahr and bukrs and len(bukrs) <= 4:
                obj_key = belnr + bukrs.ljust(4)[:4] + gjahr
                if len(obj_key) == 18:
                    cur.execute("UPDATE payroll_imports SET booking_doc_nr = ? WHERE id = ?",
                                (obj_key, imp_id))
                    log.info("Migration P-9: Import #%s booking_doc_nr %s → %s", imp_id, belnr, obj_key)

        # ── 10W-Erweiterung: three_w_answers (W4–W10 + wann_geplant + wo + womit) ──
        cur.execute("PRAGMA table_info(three_w_answers)")
        tw_cols = {row[1] for row in cur.fetchall()}
        for col, typedef in [
            ("wie",          "TEXT NOT NULL DEFAULT ''"),
            ("wann_geplant", "DATETIME"),
            ("wo",           "TEXT NOT NULL DEFAULT ''"),
            ("womit",        "TEXT NOT NULL DEFAULT ''"),
            ("risiken",      "TEXT NOT NULL DEFAULT ''"),
            ("alternativen", "TEXT NOT NULL DEFAULT ''"),
            ("erfahrung",    "TEXT NOT NULL DEFAULT ''"),
            ("ergebnis",     "TEXT NOT NULL DEFAULT ''"),
        ]:
            if col not in tw_cols:
                cur.execute(f"ALTER TABLE three_w_answers ADD COLUMN {col} {typedef}")
                log.info("Migration 10W: Spalte '%s' zu three_w_answers hinzugefuegt.", col)

        # ── mgmt_invoice_runs: transaction → sap_transaction (Task #151) ──
        cur.execute("PRAGMA table_info(mgmt_invoice_runs)")
        mir_cols = {row[1] for row in cur.fetchall()}
        if "transaction" in mir_cols and "sap_transaction" not in mir_cols:
            # Umbenennung via temporäre Tabelle (SQLite unterstützt kein RENAME COLUMN < 3.25)
            cur.execute("ALTER TABLE mgmt_invoice_runs ADD COLUMN sap_transaction TEXT NOT NULL DEFAULT ''")
            cur.execute("UPDATE mgmt_invoice_runs SET sap_transaction = \"transaction\"")
            log.info("Migration: mgmt_invoice_runs.transaction → sap_transaction kopiert.")
        elif "sap_transaction" not in mir_cols and "transaction" not in mir_cols:
            pass  # Tabelle noch nicht vorhanden – create_all legt sie korrekt an
        for col, typedef in [
            ("sap_system",     "TEXT NOT NULL DEFAULT 'SEQ'"),
            ("vendor",         "TEXT NOT NULL DEFAULT 'L372961'"),
            ("tax_code",       "TEXT NOT NULL DEFAULT ''"),
            ("business_place", "TEXT NOT NULL DEFAULT ''"),
        ]:
            if col not in mir_cols:
                cur.execute(f"ALTER TABLE mgmt_invoice_runs ADD COLUMN {col} {typedef}")
                log.info("Migration: mgmt_invoice_runs Spalte '%s' hinzugefuegt.", col)

        # ── Knowledge Universe + Decision Engine: per create_all (neue Tabellen) ──
        # Neue Tabellen werden automatisch via Base.metadata.create_all angelegt.
        # Hier kein ALTER TABLE noetig — reine Neuerstellung.
        con.commit()
        con.close()
        log.info("Migration abgeschlossen (10W + Knowledge Universe + Decision Engine).")

    except Exception as exc:
        log.warning("DB-Migration fehlgeschlagen (ignoriert): %s", exc)

    # ── Rückstellungs-Tabellen: explizit anlegen falls fehlend (Task #121) ──
    # create_all() legt neue Tabellen nur beim ersten Start an. Falls die DB
    # vor Task #121 erzeugt wurde, fehlen diese Tabellen — daher hier nachholen.
    try:
        import sqlite3 as _sq2
        _db2 = DATABASE_URL.replace("sqlite:///", "")
        _con2 = _sq2.connect(_db2, timeout=5)
        _cur2 = _con2.cursor()
        _cur2.executescript("""
            CREATE TABLE IF NOT EXISTS rueckstellung_requests (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                workflow         VARCHAR(16)  NOT NULL DEFAULT 'direkt',
                kategorie        VARCHAR(64)  NOT NULL DEFAULT '',
                beschreibung     TEXT         NOT NULL DEFAULT '',
                betrag           REAL         NOT NULL DEFAULT 0.0,
                waehrung         VARCHAR(4)   NOT NULL DEFAULT 'EUR',
                konto_soll       VARCHAR(16)  NOT NULL DEFAULT '',
                konto_haben      VARCHAR(16)  NOT NULL DEFAULT '',
                kostenstelle     VARCHAR(16)  NOT NULL DEFAULT '',
                bukrs_json       TEXT         NOT NULL DEFAULT '[]',
                periode          VARCHAR(16)  NOT NULL DEFAULT '',
                buchungsdatum    VARCHAR(10)  NOT NULL DEFAULT '',
                referenz         VARCHAR(128) NOT NULL DEFAULT '',
                sap_system       VARCHAR(16)  NOT NULL DEFAULT 'SEQ',
                blart            VARCHAR(4)   NOT NULL DEFAULT 'SA',
                status           VARCHAR(32)  NOT NULL DEFAULT 'offen',
                requires_approval INTEGER     NOT NULL DEFAULT 0,
                submitted_by     VARCHAR(128) NOT NULL DEFAULT '',
                approved_by      VARCHAR(128) NOT NULL DEFAULT '',
                approval_comment TEXT         NOT NULL DEFAULT '',
                sap_doc_nrs      TEXT         NOT NULL DEFAULT '',
                created_at       DATETIME,
                updated_at       DATETIME
            );
            CREATE TABLE IF NOT EXISTS rueckstellung_approvals (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                request_id INTEGER NOT NULL REFERENCES rueckstellung_requests(id),
                action     VARCHAR(32) NOT NULL,
                actor      VARCHAR(128) NOT NULL DEFAULT '',
                comment    TEXT         NOT NULL DEFAULT '',
                ts         DATETIME
            );
            CREATE TABLE IF NOT EXISTS rueckstellung_kontenplan (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                kategorie    VARCHAR(64)  NOT NULL DEFAULT '',
                konto_soll   VARCHAR(16)  NOT NULL DEFAULT '',
                konto_haben  VARCHAR(16)  NOT NULL DEFAULT '',
                kostenstelle VARCHAR(16)  NOT NULL DEFAULT '',
                blart        VARCHAR(4)   NOT NULL DEFAULT 'SA',
                aktiv        INTEGER      NOT NULL DEFAULT 1,
                sort_order   INTEGER      NOT NULL DEFAULT 0
            );
            CREATE TABLE IF NOT EXISTS rueckstellung_templates (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                bukrs        VARCHAR(8)   NOT NULL DEFAULT '',
                kategorie    VARCHAR(64)  NOT NULL DEFAULT '',
                name         VARCHAR(128) NOT NULL DEFAULT '',
                beschreibung TEXT         NOT NULL DEFAULT '',
                spalten_json TEXT         NOT NULL DEFAULT '[]',
                aktiv        INTEGER      NOT NULL DEFAULT 1
            );
            CREATE TABLE IF NOT EXISTS afa_runs (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                fire_at      DATETIME     NOT NULL,
                bukrs        VARCHAR(8)   NOT NULL DEFAULT '',
                fiscal_year  VARCHAR(4)   NOT NULL DEFAULT '',
                period       VARCHAR(2)   NOT NULL DEFAULT '01',
                reason       VARCHAR(1)   NOT NULL DEFAULT '1',
                testrun      INTEGER      NOT NULL DEFAULT 1,
                stop_on_warn INTEGER      NOT NULL DEFAULT 0,
                posting_date VARCHAR(8)   NOT NULL DEFAULT '',
                sap_system   VARCHAR(8)   NOT NULL DEFAULT 'SEQ',
                status       VARCHAR(16)  NOT NULL DEFAULT 'pending',
                result       TEXT         NOT NULL DEFAULT '',
                created_by   VARCHAR(64)  NOT NULL DEFAULT '',
                created_at   DATETIME     NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
        """)
        _con2.commit()
        _con2.close()
        log.info("Migration Rueckstellung: Tabellen sichergestellt (CREATE IF NOT EXISTS).")
    except Exception as _exc2:
        log.warning("Migration Rueckstellung fehlgeschlagen: %s", _exc2)


scheduler = BackgroundScheduler(
    job_defaults={"misfire_grace_time": 3600, "coalesce": True}
)


def trigger_run(task_id: int) -> None:
    """Wird vom Scheduler aufgerufen: legt einen neuen Run in den Status 'queued'."""
    with SessionLocal() as db:
        run = Run(task_id=task_id, status="queued")
        db.add(run)
        db.commit()
        log.info("Run %d fuer Task %d eingereiht (scheduler).", run.id, task_id)


def trigger_zinsen_run(sap_system: str, periode: str) -> None:
    """Zinsen-Buchungslauf: legt Run mit Payload (sap_system, periode) an."""
    import json
    with SessionLocal() as db:
        zs = db.get(ZinsenSchedule, 1)
        # Zinsen-Task per Name suchen
        task = db.query(Task).filter(Task.tcode == "ZINSEN_BAPI").first()
        if not task:
            log.error("Zinsen-Task nicht gefunden – Buchung abgebrochen.")
            return
        payload = json.dumps({"sap_system": sap_system, "periode": periode})
        run = Run(task_id=task.id, status="queued", payload=payload)
        db.add(run)
        db.commit()
        log.info("Zinsen-Run %d eingereiht (System=%s, Periode=%s).", run.id, sap_system, periode)


def _bridge_post(path: str, body: dict, timeout: int = 90) -> dict:
    """HTTP-POST an die Bridge via urllib (stdlib – kein requests noetig)."""
    import json as _json
    import urllib.request as _ur
    import urllib.error as _ue
    url  = f"http://localhost:8765{path}"
    data = _json.dumps(body).encode("utf-8")
    req  = _ur.Request(url, data=data,
                       headers={"Content-Type": "application/json"},
                       method="POST")
    try:
        with _ur.urlopen(req, timeout=timeout) as resp:
            return _json.loads(resp.read().decode("utf-8"))
    except _ue.HTTPError as e:
        body_txt = e.read().decode("utf-8", errors="replace")
        return {"status": "error", "message": f"HTTP {e.code}: {body_txt}"}
    except Exception as exc:
        return {"status": "error", "message": f"Bridge nicht erreichbar: {exc}"}


def trigger_zinsen_run_once(run_once_id: int) -> None:
    """Einmalige Zinsbuchung: ruft Bridge direkt fuer jeden BAPI-Job auf."""
    import json as _json

    # Status auf "running" setzen
    try:
        with SessionLocal() as db:
            ro = db.get(ZinsenRunOnce, run_once_id)
            if not ro or ro.status not in ("pending", "error"):
                log.warning("RunOnce %d: unerwarteter Status '%s', Abbruch.",
                            run_once_id, ro.status if ro else "not found")
                return
            ro.status = "running"
            db.commit()
            jobs_raw = ro.bapi_jobs or "[]"
            periode  = ro.periode
            sap_sys  = ro.sap_system
    except Exception as exc:
        log.error("RunOnce %d: DB-Fehler beim Statuswechsel: %s", run_once_id, exc)
        return

    # Jobs deserialisieren
    try:
        jobs = _json.loads(jobs_raw)
    except Exception as exc:
        jobs = []
        log.error("RunOnce %d: bapi_jobs nicht parsebar: %s", run_once_id, exc)

    if not jobs:
        with SessionLocal() as db:
            ro = db.get(ZinsenRunOnce, run_once_id)
            if ro:
                ro.status = "error"
                ro.result = _json.dumps({"errors": [
                    {"loan": "–", "msg": "Keine BAPI-Jobs gespeichert. Eintrag neu anlegen."}
                ]})
                db.commit()
        log.error("RunOnce %d: bapi_jobs leer – keine Buchung.", run_once_id)
        return

    log.info("RunOnce %d: starte %d BAPI-Jobs (Periode %s, System %s).",
             run_once_id, len(jobs), periode, sap_sys)

    ok_list, err_list = [], []
    for i, job in enumerate(jobs):
        loan_ref = job.get("_loan_ref", f"Job-{i}")
        log.info("RunOnce %d – Job %d/%d: %s", run_once_id, i+1, len(jobs), loan_ref)
        data = _bridge_post("/run", job)
        if data.get("status") == "ok":
            doc = data.get("doc_number") or data.get("message") or "?"
            ok_list.append({"loan": loan_ref, "doc": doc})
            log.info("RunOnce %d – OK: %s Belegnr. %s", run_once_id, loan_ref, doc)
        else:
            msg = data.get("message") or data.get("error") or str(data)
            err_list.append({"loan": loan_ref, "msg": msg})
            log.error("RunOnce %d – FEHLER: %s → %s", run_once_id, loan_ref, msg)

    final_status = "error" if err_list else "ok"
    result_json  = _json.dumps({"ok": ok_list, "errors": err_list})
    try:
        with SessionLocal() as db:
            ro = db.get(ZinsenRunOnce, run_once_id)
            if ro:
                ro.status = final_status
                ro.result = result_json
                db.commit()
    except Exception as exc:
        log.error("RunOnce %d: DB-Fehler beim Ergebnis-Speichern: %s", run_once_id, exc)

    log.info("RunOnce %d abgeschlossen: Status=%s, %d OK, %d Fehler.",
             run_once_id, final_status, len(ok_list), len(err_list))


def _auto_periode() -> str:
    """Gibt die aktuelle Buchungsperiode als MM/YYYY zurück."""
    now = datetime.utcnow()
    return f"{now.month:02d}/{now.year}"


def trigger_afa_run(afa_run_id: int) -> None:
    """Wird vom APScheduler aufgerufen – führt den AfA-Lauf aus."""
    log.info("AfA-Lauf %d gestartet.", afa_run_id)
    with SessionLocal() as db:
        try:
            ar = db.get(AfaRun, afa_run_id)
            if not ar or ar.status != "pending":
                log.warning("AfA-Lauf %d nicht mehr pending (%s) – übersprungen.",
                            afa_run_id, ar.status if ar else "not found")
                return
            ar.status = "running"
            db.commit()
        except Exception as exc:
            log.error("AfA-Lauf %d: DB-Fehler beim Start: %s", afa_run_id, exc)
            return

    import json as _json
    sap_auth = None   # Fallback auf .env Service-Account
    try:
        from worker.handlers import batch_afab_depreciation  # type: ignore
        payload = {
            "bukrs":        ar.bukrs,
            "fiscal_year":  ar.fiscal_year,
            "period":       ar.period,
            "reason":       ar.reason,
            "testrun":      bool(ar.testrun),
            "stop_on_warn": bool(ar.stop_on_warn),
            "posting_date": ar.posting_date,
            "_sap_auth":    sap_auth,
        }
        result = batch_afab_depreciation({}, payload)
        final_status = "ok"
        result_json = _json.dumps(result, ensure_ascii=False, default=str)
    except Exception as exc:
        log.error("AfA-Lauf %d: Fehler: %s", afa_run_id, exc)
        final_status = "error"
        result_json = _json.dumps({"error": str(exc)}, ensure_ascii=False)

    with SessionLocal() as db:
        try:
            ar = db.get(AfaRun, afa_run_id)
            if ar:
                ar.status = final_status
                ar.result = result_json
                db.commit()
        except Exception as exc:
            log.error("AfA-Lauf %d: DB-Fehler beim Speichern: %s", afa_run_id, exc)
    log.info("AfA-Lauf %d abgeschlossen: %s", afa_run_id, final_status)


def reload_schedules() -> None:
    """Liest alle aktiven Schedules aus der DB und plant sie im APScheduler ein."""
    scheduler.remove_all_jobs()
    with SessionLocal() as db:
        # Standard-Schedules (Robots/Tasks)
        for s in db.query(Schedule).filter(Schedule.active == 1).all():
            try:
                trigger = CronTrigger.from_crontab(s.cron)
                scheduler.add_job(trigger_run, trigger=trigger, args=[s.task_id], id=f"sched-{s.id}")
            except Exception as exc:    # noqa: BLE001
                log.warning("Schedule %s ungueltig (%s): %s", s.id, s.cron, exc)
        # Zinsen Monatsplan
        zs = db.get(ZinsenSchedule, 1)
        if zs and zs.active:
            try:
                trigger = CronTrigger.from_crontab(zs.cron)
                scheduler.add_job(trigger_zinsen_run, trigger=trigger,
                                  args=[zs.sap_system, "auto"],
                                  id="zinsen-monthly")
                log.info("Zinsen-Monatsplan registriert: %s (%s)", zs.cron, zs.sap_system)
            except Exception as exc:    # noqa: BLE001
                log.warning("Zinsen-Monatsplan ungueltig: %s", exc)
        # Einmalige Zinsbuchungen (pending, noch in der Zukunft)
        for ro in db.query(ZinsenRunOnce).filter(ZinsenRunOnce.status == "pending").all():
            if ro.fire_at > datetime.utcnow():
                scheduler.add_job(trigger_zinsen_run_once,
                                  trigger=DateTrigger(run_date=ro.fire_at),
                                  args=[ro.id], id=f"zinsen-once-{ro.id}")
                log.info("Einmalige Zinsbuchung %d eingeplant fuer %s.", ro.id, ro.fire_at)
        # AfA-Laeufe (pending, noch in der Zukunft)
        for ar in db.query(AfaRun).filter(AfaRun.status == "pending").all():
            if ar.fire_at > datetime.utcnow():
                scheduler.add_job(trigger_afa_run,
                                  trigger=DateTrigger(run_date=ar.fire_at),
                                  args=[ar.id], id=f"afa-run-{ar.id}",
                                  replace_existing=True)
                log.info("AfA-Lauf %d eingeplant fuer %s.", ar.id, ar.fire_at)


@asynccontextmanager
async def lifespan(_: FastAPI):
    Base.metadata.create_all(engine)   # neue Tabellen anlegen
    _migrate_db()                       # fehlende Spalten in bestehenden Tabellen ergaenzen
    seed_if_empty()
    scheduler.start()
    reload_schedules()
    log.info("Orchestrator gestartet.")
    yield
    scheduler.shutdown()


app = FastAPI(title="SAP Robot Orchestrator", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],            # in Produktion einschraenken
    allow_methods=["*"],
    allow_headers=["*"],
)

# Cockpit-Seiten über http://localhost:8000/cockpit/ erreichbar machen
# (nötig für Mikrofon-Zugriff in Browsern – file:// wird teilweise blockiert)
# HTML-Dateien werden mit no-cache ausgeliefert, damit Änderungen sofort sichtbar sind.
_COCKPIT_DIR = pathlib.Path(__file__).parent.parent / "cockpit"
if _COCKPIT_DIR.exists():
    from starlette.staticfiles import StaticFiles as _SF
    from starlette.responses import FileResponse as _FR

    class NoCacheStaticFiles(_SF):
        """StaticFiles mit no-cache für .html, damit Änderungen sofort greifen."""
        async def get_response(self, path: str, scope):
            resp = await super().get_response(path, scope)
            if isinstance(resp, _FR) and path.endswith(".html"):
                resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
                resp.headers["Pragma"] = "no-cache"
                resp.headers["Expires"] = "0"
            return resp

    app.mount("/cockpit", NoCacheStaticFiles(directory=str(_COCKPIT_DIR), html=True), name="cockpit")


# ----------------------------------------------------------------------------
# Endpoints: Robots & Tasks (Cockpit)
# ----------------------------------------------------------------------------
@app.get("/robots")
def list_robots(db: Session = Depends(get_db)):
    return [
        {
            "id": r.id, "name": r.name, "domain": r.domain,
            "task_count": len(r.tasks),
        }
        for r in db.query(Robot).all()
    ]


@app.get("/tasks", response_model=List[TaskOut])
def list_tasks(robot_id: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(Task)
    if robot_id:
        q = q.filter(Task.robot_id == robot_id)
    return q.order_by(Task.id).all()


@app.post("/tasks", response_model=TaskOut, status_code=201)
def create_task(payload: TaskIn, db: Session = Depends(get_db)):
    if not db.get(Robot, payload.robot_id):
        raise HTTPException(404, f"Robot {payload.robot_id} unbekannt.")
    t = Task(**payload.model_dump())
    db.add(t); db.commit(); db.refresh(t)
    return t


@app.put("/tasks/{task_id}", response_model=TaskOut)
def update_task(task_id: int, payload: TaskIn, db: Session = Depends(get_db)):
    t = db.get(Task, task_id)
    if not t:
        raise HTTPException(404)
    for k, v in payload.model_dump().items():
        setattr(t, k, v)
    db.commit(); db.refresh(t)
    return t


@app.delete("/tasks/{task_id}", status_code=204)
def delete_task(task_id: int, db: Session = Depends(get_db)):
    t = db.get(Task, task_id)
    if not t:
        raise HTTPException(404)
    db.delete(t); db.commit()


# ----------------------------------------------------------------------------
# Endpoints: Runs (Cockpit -> startet; Worker -> aktualisiert)
# ----------------------------------------------------------------------------
@app.post("/tasks/{task_id}/run", response_model=RunOut, status_code=201)
def enqueue_run(task_id: int, db: Session = Depends(get_db)):
    """Cockpit-Button 'Jetzt ausfuehren' -> legt Run in den Status 'queued'."""
    if not db.get(Task, task_id):
        raise HTTPException(404, "Task unbekannt.")
    run = Run(task_id=task_id, status="queued")
    db.add(run); db.commit(); db.refresh(run)
    log.info("Run %d fuer Task %d manuell eingereiht.", run.id, task_id)
    return run


@app.get("/runs", response_model=List[RunOut])
def list_runs(task_id: Optional[int] = None, status_filter: Optional[str] = None,
              limit: int = 50, db: Session = Depends(get_db)):
    q = db.query(Run)
    if task_id:
        q = q.filter(Run.task_id == task_id)
    if status_filter:
        q = q.filter(Run.status == status_filter)
    return q.order_by(Run.id.desc()).limit(limit).all()


# --- Worker-Endpoints ---
@app.post("/worker/claim", response_model=Optional[RunOut])
def claim_next_run(worker_id: str, robot_id: Optional[str] = None,
                   db: Session = Depends(get_db)):
    """Worker fragt nach naechstem offenen Job. Optional gefiltert auf seinen Robot."""
    q = db.query(Run).filter(Run.status == "queued")
    if robot_id:
        q = q.join(Task).filter(Task.robot_id == robot_id)
    run = q.order_by(Run.id.asc()).first()
    if not run:
        return None
    run.status = "running"
    run.started_at = datetime.utcnow()
    run.worker_id = worker_id
    db.commit(); db.refresh(run)
    log.info("Run %d an Worker %s vergeben.", run.id, worker_id)
    return run


@app.post("/worker/runs/{run_id}", response_model=RunOut)
def update_run(run_id: int, update: RunUpdate, db: Session = Depends(get_db)):
    """Worker meldet Fortschritt/Ergebnis zurueck."""
    run = db.get(Run, run_id)
    if not run:
        raise HTTPException(404)
    if update.status not in ("running", "ok", "error"):
        raise HTTPException(400, "Ungueltiger Status.")
    run.status = update.status
    if update.log:
        run.log = (run.log or "") + update.log + "\n"
    if update.worker_id:
        run.worker_id = update.worker_id
    if update.status in ("ok", "error"):
        run.finished_at = datetime.utcnow()
    db.commit(); db.refresh(run)
    return run


# ----------------------------------------------------------------------------
# Endpoints: Schedules
# ----------------------------------------------------------------------------
class ScheduleIn(BaseModel):
    task_id: int
    cron: str        # Standard-Crontab, z. B. "0 6 * * *"
    active: int = 1


@app.get("/schedules")
def list_schedules(db: Session = Depends(get_db)):
    return [
        {"id": s.id, "task_id": s.task_id, "cron": s.cron, "active": s.active}
        for s in db.query(Schedule).all()
    ]


@app.post("/schedules", status_code=201)
def create_schedule(payload: ScheduleIn, db: Session = Depends(get_db)):
    if not db.get(Task, payload.task_id):
        raise HTTPException(404, "Task unbekannt.")
    s = Schedule(**payload.model_dump())
    db.add(s); db.commit(); db.refresh(s)
    reload_schedules()
    return {"id": s.id}


@app.delete("/schedules/{schedule_id}", status_code=204)
def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    s = db.get(Schedule, schedule_id)
    if not s:
        raise HTTPException(404)
    db.delete(s); db.commit()
    reload_schedules()


# ----------------------------------------------------------------------------
# Endpoints: SAP-Benutzerverwaltung
# ----------------------------------------------------------------------------
def _sap_user_decrypt(u: "SapUser") -> "SapUser":
    """Entschlüsselt sensitive Felder eines SapUser-ORM-Objekts in-place."""
    u.sap_username = _dec(u.sap_username)
    u.ashost       = _dec(u.ashost)
    return u


@app.get("/sap-users", response_model=List[SapUserOut])
def list_sap_users(active_only: bool = False, db: Session = Depends(get_db)):
    q = db.query(SapUser)
    if active_only:
        q = q.filter(SapUser.active == 1)
    users = q.order_by(SapUser.display_name).all()
    return [_sap_user_decrypt(u) for u in users]


@app.get("/sap-users/{user_id}", response_model=SapUserOut)
def get_sap_user(user_id: int, db: Session = Depends(get_db)):
    u = db.get(SapUser, user_id)
    if not u:
        raise HTTPException(404, "Benutzer nicht gefunden.")
    return _sap_user_decrypt(u)


@app.post("/sap-users", response_model=SapUserOut, status_code=201)
def create_sap_user(payload: SapUserIn, db: Session = Depends(get_db)):
    data = payload.model_dump()
    data["sap_username"] = _enc(data.get("sap_username", ""))
    data["ashost"]       = _enc(data.get("ashost", ""))
    u = SapUser(**data)
    db.add(u); db.commit(); db.refresh(u)
    log.info("SAP-Benutzer angelegt: %s", u.display_name)
    return _sap_user_decrypt(u)


@app.put("/sap-users/{user_id}", response_model=SapUserOut)
def update_sap_user(user_id: int, payload: SapUserIn, db: Session = Depends(get_db)):
    u = db.get(SapUser, user_id)
    if not u:
        raise HTTPException(404, "Benutzer nicht gefunden.")
    data = payload.model_dump()
    data["sap_username"] = _enc(data.get("sap_username", ""))
    data["ashost"]       = _enc(data.get("ashost", ""))
    for k, v in data.items():
        setattr(u, k, v)
    db.commit(); db.refresh(u)
    log.info("SAP-Benutzer aktualisiert: %s", u.display_name)
    return _sap_user_decrypt(u)


@app.delete("/sap-users/{user_id}", status_code=204)
def delete_sap_user(user_id: int, db: Session = Depends(get_db)):
    u = db.get(SapUser, user_id)
    if not u:
        raise HTTPException(404, "Benutzer nicht gefunden.")
    db.delete(u); db.commit()
    log.info("SAP-Benutzer geloescht: ID %d", user_id)


@app.get("/health")
def health():
    return {"status": "ok", "time": datetime.utcnow().isoformat(), "env": _APP_ENV}

@app.get("/env")
def get_env():
    labels = {"dev": "🟡 ENTWICKLUNG", "test": "🔵 TEST", "prod": "🔴 PRODUKTION"}
    return {
        "env":   _APP_ENV,
        "label": labels.get(_APP_ENV, _APP_ENV),
        "db":    os.getenv("ORCHESTRATOR_DB", "orchestrator.db"),
    }


# ----------------------------------------------------------------------------
# Zahlungsbedingungen-Protokoll
# ----------------------------------------------------------------------------
class RueckstellungIn(BaseModel):
    workflow:      str = "direkt"
    kategorie:     str = ""
    beschreibung:  str = ""
    betrag:        float = 0.0
    waehrung:      str = "EUR"
    konto_soll:    str = ""
    konto_haben:   str = ""
    kostenstelle:  str = ""
    bukrs_json:    str = "[]"
    periode:       str = ""
    buchungsdatum: str = ""
    referenz:      str = ""
    sap_system:    str = "SEQ"
    blart:         str = "SA"
    submitted_by:  str = ""

class RueckstellungApprovalLog(BaseModel):
    id:      int
    action:  str
    actor:   str = ""
    comment: str = ""
    ts:      datetime
    model_config = {"from_attributes": True}

class RueckstellungOut(RueckstellungIn):
    id:               int
    status:           str
    requires_approval: int
    approved_by:      str
    approval_comment: str
    sap_doc_nrs:      str
    created_at:       datetime
    updated_at:       datetime
    approvals:        List[RueckstellungApprovalLog] = []
    model_config = {"from_attributes": True}


class KontenplanIn(BaseModel):
    kategorie:    str
    konto_soll:   str = ""
    konto_haben:  str = ""
    kostenstelle: str = ""
    blart:        str = "SA"
    aktiv:        int = 1
    sort_order:   int = 0

class KontenplanOut(KontenplanIn):
    id: int
    model_config = {"from_attributes": True}


class RueckstellungTemplateIn(BaseModel):
    bukrs:        str = ""
    kategorie:    str = ""
    name:         str = ""
    beschreibung: str = ""
    spalten_json: str = "[]"
    aktiv:        int = 1

class RueckstellungTemplateOut(RueckstellungTemplateIn):
    id:         int
    created_at: datetime
    updated_at: datetime
    model_config = {"from_attributes": True}


class ZtermLogIn(BaseModel):
    kunnr:      str
    kname:      str = ""
    bukrs:      str = ""
    tcode:      str
    vbeln:      str = ""
    zterm_old:  str = ""
    zterm_new:  str
    sap_system: str = ""
    changed_by: str
    status:     str = "ok"
    note:       str = ""

class ZtermLogOut(ZtermLogIn):
    id:         int
    changed_at: datetime
    class Config:
        from_attributes = True

@app.post("/zterm-log", response_model=ZtermLogOut, status_code=201, tags=["Protokoll"])
def create_zterm_log(payload: ZtermLogIn, db: Session = Depends(get_db)):
    entry = ZtermChangeLog(**payload.model_dump())
    db.add(entry); db.commit(); db.refresh(entry)
    return entry

@app.get("/zterm-log", response_model=List[ZtermLogOut], tags=["Protokoll"])
def get_zterm_log(
    kunnr: Optional[str] = None,
    tcode: Optional[str] = None,
    limit: int = 200,
    db: Session = Depends(get_db)
):
    from sqlalchemy import desc
    q = db.query(ZtermChangeLog)
    if kunnr:
        q = q.filter(ZtermChangeLog.kunnr == kunnr)
    if tcode:
        q = q.filter(ZtermChangeLog.tcode == tcode)
    return q.order_by(desc(ZtermChangeLog.changed_at)).limit(limit).all()


# ----------------------------------------------------------------------------
# Endpoints: PayRoll-Buchungen
# ----------------------------------------------------------------------------
@app.post("/payroll/imports", status_code=201)
def create_payroll_import(payload: PayrollImportIn, db: Session = Depends(get_db)):
    """Speichert importierte FIBU-LOGA-Daten sofort nach dem CSV-Upload."""
    imp = PayrollImport(
        doc_id=payload.doc_id,       run_date=payload.run_date,
        comp_code=payload.comp_code, currency=payload.currency,
        row_count=payload.row_count, file_name=payload.file_name,
        imported_by=payload.imported_by, sap_system=payload.sap_system,
    )
    db.add(imp); db.flush()          # imp.id sofort verfügbar
    for ln in payload.lines:
        db.add(PayrollImportLine(
            import_id=imp.id,  nr=ln.nr,
            comp_code=ln.comp_code, currency=ln.currency,
            pk=ln.pk,          account=ln.account,
            amount=str(ln.amount),
            costcenter=ln.costcenter, assignment=ln.assignment,
            item_text=ln.item_text,
            tax_code=ln.tax_code,
        ))
    db.commit(); db.refresh(imp)
    log.info("PayRoll-Import %d gespeichert: %s (%s, %d Zeilen).",
             imp.id, payload.doc_id, payload.comp_code, payload.row_count)
    return {"id": imp.id, "doc_id": imp.doc_id,
            "row_count": imp.row_count, "status": imp.status}


@app.get("/payroll/imports")
def list_payroll_imports(limit: int = 50, db: Session = Depends(get_db)):
    """Listet alle gespeicherten PayRoll-Importe (neueste zuerst)."""
    rows = (db.query(PayrollImport)
              .order_by(PayrollImport.created_at.desc())
              .limit(limit).all())
    return [
        {"id": r.id, "doc_id": r.doc_id or "", "run_date": r.run_date or "",
         "comp_code": r.comp_code or "", "currency": r.currency or "EUR",
         "row_count": r.row_count or 0, "file_name": r.file_name or "",
         "imported_by": r.imported_by or "", "sap_system": r.sap_system or "",
         "status": r.status or "imported", "booking_doc_nr": r.booking_doc_nr or "",
         "storno_doc_nr": r.storno_doc_nr or "",
         "created_at": r.created_at.isoformat() if r.created_at else ""}
        for r in rows
    ]


@app.get("/payroll/imports/{import_id}")
def get_payroll_import(import_id: int, db: Session = Depends(get_db)):
    """Gibt einen Import inkl. aller Buchungszeilen zurück."""
    imp = db.get(PayrollImport, import_id)
    if not imp:
        raise HTTPException(404, "Import nicht gefunden.")
    lines = [
        {"nr": ln.nr, "comp_code": ln.comp_code, "currency": ln.currency,
         "pk": ln.pk, "account": ln.account, "amount": float(ln.amount),
         "costcenter": ln.costcenter, "assignment": ln.assignment,
         "item_text": ln.item_text, "tax_code": ln.tax_code or ""}
        for ln in sorted(imp.lines, key=lambda x: x.nr)
    ]
    return {"id": imp.id, "doc_id": imp.doc_id or "", "run_date": imp.run_date or "",
            "comp_code": imp.comp_code or "", "currency": imp.currency or "EUR",
            "row_count": imp.row_count or 0, "file_name": imp.file_name or "",
            "imported_by": imp.imported_by or "", "sap_system": imp.sap_system or "",
            "status": imp.status or "imported", "booking_doc_nr": imp.booking_doc_nr or "",
            "storno_doc_nr": imp.storno_doc_nr or "",
            "created_at": imp.created_at.isoformat() if imp.created_at else "",
            "lines": lines}


@app.patch("/payroll/imports/{import_id}/status")
def update_payroll_status(import_id: int, body: PayrollStatusIn,
                          db: Session = Depends(get_db)):
    imp = db.get(PayrollImport, import_id)
    if not imp:
        raise HTTPException(404, "Import nicht gefunden.")
    imp.status = body.status
    if body.booking_doc_nr:
        imp.booking_doc_nr = body.booking_doc_nr
    db.commit()
    return {"id": imp.id, "status": imp.status, "booking_doc_nr": imp.booking_doc_nr}


@app.patch("/payroll/imports/{import_id}/storno")
def update_payroll_storno(import_id: int, body: PayrollStornoIn,
                          db: Session = Depends(get_db)):
    imp = db.get(PayrollImport, import_id)
    if not imp:
        raise HTTPException(404, "Import nicht gefunden.")
    imp.status = "reversed"
    if body.storno_doc_nr:
        imp.storno_doc_nr = body.storno_doc_nr
    db.commit()
    return {"id": imp.id, "status": imp.status, "storno_doc_nr": imp.storno_doc_nr}


@app.delete("/payroll/imports/{import_id}", status_code=204)
def delete_payroll_import(import_id: int, db: Session = Depends(get_db)):
    imp = db.get(PayrollImport, import_id)
    if not imp:
        raise HTTPException(404, "Import nicht gefunden.")
    if imp.status not in ("imported", "error", "validated"):
        raise HTTPException(409, f"Import im Status '{imp.status}' kann nicht geloescht werden.")
    db.delete(imp)
    db.commit()


@app.patch("/payroll/imports/{import_id}/lines/{line_nr}")
def update_payroll_line(import_id: int, line_nr: int,
                        payload: PayrollLineUpdate, db: Session = Depends(get_db)):
    line = (db.query(PayrollImportLine)
              .filter(PayrollImportLine.import_id == import_id,
                      PayrollImportLine.nr == line_nr)
              .first())
    if not line:
        raise HTTPException(404, "Zeile nicht gefunden.")
    if payload.costcenter is not None: line.costcenter = payload.costcenter
    if payload.item_text  is not None: line.item_text  = payload.item_text
    if payload.assignment is not None: line.assignment  = payload.assignment
    db.commit()
    return {"nr": line.nr, "costcenter": line.costcenter,
            "item_text": line.item_text, "assignment": line.assignment}


# ----------------------------------------------------------------------------
# Endpoints: InvoiceRecord (Rechnungserstellung)
# ----------------------------------------------------------------------------
import json as _json


@app.get("/invoice_records")
def list_invoice_records(
    status:  Optional[str] = None,
    periode: Optional[str] = None,
    bukrs:   Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(InvoiceRecord)
    if status:  q = q.filter(InvoiceRecord.status  == status)
    if periode: q = q.filter(InvoiceRecord.periode == periode)
    if bukrs:   q = q.filter(InvoiceRecord.bukrs   == bukrs)
    rows = q.order_by(InvoiceRecord.periode.desc(), InvoiceRecord.kunden_nr).all()
    return [_invoice_to_dict(r) for r in rows]


@app.post("/invoice_records", status_code=201)
def create_invoice_record(rec: InvoiceRecordIn, db: Session = Depends(get_db)):
    """Einzelnen Datensatz manuell anlegen (aus Frontend-Formular)."""
    positions_json = _json.dumps(
        [p.model_dump() for p in rec.positions], ensure_ascii=False
    )
    existing = db.query(InvoiceRecord).filter_by(group_key=rec.group_key).first()
    if existing:
        raise HTTPException(409, f"group_key '{rec.group_key}' existiert bereits")
    obj = InvoiceRecord(
        group_key=rec.group_key, kunden_nr=rec.kunden_nr, name=rec.name,
        periode=rec.periode, sap_date=rec.sap_date, leistungsart=rec.leistungsart,
        mwst_code=rec.mwst_code, waehrung=rec.waehrung, total=rec.total,
        status=rec.status, invoice_nr=rec.invoice_nr, order_nr=rec.order_nr,
        positions_json=positions_json, is_storno=rec.is_storno, bukrs=rec.bukrs,
    )
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return _invoice_to_dict(obj)


@app.post("/invoice_records/bulk", status_code=200)
def upsert_invoice_records(records: List[InvoiceRecordIn], db: Session = Depends(get_db)):
    deduped: dict = {}
    for rec in records:
        deduped[rec.group_key] = rec
    records = list(deduped.values())
    inserted = updated = 0
    for rec in records:
        positions_json = _json.dumps(
            [p.model_dump() for p in rec.positions], ensure_ascii=False
        )
        existing = db.query(InvoiceRecord).filter_by(group_key=rec.group_key).first()
        if existing:
            if existing.status == "offen":
                existing.name         = rec.name
                existing.sap_date     = rec.sap_date
                existing.leistungsart = rec.leistungsart
                existing.mwst_code    = rec.mwst_code
                existing.waehrung     = rec.waehrung
                existing.total        = rec.total
                existing.positions_json = positions_json
                existing.is_storno    = rec.is_storno
                existing.updated_at   = datetime.utcnow()
                updated += 1
        else:
            db.add(InvoiceRecord(
                group_key=rec.group_key, kunden_nr=rec.kunden_nr, name=rec.name,
                periode=rec.periode, sap_date=rec.sap_date, leistungsart=rec.leistungsart,
                mwst_code=rec.mwst_code, waehrung=rec.waehrung, total=rec.total,
                status=rec.status, invoice_nr=rec.invoice_nr, order_nr=rec.order_nr,
                positions_json=positions_json, is_storno=rec.is_storno, bukrs=rec.bukrs,
            ))
            inserted += 1
    db.commit()
    return {"inserted": inserted, "updated": updated}


@app.patch("/invoice_records/{record_id}")
def update_invoice_record(
    record_id: int,
    body: InvoiceStatusIn,
    db: Session = Depends(get_db),
):
    rec = db.get(InvoiceRecord, record_id)
    if not rec:
        raise HTTPException(404, "Datensatz nicht gefunden")
    rec.status     = body.status
    rec.invoice_nr = body.invoice_nr
    rec.order_nr   = body.order_nr
    rec.updated_at = datetime.utcnow()
    db.commit()
    return _invoice_to_dict(rec)


@app.delete("/invoice_records/{record_id}", status_code=204)
def delete_invoice_record(record_id: int, db: Session = Depends(get_db)):
    """Löscht einen offenen Rechnungsdatensatz (nur Status 'offen' erlaubt)."""
    rec = db.get(InvoiceRecord, record_id)
    if not rec:
        raise HTTPException(404, "Datensatz nicht gefunden")
    if rec.status not in ("offen", "open"):
        raise HTTPException(409, f"Nur offene Rechnungen können gelöscht werden (Status: {rec.status})")
    db.delete(rec)
    db.commit()
    return None


def _invoice_to_dict(r: InvoiceRecord) -> dict:
    try:
        positions = _json.loads(r.positions_json or "[]")
    except Exception:
        positions = []
    # PDF vorhanden? Suche nach invoice_nr.pdf im Rechnungen-Verzeichnis
    has_pdf = False
    if r.invoice_nr:
        has_pdf = (RECHNUNGEN_DIR / f"{r.invoice_nr}.pdf").exists()
    return {
        "id": r.id, "group_key": r.group_key, "kunden_nr": r.kunden_nr,
        "name": r.name, "periode": r.periode, "sap_date": r.sap_date,
        "leistungsart": r.leistungsart, "mwst_code": r.mwst_code,
        "waehrung": r.waehrung, "total": r.total, "status": r.status,
        "invoice_nr": r.invoice_nr, "order_nr": r.order_nr,
        "positions": positions, "is_storno": bool(r.is_storno),
        "bukrs": getattr(r, "bukrs", "0435") or "0435",
        "has_pdf": has_pdf,
        "imported_at": r.imported_at.isoformat() if r.imported_at else "",
        "updated_at":  r.updated_at.isoformat()  if r.updated_at  else "",
    }


# ─── Rechnungs-PDF Upload / Download ────────────────────────────────────────

@app.post("/invoice_records/{record_id}/pdf", status_code=200)
async def upload_invoice_pdf(record_id: int, file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Speichert eine Rechnungs-PDF unter rechnungen/{invoice_nr}.pdf."""
    rec = db.get(InvoiceRecord, record_id)
    if not rec:
        raise HTTPException(404, "Datensatz nicht gefunden")
    if not rec.invoice_nr:
        raise HTTPException(400, "Faktura-Nr. fehlt — bitte zuerst eintragen")
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Nur PDF-Dateien erlaubt")

    dest = RECHNUNGEN_DIR / f"{rec.invoice_nr}.pdf"
    content = await file.read()
    dest.write_bytes(content)
    log.info("PDF gespeichert: %s (%d Bytes)", dest, len(content))
    return {"ok": True, "path": str(dest), "size": len(content)}


@app.get("/invoice_records/{record_id}/pdf")
def download_invoice_pdf(record_id: int, db: Session = Depends(get_db)):
    """Liefert die gespeicherte PDF-Rechnung."""
    from fastapi.responses import FileResponse
    rec = db.get(InvoiceRecord, record_id)
    if not rec:
        raise HTTPException(404, "Datensatz nicht gefunden")
    if not rec.invoice_nr:
        raise HTTPException(404, "Keine Faktura-Nr. vorhanden")
    dest = RECHNUNGEN_DIR / f"{rec.invoice_nr}.pdf"
    if not dest.exists():
        raise HTTPException(404, "PDF nicht vorhanden")
    return FileResponse(
        path=str(dest),
        media_type="application/pdf",
        filename=f"Rechnung_{rec.invoice_nr}.pdf",
    )


# ============================================================================
# INVOICE E-MAIL — Konfiguration & Versand
# ============================================================================

def _email_config_to_dict(c: "InvoiceEmailConfig") -> dict:
    return {
        "id": c.id, "kunden_nr": c.kunden_nr, "name": c.name,
        "recipient_email": c.recipient_email, "sender_email": c.sender_email,
        "smtp_host": c.smtp_host, "smtp_port": c.smtp_port,
        "smtp_user": c.smtp_user, "smtp_pass": c.smtp_pass,
        "smtp_tls": c.smtp_tls,
        "subject_template": c.subject_template, "body_template": c.body_template,
        "status": c.status, "auto_send": c.auto_send,
        "created_at": c.created_at.isoformat(), "updated_at": c.updated_at.isoformat(),
    }


def _email_history_to_dict(h: "InvoiceEmailHistory") -> dict:
    return {
        "id": h.id, "invoice_record_id": h.invoice_record_id,
        "kunden_nr": h.kunden_nr, "invoice_nr": h.invoice_nr,
        "periode": h.periode, "recipient_email": h.recipient_email,
        "sender_email": h.sender_email, "subject": h.subject,
        "status": h.status, "error_msg": h.error_msg,
        "sent_at": h.sent_at.isoformat(),
    }


def _send_invoice_email(cfg: "InvoiceEmailConfig", rec: "InvoiceRecord", db: "Session") -> dict:
    """Versendet die Rechnung per SMTP und schreibt Eintrag in invoice_email_history."""
    import smtplib, ssl
    from email.message import EmailMessage

    invoice_nr = rec.invoice_nr or ""
    periode    = rec.periode    or ""
    name       = rec.name       or cfg.name or ""

    subject = cfg.subject_template.format(invoice_nr=invoice_nr, name=name, periode=periode,
                                          kunden_nr=rec.kunden_nr)
    body    = cfg.body_template.format(invoice_nr=invoice_nr, name=name, periode=periode,
                                       kunden_nr=rec.kunden_nr)

    hist = InvoiceEmailHistory(
        invoice_record_id=rec.id, kunden_nr=rec.kunden_nr,
        invoice_nr=invoice_nr, periode=periode,
        recipient_email=cfg.recipient_email, sender_email=cfg.sender_email,
        subject=subject, status="ausstehend",
    )
    db.add(hist); db.commit(); db.refresh(hist)

    try:
        # SMTP-Konfiguration prüfen
        if not cfg.smtp_host or not cfg.smtp_host.strip():
            raise ValueError(
                "SMTP-Server nicht konfiguriert. "
                "Bitte smtp_host im E-Mail-Tab (Rechnungserstellung → E-Mail → Konfiguration bearbeiten) eintragen."
            )
        if not cfg.smtp_user or not cfg.smtp_pass:
            raise ValueError(
                "SMTP-Zugangsdaten fehlen (Benutzername / Passwort). "
                "Bitte im E-Mail-Tab eintragen."
            )

        msg = EmailMessage()
        msg["From"]    = cfg.sender_email
        msg["To"]      = cfg.recipient_email
        msg["Subject"] = subject
        msg.set_content(body)

        # PDF anhängen wenn vorhanden
        pdf_path = RECHNUNGEN_DIR / f"{invoice_nr}.pdf"
        if pdf_path.exists():
            msg.add_attachment(pdf_path.read_bytes(),
                               maintype="application", subtype="pdf",
                               filename=f"Rechnung_{invoice_nr}.pdf")

        smtp_host = cfg.smtp_host.strip()
        if cfg.smtp_tls:
            ctx = ssl.create_default_context()
            with smtplib.SMTP(smtp_host, cfg.smtp_port, timeout=15) as s:
                s.starttls(context=ctx)
                s.login(cfg.smtp_user, cfg.smtp_pass)
                s.send_message(msg)
        else:
            with smtplib.SMTP_SSL(smtp_host, cfg.smtp_port, context=ssl.create_default_context(), timeout=15) as s:
                s.login(cfg.smtp_user, cfg.smtp_pass)
                s.send_message(msg)

        hist.status = "gesendet"
        hist.error_msg = ""
        db.commit()
        log.info("E-Mail gesendet: %s → %s", invoice_nr, cfg.recipient_email)
        return {"ok": True, "history_id": hist.id}

    except Exception as e:
        hist.status   = "fehler"
        hist.error_msg = str(e)
        db.commit()
        log.error("E-Mail Fehler: %s", e)
        return {"ok": False, "error": str(e), "history_id": hist.id}


# ── E-Mail Konfiguration CRUD ────────────────────────────────────────────────

@app.get("/invoice_email_configs", tags=["E-Mail"])
def list_email_configs(db: Session = Depends(get_db)):
    rows = db.query(InvoiceEmailConfig).order_by(InvoiceEmailConfig.kunden_nr).all()
    return [_email_config_to_dict(r) for r in rows]


@app.post("/invoice_email_configs", status_code=201, tags=["E-Mail"])
def create_email_config(payload: InvoiceEmailConfigIn, db: Session = Depends(get_db)):
    existing = db.query(InvoiceEmailConfig).filter_by(kunden_nr=payload.kunden_nr).first()
    if existing:
        raise HTTPException(409, f"Konfiguration für Kd. {payload.kunden_nr} existiert bereits")
    cfg = InvoiceEmailConfig(**payload.model_dump())
    db.add(cfg); db.commit(); db.refresh(cfg)
    return _email_config_to_dict(cfg)


@app.put("/invoice_email_configs/{cfg_id}", tags=["E-Mail"])
def update_email_config(cfg_id: int, payload: InvoiceEmailConfigIn, db: Session = Depends(get_db)):
    cfg = db.get(InvoiceEmailConfig, cfg_id)
    if not cfg:
        raise HTTPException(404, "Konfiguration nicht gefunden")
    for k, v in payload.model_dump().items():
        setattr(cfg, k, v)
    cfg.updated_at = datetime.utcnow()
    db.commit(); db.refresh(cfg)
    return _email_config_to_dict(cfg)


@app.delete("/invoice_email_configs/{cfg_id}", status_code=204, tags=["E-Mail"])
def delete_email_config(cfg_id: int, db: Session = Depends(get_db)):
    cfg = db.get(InvoiceEmailConfig, cfg_id)
    if not cfg:
        raise HTTPException(404, "Konfiguration nicht gefunden")
    db.delete(cfg); db.commit()


# ── E-Mail Versand ───────────────────────────────────────────────────────────

@app.post("/invoice_records/{record_id}/send_email", tags=["E-Mail"])
def send_invoice_email(record_id: int, db: Session = Depends(get_db)):
    """Versendet die Rechnung per E-Mail an die konfigurierte Adresse."""
    rec = db.get(InvoiceRecord, record_id)
    if not rec:
        raise HTTPException(404, "Datensatz nicht gefunden")
    if not rec.invoice_nr:
        raise HTTPException(400, "Faktura-Nr. fehlt")
    cfg = db.query(InvoiceEmailConfig).filter_by(kunden_nr=rec.kunden_nr).first()
    if not cfg:
        raise HTTPException(404, f"Keine E-Mail-Konfiguration für Kunde {rec.kunden_nr}")
    if cfg.status != "aktiv":
        raise HTTPException(400, "E-Mail-Konfiguration ist inaktiv")
    return _send_invoice_email(cfg, rec, db)


# ── E-Mail Historie ──────────────────────────────────────────────────────────

@app.get("/invoice_email_history", tags=["E-Mail"])
def list_email_history(
    kunden_nr: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(InvoiceEmailHistory)
    if kunden_nr:
        q = q.filter(InvoiceEmailHistory.kunden_nr == kunden_nr)
    rows = q.order_by(InvoiceEmailHistory.sent_at.desc()).limit(limit).all()
    return [_email_history_to_dict(r) for r in rows]


# ── KI-Provider-Konfiguration ───────────────────────────────────────────────

@app.get("/ai_config", tags=["KI-Konfiguration"])
def get_ai_config(db: Session = Depends(get_db)):
    """Aktuelle KI-Konfiguration (API-Key wird maskiert zurückgegeben)."""
    row = db.query(AiConfig).first()
    if not row:
        return {"provider": "claude", "api_key": "", "api_key_set": False, "model": "", "updated_at": None}
    masked = ("*" * (len(row.api_key) - 4) + row.api_key[-4:]) if len(row.api_key) > 4 else ("*" * len(row.api_key))
    return {
        "provider":    row.provider,
        "api_key":     masked,
        "api_key_set": bool(row.api_key),
        "model":       row.model,
        "updated_at":  row.updated_at.isoformat() if row.updated_at else None,
    }


@app.put("/ai_config", tags=["KI-Konfiguration"])
def save_ai_config(
    provider: str = "",
    api_key:  str = "",
    model:    str = "",
    db: Session = Depends(get_db),
):
    """KI-Konfiguration speichern. Leeres api_key → bestehenden Key behalten."""
    row = db.query(AiConfig).first()
    if not row:
        row = AiConfig()
        db.add(row)
    if provider: row.provider = provider
    if api_key:  row.api_key  = api_key
    if model:    row.model    = model
    row.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True, "provider": row.provider, "model": row.model}


@app.get("/ai_config/full", tags=["KI-Konfiguration"])
def get_ai_config_full(db: Session = Depends(get_db)):
    """Vollständige Konfiguration inkl. API-Key – nur für internen voice_bot-Aufruf."""
    row = db.query(AiConfig).first()
    if not row:
        return {"provider": "claude", "api_key": os.getenv("ANTHROPIC_API_KEY", ""), "model": ""}
    # Fallback: wenn kein Key in DB, Env-Variable nehmen
    key = row.api_key or os.getenv("ANTHROPIC_API_KEY", "")
    return {"provider": row.provider, "api_key": key, "model": row.model}


@app.get("/ai_config/test", tags=["KI-Konfiguration"])
def test_ai_config(db: Session = Depends(get_db)):
    """Testet die gespeicherte KI-Konfiguration mit einem Mini-Aufruf."""
    import urllib.request as _ur, json as _js
    row = db.query(AiConfig).first()
    if not row:
        return {"ok": False, "error": "Keine Konfiguration in der Datenbank", "provider": None}
    key      = row.api_key or os.getenv("ANTHROPIC_API_KEY", "")
    provider = row.provider or "claude"
    model    = row.model or ""
    if not key:
        return {"ok": False, "error": "API-Key fehlt", "provider": provider}
    try:
        if provider == "openai":
            model = model or "gpt-4o-mini"
            payload = _js.dumps({
                "model": model, "max_tokens": 5,
                "messages": [{"role": "user", "content": "Ping"}],
            }).encode()
            req = _ur.Request(
                "https://api.openai.com/v1/chat/completions",
                data=payload,
                headers={"Content-Type": "application/json", "Authorization": f"Bearer {key}"},
                method="POST",
            )
            with _ur.urlopen(req, timeout=15) as r:
                result = _js.loads(r.read().decode())
            text = result["choices"][0]["message"]["content"]
        else:
            # Claude via urllib (kein anthropic-Paket nötig)
            model = model or "claude-haiku-4-5-20251001"
            payload = _js.dumps({
                "model": model, "max_tokens": 5,
                "messages": [{"role": "user", "content": "Ping"}],
            }).encode()
            req = _ur.Request(
                "https://api.anthropic.com/v1/messages",
                data=payload,
                headers={
                    "Content-Type": "application/json",
                    "x-api-key": key,
                    "anthropic-version": "2023-06-01",
                },
                method="POST",
            )
            with _ur.urlopen(req, timeout=15) as r:
                result = _js.loads(r.read().decode())
            text = result["content"][0]["text"]
        return {"ok": True, "provider": provider, "model": model, "response": text}
    except Exception as exc:
        return {"ok": False, "provider": provider, "model": model, "error": str(exc)}


# ============================================================================
# 3W SMART DATABASE ENGINE -- API-Endpunkte
# ============================================================================

@app.post("/3w/answers", response_model=ThreeWAnswerOut, status_code=201, tags=["3W-Engine"])
def create_3w_answer(payload: ThreeWAnswerIn, db: Session = Depends(get_db)):
    item = ThreeWAnswer(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item


@app.get("/3w/answers", tags=["3W-Engine"])
def list_3w_answers(
    module: Optional[str] = None, who: Optional[str] = None,
    action_type: Optional[str] = None, sap_system: Optional[str] = None,
    periode: Optional[str] = None, limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(ThreeWAnswer)
    if module:      q = q.filter(ThreeWAnswer.module      == module)
    if who:         q = q.filter(ThreeWAnswer.who         == who)
    if action_type: q = q.filter(ThreeWAnswer.action_type == action_type)
    if sap_system:  q = q.filter(ThreeWAnswer.sap_system  == sap_system)
    if periode:     q = q.filter(ThreeWAnswer.periode      == periode)
    rows = q.order_by(ThreeWAnswer.when_ts.desc()).limit(limit).all()
    return [{"id": r.id, "who": r.who, "action_type": r.action_type,
             "action_desc": r.action_desc, "object_ref": r.object_ref,
             "module": r.module, "business_rule": r.business_rule,
             "why_desc": r.why_desc, "wie": r.wie, "wo": r.wo,
             "womit": r.womit, "risiken": r.risiken, "alternativen": r.alternativen,
             "erfahrung": r.erfahrung, "ergebnis": r.ergebnis,
             "sap_system": r.sap_system, "comp_code": r.comp_code,
             "periode": r.periode, "status": r.status,
             "when_ts": r.when_ts.isoformat()} for r in rows]


@app.get("/3w/stats", tags=["3W-Engine"])
def get_3w_stats(db: Session = Depends(get_db)):
    return {
        "three_w_answers":       db.query(ThreeWAnswer).count(),
        "knowledge_items":       db.query(KnowledgeItem).count(),
        "learning_patterns":     db.query(LearningPattern).count(),
        "structure_suggestions": db.query(StructureSuggestion).count(),
        "governance_rules":      db.query(GovernanceRule).count(),
        "knowledge_universe":    db.query(KnowledgeUniverse).count(),
        "knowledge_relations":   db.query(KnowledgeRelation).count(),
        "decision_requests":     db.query(DecisionRequest).count(),
    }


@app.post("/3w/knowledge", response_model=KnowledgeItemOut, status_code=201, tags=["3W-Engine"])
def create_knowledge_item(payload: KnowledgeItemIn, db: Session = Depends(get_db)):
    existing = db.query(KnowledgeItem).filter(KnowledgeItem.key == payload.key).first()
    if existing:
        for k, v in payload.model_dump().items():
            setattr(existing, k, v)
        existing.updated_at = datetime.utcnow()
        db.commit(); db.refresh(existing)
        return existing
    item = KnowledgeItem(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item


@app.get("/3w/knowledge", tags=["3W-Engine"])
def list_knowledge_items(
    category: Optional[str] = None, source_module: Optional[str] = None,
    active: Optional[int] = 1, db: Session = Depends(get_db),
):
    q = db.query(KnowledgeItem)
    if category:      q = q.filter(KnowledgeItem.category      == category)
    if source_module: q = q.filter(KnowledgeItem.source_module == source_module)
    if active is not None: q = q.filter(KnowledgeItem.active   == active)
    rows = q.order_by(KnowledgeItem.updated_at.desc()).limit(200).all()
    return [{"id": r.id, "category": r.category, "key": r.key, "title": r.title,
             "body": r.body, "source_module": r.source_module, "source_ref": r.source_ref,
             "tags_json": r.tags_json, "confidence": r.confidence, "active": r.active,
             "created_at": r.created_at.isoformat(),
             "updated_at": r.updated_at.isoformat()} for r in rows]


@app.post("/3w/patterns", status_code=201, tags=["3W-Engine"])
def upsert_learning_pattern(
    pattern_type: str, pattern_key: str,
    description: str = "", data_json: str = "{}",
    db: Session = Depends(get_db),
):
    p = db.query(LearningPattern).filter(
        LearningPattern.pattern_type == pattern_type,
        LearningPattern.pattern_key  == pattern_key,
    ).first()
    if p:
        p.occurrence_count += 1
        p.last_seen = datetime.utcnow()
        if description: p.description = description
        if data_json:   p.data_json   = data_json
        db.commit()
        return {"id": p.id, "occurrence_count": p.occurrence_count}
    p = LearningPattern(
        pattern_type=pattern_type, pattern_key=pattern_key,
        description=description, data_json=data_json,
    )
    db.add(p); db.commit(); db.refresh(p)
    return {"id": p.id, "occurrence_count": 1}


@app.get("/3w/patterns", tags=["3W-Engine"])
def list_learning_patterns(pattern_type: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(LearningPattern)
    if pattern_type: q = q.filter(LearningPattern.pattern_type == pattern_type)
    rows = q.order_by(LearningPattern.occurrence_count.desc()).limit(100).all()
    return [{"id": r.id, "pattern_type": r.pattern_type, "pattern_key": r.pattern_key,
             "description": r.description, "occurrence_count": r.occurrence_count,
             "last_seen": r.last_seen.isoformat(), "data_json": r.data_json} for r in rows]


@app.post("/3w/suggestions", response_model=StructureSuggestionOut, status_code=201, tags=["3W-Engine"])
def create_suggestion(payload: StructureSuggestionIn, db: Session = Depends(get_db)):
    item = StructureSuggestion(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item


@app.get("/3w/suggestions", tags=["3W-Engine"])
def list_suggestions(status: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(StructureSuggestion)
    if status: q = q.filter(StructureSuggestion.status == status)
    rows = q.order_by(StructureSuggestion.proposed_at.desc()).limit(100).all()
    return [{"id": r.id, "stype": r.stype, "title": r.title, "description": r.description,
             "priority": r.priority, "status": r.status, "proposed_by": r.proposed_by,
             "proposed_at": r.proposed_at.isoformat(),
             "notes": r.notes} for r in rows]


@app.patch("/3w/suggestions/{suggestion_id}", tags=["3W-Engine"])
def update_suggestion(suggestion_id: int, status: str, notes: str = "",
                      db: Session = Depends(get_db)):
    item = db.get(StructureSuggestion, suggestion_id)
    if not item:
        raise HTTPException(404, "Vorschlag nicht gefunden")
    item.status = status
    if notes: item.notes = notes
    if status in ("accepted", "rejected", "done"):
        item.decided_at = datetime.utcnow()
    db.commit()
    return {"id": item.id, "status": item.status}


@app.get("/3w/rules", tags=["3W-Engine"])
def list_governance_rules(
    module: Optional[str] = None, active: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(GovernanceRule)
    if module: q = q.filter(GovernanceRule.module == module)
    if active is not None: q = q.filter(GovernanceRule.active == active)
    rows = q.order_by(GovernanceRule.rule_id).all()
    return [{"id": r.id, "rule_id": r.rule_id, "module": r.module,
             "title": r.title, "description": r.description,
             "priority": r.priority, "active": r.active,
             "created_at": r.created_at.isoformat()} for r in rows]


@app.post("/3w/rules", response_model=GovernanceRuleOut, status_code=201, tags=["3W-Engine"])
def create_governance_rule(payload: GovernanceRuleIn, db: Session = Depends(get_db)):
    existing = db.query(GovernanceRule).filter(GovernanceRule.rule_id == payload.rule_id).first()
    if existing:
        for k, v in payload.model_dump().items():
            setattr(existing, k, v)
        db.commit(); db.refresh(existing)
        return existing
    item = GovernanceRule(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item


@app.patch("/3w/rules/{rule_id_str}", tags=["3W-Engine"])
def toggle_governance_rule(rule_id_str: str, active: int, db: Session = Depends(get_db)):
    item = db.query(GovernanceRule).filter(GovernanceRule.rule_id == rule_id_str).first()
    if not item:
        raise HTTPException(404, "Regel nicht gefunden")
    item.active = active
    db.commit()
    return {"rule_id": item.rule_id, "active": item.active}


# ============================================================================
# MODULE VERSIONS API — Datei-Snapshots / Backup
# ============================================================================

@app.get("/module-versions", tags=["Versionierung"])
def list_module_versions(db: Session = Depends(get_db)):
    """Alle Module mit ihrer letzten Version."""
    rows = (
        db.query(ModuleVersion)
        .order_by(ModuleVersion.module_name, ModuleVersion.version_number.desc())
        .all()
    )
    # Pro Modul die neueste + alle älteren zurückgeben
    result = []
    for r in rows:
        result.append({
            "id":             r.id,
            "module_name":    r.module_name,
            "version_number": r.version_number,
            "file_path":      r.file_path,
            "file_size":      r.file_size,
            "file_hash":      r.file_hash[:12],
            "triggered_by":   r.triggered_by,
            "description":    r.description,
            "created_at":     r.created_at.isoformat() if r.created_at else "",
        })
    return result


@app.get("/module-versions/{version_id}/content", tags=["Versionierung"])
def get_module_version_content(version_id: int, db: Session = Depends(get_db)):
    """Vollständigen Dateiinhalt einer Version zurückgeben."""
    row = db.query(ModuleVersion).filter(ModuleVersion.id == version_id).first()
    if not row:
        raise HTTPException(404, "Version nicht gefunden")
    return {
        "id":             row.id,
        "module_name":    row.module_name,
        "version_number": row.version_number,
        "file_path":      row.file_path,
        "file_content":   row.file_content,
        "file_size":      row.file_size,
        "triggered_by":   row.triggered_by,
        "description":    row.description,
        "created_at":     row.created_at.isoformat() if row.created_at else "",
    }


@app.post("/module-versions/{version_id}/restore", tags=["Versionierung"])
def restore_module_version(version_id: int, db: Session = Depends(get_db)):
    """Datei einer Version wiederherstellen (Backup→Datei schreiben)."""
    import shutil, os
    row = db.query(ModuleVersion).filter(ModuleVersion.id == version_id).first()
    if not row:
        raise HTTPException(404, "Version nicht gefunden")
    file_path = row.file_path
    # Aktuellen Stand als Sicherung daneben speichern
    if os.path.exists(file_path):
        backup_path = file_path + ".before_restore"
        shutil.copy2(file_path, backup_path)
    try:
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(row.file_content)
    except Exception as e:
        raise HTTPException(500, f"Schreiben fehlgeschlagen: {e}")
    return {
        "status":         "restored",
        "module_name":    row.module_name,
        "version_number": row.version_number,
        "file_path":      file_path,
    }


@app.post("/module-versions/save", tags=["Versionierung"], status_code=201)
def save_module_version_manual(
    payload: dict,
    db: Session = Depends(get_db)
):
    """Manuellen Snapshot einer Datei speichern.
    Body: { file_path, triggered_by?, description? }
    """
    import hashlib, os
    file_path    = payload.get("file_path", "")
    triggered_by = payload.get("triggered_by", "manuell")
    description  = payload.get("description", "")
    if not file_path or not os.path.exists(file_path):
        raise HTTPException(400, f"Datei nicht gefunden: {file_path}")
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    file_hash   = hashlib.sha256(content.encode()).hexdigest()
    module_name = os.path.basename(file_path)
    # Duplikat-Schutz
    last = (db.query(ModuleVersion)
              .filter(ModuleVersion.module_name == module_name)
              .order_by(ModuleVersion.version_number.desc())
              .first())
    if last and last.file_hash == file_hash:
        return {"status": "unchanged", "version_number": last.version_number, "id": last.id}
    next_ver = (last.version_number + 1) if last else 1
    row = ModuleVersion(
        module_name    = module_name,
        version_number = next_ver,
        file_path      = file_path,
        file_content   = content,
        file_hash      = file_hash,
        file_size      = len(content.encode()),
        triggered_by   = triggered_by,
        description    = description,
    )
    db.add(row); db.commit(); db.refresh(row)
    return {"status": "saved", "version_number": next_ver, "id": row.id}


@app.delete("/module-versions/{version_id}", tags=["Versionierung"])
def delete_module_version(version_id: int, db: Session = Depends(get_db)):
    """Einzelne Version aus der DB löschen."""
    row = db.query(ModuleVersion).filter(ModuleVersion.id == version_id).first()
    if not row:
        raise HTTPException(404, "Version nicht gefunden")
    db.delete(row); db.commit()
    return {"status": "deleted", "id": version_id}


# ============================================================================
# KNOWLEDGE UNIVERSE API -- 13-Schichten-Wissensmodell
# ============================================================================

@app.post("/knowledge", response_model=KnowledgeUniverseOut, status_code=201, tags=["Knowledge-Universe"])
def create_knowledge(payload: KnowledgeUniverseIn, db: Session = Depends(get_db)):
    existing = db.query(KnowledgeUniverse).filter(KnowledgeUniverse.key == payload.key).first()
    if existing:
        for k, v in payload.model_dump().items():
            setattr(existing, k, v)
        existing.updated_at = datetime.utcnow()
        db.commit(); db.refresh(existing)
        return existing
    item = KnowledgeUniverse(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return item


@app.get("/knowledge", tags=["Knowledge-Universe"])
def list_knowledge(
    layer: Optional[str] = None, object_type: Optional[str] = None,
    source_module: Optional[str] = None, active: Optional[int] = 1,
    search: Optional[str] = None, limit: int = 100,
    db: Session = Depends(get_db),
):
    q = db.query(KnowledgeUniverse)
    if layer:         q = q.filter(KnowledgeUniverse.layer == layer)
    if object_type:   q = q.filter(KnowledgeUniverse.object_type == object_type)
    if source_module: q = q.filter(KnowledgeUniverse.source_module == source_module)
    if active is not None: q = q.filter(KnowledgeUniverse.active == active)
    if search:
        like = f"%{search}%"
        q = q.filter(
            (KnowledgeUniverse.title.ilike(like)) |
            (KnowledgeUniverse.summary.ilike(like)) |
            (KnowledgeUniverse.tags_json.ilike(like))
        )
    rows = q.order_by(KnowledgeUniverse.importance.desc(),
                      KnowledgeUniverse.updated_at.desc()).limit(limit).all()
    return [{"id": r.id, "layer": r.layer, "object_type": r.object_type,
             "key": r.key, "title": r.title, "summary": r.summary,
             "confidence": r.confidence, "importance": r.importance,
             "source_module": r.source_module, "source_type": r.source_type,
             "created_by": r.created_by, "tags_json": r.tags_json,
             "active": r.active, "created_at": r.created_at.isoformat(),
             "updated_at": r.updated_at.isoformat()} for r in rows]


@app.get("/knowledge/universe/{ku_id}", tags=["Knowledge-Universe"])
def get_ku_entry(ku_id: int, db: Session = Depends(get_db)):
    r = db.get(KnowledgeUniverse, ku_id)
    if not r:
        raise HTTPException(404)
    return {"id": r.id, "layer": r.layer, "object_type": r.object_type,
            "key": r.key, "title": r.title, "summary": r.summary,
            "confidence": r.confidence, "importance": r.importance,
            "source_module": r.source_module, "source_type": r.source_type,
            "created_by": r.created_by, "tags_json": r.tags_json,
            "active": r.active, "created_at": r.created_at.isoformat(),
            "updated_at": r.updated_at.isoformat()}


@app.patch("/knowledge/universe/{ku_id}", tags=["Knowledge-Universe"])
def patch_ku_entry(ku_id: int, payload: dict, db: Session = Depends(get_db)):
    r = db.get(KnowledgeUniverse, ku_id)
    if not r:
        raise HTTPException(404)
    for k, v in payload.items():
        if hasattr(r, k):
            setattr(r, k, v)
    r.updated_at = datetime.utcnow()
    db.commit()
    return {"ok": True}


@app.delete("/knowledge/universe/{ku_id}", status_code=204, tags=["Knowledge-Universe"])
def delete_ku_entry(ku_id: int, db: Session = Depends(get_db)):
    r = db.get(KnowledgeUniverse, ku_id)
    if r:
        db.delete(r); db.commit()


# ── Decision Requests ────────────────────────────────────────────────────────

class DecisionRequestIn(BaseModel):
    question:       str
    recommendation: str = ""
    confidence:     float = 0.0
    module:         str = ""
    asked_by:       str = ""


@app.post("/decisions", response_model=dict, status_code=201, tags=["Decision-Engine"])
def create_decision(payload: DecisionRequestIn, db: Session = Depends(get_db)):
    item = DecisionRequest(**payload.model_dump())
    db.add(item); db.commit(); db.refresh(item)
    return {"id": item.id, "status": item.status}


# ============================================================================
# KNOWLEDGE DECISION ENGINE -- Argumentations- und Begruendungslogik
# ============================================================================

def _run_decision_engine(question: str, context: dict, module: str, db) -> dict:
    """Regelbasierte Entscheidungslogik aus Knowledge Universe + Governance Rules."""
    import json as _json2
    keywords = [w for w in question.lower().split() if len(w) > 3][:8]
    knowledge_refs, pro_args, contra_args, risks, alternatives = [], [], [], [], []
    confidence_sum, confidence_count = 0, 0

    for kw in keywords:
        items = db.query(KnowledgeUniverse).filter(
            KnowledgeUniverse.active == 1,
            (KnowledgeUniverse.title.ilike(f"%{kw}%")) |
            (KnowledgeUniverse.summary.ilike(f"%{kw}%")) |
            (KnowledgeUniverse.tags_json.ilike(f"%{kw}%"))
        ).limit(5).all()
        for item in items:
            if item.id not in [r["id"] for r in knowledge_refs]:
                knowledge_refs.append({"id": item.id, "layer": item.layer,
                                       "title": item.title, "confidence": item.confidence})
                confidence_sum += item.confidence; confidence_count += 1
                if item.layer in ("rule", "finance", "sap", "business"):
                    pro_args.append(f"[{item.layer.upper()}] {item.title}: {item.summary[:120]}")
                elif item.layer == "risk":
                    risks.append(f"{item.title}: {item.summary[:120]}")
                elif item.layer == "experience":
                    pro_args.append(f"[ERFAHRUNG] {item.title}: {item.summary[:120]}")

    rules = db.query(GovernanceRule).filter(GovernanceRule.active == 1).all()
    matching_rules = []
    for rule in rules:
        rule_terms = (rule.title + " " + rule.description).lower()
        if any(kw in rule_terms for kw in keywords):
            matching_rules.append(rule)
            pro_args.append(f"[APP-REGEL {rule.rule_id}] {rule.title}")
            confidence_sum += 4; confidence_count += 1

    past = db.query(DecisionRequest).filter(
        DecisionRequest.status == "archived",
        DecisionRequest.outcome_correct == 1
    ).order_by(DecisionRequest.decided_at.desc()).limit(3).all()
    for p in past:
        if any(kw in p.question.lower() for kw in keywords):
            alternatives.append(f"Fruehere Entscheidung: {p.question[:80]} -> {p.recommendation[:80]}")

    risk_items = db.query(KnowledgeUniverse).filter(
        KnowledgeUniverse.layer == "risk", KnowledgeUniverse.active == 1).limit(5).all()
    for ri in risk_items:
        if any(kw in (ri.title + ri.summary).lower() for kw in keywords):
            if ri.title not in risks:
                risks.append(f"{ri.title}: {ri.summary[:100]}")
            contra_args.append(f"[RISIKO] {ri.title}")

    patterns = db.query(LearningPattern).filter(
        LearningPattern.pattern_type == "fehler").limit(3).all()
    for pat in patterns:
        if any(kw in pat.description.lower() for kw in keywords):
            contra_args.append(f"[LERNMUSTER] {pat.description[:100]}")

    if confidence_count > 0:
        avg_conf = confidence_sum / confidence_count
        rule_bonus = min(len(matching_rules) * 0.3, 1.0)
        raw_conf = min(avg_conf + rule_bonus, 5.0)
    else:
        raw_conf = 2.0
    confidence = max(1, min(5, round(raw_conf)))

    if confidence >= 4 and len(pro_args) > 0 and len(contra_args) == 0:
        recommendation = (f"Empfehlung: AUSFUEHREN. {len(pro_args)} unterstuetzende Argumente, "
                          f"keine kritischen Risiken. Konfidenz: {confidence}/5.")
    elif confidence >= 3:
        recommendation = (f"Empfehlung: MIT VORSICHT AUSFUEHREN. {len(pro_args)} Pro-Argumente, "
                          f"{len(contra_args)} Gegenargumente. Konfidenz: {confidence}/5.")
    else:
        recommendation = (f"Empfehlung: PRUEFUNG ERFORDERLICH. Unzureichendes Wissen. "
                          f"Konfidenz: {confidence}/5. Manuelle Ueberpruefung empfohlen.")

    rationale = (f"Basierend auf {len(knowledge_refs)} Wissensobjekten, "
                 f"{len(matching_rules)} App-Regeln, {len(past)} frueheren Entscheidungen. "
                 f"Modul: {module or 'global'}.")

    return {
        "recommendation":    recommendation,
        "arguments_pro":     _json2.dumps(pro_args,       ensure_ascii=False),
        "arguments_contra":  _json2.dumps(contra_args,    ensure_ascii=False),
        "risks_json":        _json2.dumps(risks,           ensure_ascii=False),
        "alternatives_json": _json2.dumps(alternatives,   ensure_ascii=False),
        "knowledge_refs":    _json2.dumps(knowledge_refs,  ensure_ascii=False),
        "confidence":        confidence,
        "rationale":         rationale,
    }


@app.post("/decisions", response_model=DecisionOut, status_code=201, tags=["Decision-Engine"])
def request_decision(payload: DecisionRequestIn, db: Session = Depends(get_db)):
    """Knowledge Decision Engine: Fragt Knowledge Universe ab und gibt begruendete Empfehlung."""
    import json as _json2
    ctx = {}
    try: ctx = _json2.loads(payload.context_json)
    except Exception: pass
    result = _run_decision_engine(payload.question, ctx, payload.module, db)
    req = DecisionRequest(
        question=payload.question, context_json=payload.context_json,
        asked_by=payload.asked_by, module=payload.module,
        status="decided", decided_by="engine", decided_at=datetime.utcnow(),
        **result,
    )
    db.add(req); db.commit(); db.refresh(req)
    return req


@app.get("/decisions", tags=["Decision-Engine"])
def list_decisions(
    module: Optional[str] = None, asked_by: Optional[str] = None,
    status: Optional[str] = None, limit: int = 50,
    db: Session = Depends(get_db),
):
    q = db.query(DecisionRequest)
    if module:   q = q.filter(DecisionRequest.module   == module)
    if asked_by: q = q.filter(DecisionRequest.asked_by == asked_by)
    if status:   q = q.filter(DecisionRequest.status   == status)
    rows = q.order_by(DecisionRequest.created_at.desc()).limit(limit).all()
    return [{"id": r.id, "question": r.question, "recommendation": r.recommendation,
             "confidence": r.confidence, "status": r.status, "module": r.module,
             "asked_by": r.asked_by, "created_at": r.created_at.isoformat(),
             "decided_at": r.decided_at.isoformat() if r.decided_at else None} for r in rows]


@app.get("/decisions/stats/summary", tags=["Decision-Engine"])
def decision_stats(db: Session = Depends(get_db)):
    total    = db.query(DecisionRequest).count()
    archived = db.query(DecisionRequest).filter(DecisionRequest.status == "archived").count()
    correct  = db.query(DecisionRequest).filter(DecisionRequest.outcome_correct == 1).count()
    return {"total": total, "archived": archived, "correct_outcomes": correct,
            "accuracy": round(correct / max(archived, 1) * 100, 1)}


@app.get("/decisions/{decision_id}", response_model=DecisionOut, tags=["Decision-Engine"])
def get_decision(decision_id: int, db: Session = Depends(get_db)):
    item = db.get(DecisionRequest, decision_id)
    if not item:
        raise HTTPException(404, "Entscheidung nicht gefunden")
    return item


@app.patch("/decisions/{decision_id}/feedback", tags=["Decision-Engine"])
def decision_feedback(
    decision_id: int, outcome_correct: int, outcome_note: str = "",
    db: Session = Depends(get_db),
):
    """Feedback-Loop: War die Entscheidung korrekt? Lerneffekt -> Experience Layer."""
    import json as _json2
    item = db.get(DecisionRequest, decision_id)
    if not item:
        raise HTTPException(404, "Entscheidung nicht gefunden")
    item.outcome_correct = outcome_correct
    item.outcome_note    = outcome_note
    item.status          = "archived"
    db.commit()
    if outcome_correct == 1:
        exp_key = f"decision_outcome_{item.id}"
        if not db.query(KnowledgeUniverse).filter(KnowledgeUniverse.key == exp_key).first():
            exp = KnowledgeUniverse(
                layer="experience", object_type="decision",
                key=exp_key,
                title=f"Entscheidung: {item.question[:80]}",
                summary=item.recommendation[:200],
                body_json=_json2.dumps({"question": item.question,
                    "recommendation": item.recommendation,
                    "confidence": item.confidence, "module": item.module,
                    "outcome_note": outcome_note}, ensure_ascii=False),
                confidence=item.confidence, importance=3,
                source_module=item.module or "decision_engine",
                source_ref=f"decision:{item.id}", source_type="automatic",
                created_by=item.asked_by or "system",
            )
            db.add(exp); db.commit()
            log.info("Lerneffekt: Decision #%d als Erfahrung gespeichert.", item.id)
    return {"id": item.id, "outcome_correct": item.outcome_correct, "status": item.status}


# ============================================================================
# DB-KONFIGURATION  –  GET /db-config  ·  /test  ·  /set  ·  /migrate
# ============================================================================

@app.get("/db-config", tags=["DB-Konfiguration"])
def get_db_config():
    """Aktuelle DB-Verbindung + Tabellenstatistiken."""
    import sqlite3 as _sq
    p = pathlib.Path(_DB_PATH)
    tables: dict = {}
    conn_ok = False
    conn_err = ""
    try:
        con = _sq.connect(_DB_PATH, timeout=5)
        cur = con.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")
        for (tname,) in cur.fetchall():
            try:
                cur.execute(f'SELECT COUNT(*) FROM "{tname}"')
                tables[tname] = cur.fetchone()[0]
            except Exception:
                tables[tname] = -1
        con.close()
        conn_ok = True
    except Exception as exc:
        conn_err = str(exc)

    cfg_meta: dict = {}
    if _CONFIG_FILE.exists():
        try:
            cfg_meta = _json.loads(_CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass

    return {
        "db_path":          _DB_PATH,
        "db_exists":        p.exists(),
        "db_size_bytes":    p.stat().st_size if p.exists() else 0,
        "connection_ok":    conn_ok,
        "connection_error": conn_err,
        "table_count":      len(tables),
        "tables":           tables,
        "config_file":      str(_CONFIG_FILE),
        "config_active":    _CONFIG_FILE.exists(),
        "config_meta":      cfg_meta,
        "default_target":   "Z:\\DB\\orchestrator.db",
    }


@app.post("/db-config/test", tags=["DB-Konfiguration"])
def test_db_path(payload: dict):
    """Testet, ob ein Pfad beschreibbar und als SQLite-DB nutzbar ist."""
    import sqlite3 as _sq
    db_path = payload.get("db_path", "").strip()
    if not db_path:
        raise HTTPException(400, "db_path fehlt")
    p = pathlib.Path(db_path)
    if not p.parent.exists():
        return {"status": "error",
                "message": f"Verzeichnis nicht gefunden: {p.parent}"}
    try:
        con = _sq.connect(str(p), timeout=5)
        con.execute("PRAGMA journal_mode=WAL")
        con.execute("CREATE TABLE IF NOT EXISTS _yconn_probe (id INTEGER PRIMARY KEY)")
        con.execute("INSERT INTO _yconn_probe VALUES (NULL)")
        con.execute("DELETE FROM _yconn_probe")
        con.execute("DROP TABLE _yconn_probe")
        con.commit()
        con.close()
        # Temporäre Testdatei wieder löschen (nur wenn neu angelegt)
        is_new = p.exists() and p.stat().st_size < 4096
        exists_before = p.exists()
        return {
            "status":       "ok",
            "message":      "Verbindung und Schreibzugriff OK",
            "db_path":      str(p),
            "dir_exists":   True,
            "file_exists":  exists_before,
            "size_bytes":   p.stat().st_size if p.exists() else 0,
        }
    except Exception as exc:
        return {"status": "error", "message": str(exc)}


@app.post("/db-config/set", tags=["DB-Konfiguration"])
def set_db_config(payload: dict):
    """Speichert neuen DB-Pfad in db_config.json. Neustart erforderlich."""
    db_path    = payload.get("db_path", "").strip()
    updated_by = payload.get("updated_by", "manuell")
    if not db_path:
        raise HTTPException(400, "db_path fehlt")
    p = pathlib.Path(db_path)
    if not p.parent.exists():
        raise HTTPException(400, f"Verzeichnis nicht gefunden: {p.parent}")
    cfg = {
        "db_path":       str(p),
        "db_dir":        str(p.parent),
        "previous_path": _DB_PATH,
        "updated_at":    datetime.utcnow().isoformat(),
        "updated_by":    updated_by,
    }
    try:
        _CONFIG_FILE.write_text(
            _json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8"
        )
        log.info("DB-Konfiguration geändert → %s (durch %s)", db_path, updated_by)
    except Exception as exc:
        raise HTTPException(500, f"Config-Datei konnte nicht gespeichert werden: {exc}")
    return {
        "status":          "saved",
        "message":         "Konfiguration gespeichert. Orchestrator jetzt neu starten.",
        "db_path":         str(p),
        "restart_required": True,
    }


@app.get("/server-info", tags=["System"])
def server_info():
    """Gibt Netzwerk-IP, Hostname und Zugriffs-URL des Servers zurück."""
    import socket as _sock
    hostname = _sock.gethostname()
    # Primäre Netzwerk-IP ermitteln (nicht 127.0.0.1)
    primary_ip = "127.0.0.1"
    try:
        s = _sock.socket(_sock.AF_INET, _sock.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        primary_ip = s.getsockname()[0]
        s.close()
    except Exception:
        pass
    # Alle Netzwerk-IPs sammeln
    all_ips = []
    try:
        for info in _sock.getaddrinfo(hostname, None, _sock.AF_INET):
            ip = info[4][0]
            if ip not in all_ips:
                all_ips.append(ip)
    except Exception:
        all_ips = [primary_ip]
    port = 8000
    return {
        "hostname":    hostname,
        "primary_ip":  primary_ip,
        "all_ips":     all_ips,
        "port":        port,
        "cockpit_url": f"http://{primary_ip}:{port}/cockpit/startseite.html",
        "login_url":   f"http://{primary_ip}:{port}/cockpit/login.html",
        "api_docs":    f"http://{primary_ip}:{port}/docs",
        "db_path":     _DB_PATH,
    }


# ── Backup ────────────────────────────────────────────────────────────────────
_BACKUP_VERSION_FILE = pathlib.Path(__file__).parent.parent / "backups" / "backup_version.json"
_BACKUP_DIR          = pathlib.Path(__file__).parent.parent / "backups"
_BACKUP_SCRIPT       = pathlib.Path(__file__).parent.parent / "backup.py"


@app.get("/backup/list", tags=["Backup"])
def backup_list():
    """Gibt die Backup-Historie zurück."""
    if not _BACKUP_VERSION_FILE.exists():
        return {"version": 0, "backups": []}
    import json as _j
    try:
        data = _j.loads(_BACKUP_VERSION_FILE.read_text(encoding="utf-8"))
        # Datei-Existenz prüfen
        for b in data.get("backups", []):
            f = _BACKUP_DIR / b["filename"]
            b["exists"] = f.exists()
            b["path"]   = str(f)
        return data
    except Exception as exc:
        raise HTTPException(500, f"Fehler beim Lesen der Backup-Historie: {exc}")


@app.post("/backup/create", tags=["Backup"])
def backup_create(payload: dict = {}):
    """Erstellt einen neuen App-Backup (läuft synchron, dauert ~5s)."""
    import subprocess as _sp
    import sys as _sys
    note = (payload or {}).get("note", "")
    args = [_sys.executable, str(_BACKUP_SCRIPT)]
    if note:
        args += ["--note", note]
    try:
        result = _sp.run(
            args,
            capture_output=True, text=True, timeout=120,
            cwd=str(_BACKUP_SCRIPT.parent),
        )
        if result.returncode != 0:
            raise HTTPException(500, f"Backup fehlgeschlagen: {result.stderr or result.stdout}")
        # Neu gelesene Version zurückgeben
        if _BACKUP_VERSION_FILE.exists():
            import json as _j
            data = _j.loads(_BACKUP_VERSION_FILE.read_text(encoding="utf-8"))
            last = data["backups"][-1] if data["backups"] else {}
            return {"status": "ok", "backup": last, "output": result.stdout}
        return {"status": "ok", "output": result.stdout}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(500, str(exc))


@app.delete("/backup/{filename}", tags=["Backup"])
def backup_delete(filename: str):
    """Löscht eine Backup-ZIP-Datei und den Eintrag aus der Historie."""
    import json as _j
    # Sicherheitsprüfung: nur .zip, kein Pfad-Traversal
    if "/" in filename or "\\" in filename or not filename.endswith(".zip"):
        raise HTTPException(400, "Ungültiger Dateiname")
    f = _BACKUP_DIR / filename
    if f.exists():
        f.unlink()
    # Aus JSON entfernen
    if _BACKUP_VERSION_FILE.exists():
        data = _j.loads(_BACKUP_VERSION_FILE.read_text(encoding="utf-8"))
        data["backups"] = [b for b in data["backups"] if b["filename"] != filename]
        _BACKUP_VERSION_FILE.write_text(_j.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    return {"status": "ok"}


@app.post("/db-config/migrate", tags=["DB-Konfiguration"])
def migrate_db(payload: dict):
    """Kopiert alle Daten der aktuellen DB in eine neue Zieldatei (SQLite backup API)."""
    import sqlite3 as _sq
    target = payload.get("target_path", "").strip()
    if not target:
        raise HTTPException(400, "target_path fehlt")
    tp = pathlib.Path(target)
    if not tp.parent.exists():
        raise HTTPException(400, f"Zielverzeichnis nicht gefunden: {tp.parent}")
    if tp.resolve() == pathlib.Path(_DB_PATH).resolve():
        raise HTTPException(400, "Quelle und Ziel sind identisch")
    try:
        src = _sq.connect(_DB_PATH, timeout=10)
        dst = _sq.connect(str(tp), timeout=10)
        src.backup(dst, pages=200)   # pages=-1 = alles auf einmal; 200 = schonend
        src.close()
        dst.close()
        size = tp.stat().st_size
        log.info("DB-Migration abgeschlossen: %s → %s (%d bytes)", _DB_PATH, target, size)
        return {
            "status":           "ok",
            "message":          f"Migration erfolgreich – {size:,} Bytes kopiert",
            "source_path":      _DB_PATH,
            "target_path":      str(tp),
            "target_size_bytes": size,
        }
    except Exception as exc:
        raise HTTPException(500, f"Migration fehlgeschlagen: {exc}")

# ── Rückstellungs-Request-Endpoints ──────────────────────────────────────────
@app.get("/rueckstellung/requests", response_model=List[RueckstellungOut], tags=["Rückstellung"])
def list_rueckstellungen(status: str = "", db: Session = Depends(get_db)):
    q = db.query(RueckstellungRequest)
    if status:
        q = q.filter(RueckstellungRequest.status == status)
    return q.order_by(RueckstellungRequest.created_at.desc()).all()


@app.post("/rueckstellung/requests", response_model=RueckstellungOut, status_code=201, tags=["Rückstellung"])
def create_rueckstellung(payload: RueckstellungIn, db: Session = Depends(get_db)):
    requires = int(payload.workflow == "indirekt" and payload.betrag > 50000)
    status   = "ausstehend" if requires else "offen"
    req = RueckstellungRequest(**payload.model_dump(), requires_approval=requires, status=status)
    db.add(req)
    db.commit()
    db.refresh(req)
    db.add(RueckstellungApproval(request_id=req.id, action="submitted",
                                  actor=payload.submitted_by, comment="Antrag eingereicht"))
    db.commit()
    db.refresh(req)
    return req


@app.get("/rueckstellung/requests/{request_id}", response_model=RueckstellungOut, tags=["Rückstellung"])
def get_rueckstellung(request_id: int, db: Session = Depends(get_db)):
    req = db.get(RueckstellungRequest, request_id)
    if not req:
        raise HTTPException(404, "Antrag nicht gefunden")
    return req


@app.patch("/rueckstellung/requests/{request_id}/approve", response_model=RueckstellungOut, tags=["Rückstellung"])
def approve_rueckstellung(request_id: int, payload: dict, db: Session = Depends(get_db)):
    req = db.get(RueckstellungRequest, request_id)
    if not req:
        raise HTTPException(404, "Antrag nicht gefunden")
    if req.status not in ("ausstehend", "offen"):
        raise HTTPException(400, f"Antrag kann nicht genehmigt werden (Status: {req.status})")
    actor   = payload.get("actor", "")
    comment = payload.get("comment", "")
    req.status = "genehmigt"; req.approved_by = actor
    req.approval_comment = comment; req.updated_at = datetime.utcnow()
    db.add(RueckstellungApproval(request_id=req.id, action="approved", actor=actor, comment=comment))
    db.commit(); db.refresh(req)
    return req


@app.patch("/rueckstellung/requests/{request_id}/reject", response_model=RueckstellungOut, tags=["Rückstellung"])
def reject_rueckstellung(request_id: int, payload: dict, db: Session = Depends(get_db)):
    req = db.get(RueckstellungRequest, request_id)
    if not req:
        raise HTTPException(404, "Antrag nicht gefunden")
    if req.status not in ("ausstehend", "offen", "genehmigt"):
        raise HTTPException(400, f"Antrag kann nicht abgelehnt werden (Status: {req.status})")
    actor = payload.get("actor", ""); comment = payload.get("comment", "")
    req.status = "abgelehnt"; req.updated_at = datetime.utcnow()
    db.add(RueckstellungApproval(request_id=req.id, action="rejected", actor=actor, comment=comment))
    db.commit(); db.refresh(req)
    return req


@app.patch("/rueckstellung/requests/{request_id}/booked", response_model=RueckstellungOut, tags=["Rückstellung"])
def mark_rueckstellung_booked(request_id: int, payload: dict, db: Session = Depends(get_db)):
    req = db.get(RueckstellungRequest, request_id)
    if not req:
        raise HTTPException(404, "Antrag nicht gefunden")
    doc_nrs = payload.get("sap_doc_nrs", ""); actor = payload.get("actor", "system")
    req.status = "gebucht"; req.sap_doc_nrs = doc_nrs; req.updated_at = datetime.utcnow()
    db.add(RueckstellungApproval(request_id=req.id, action="booked",
                                  actor=actor, comment=f"SAP-Belege: {doc_nrs}"))
    db.commit(); db.refresh(req)
    return req


@app.patch("/rueckstellung/requests/{request_id}/cancel", response_model=RueckstellungOut, tags=["Rückstellung"])
def cancel_rueckstellung(request_id: int, payload: dict, db: Session = Depends(get_db)):
    req = db.get(RueckstellungRequest, request_id)
    if not req:
        raise HTTPException(404, "Antrag nicht gefunden")
    if req.status == "gebucht":
        raise HTTPException(400, "Gebuchte Anträge können nicht storniert werden")
    actor = payload.get("actor", ""); req.status = "storniert"; req.updated_at = datetime.utcnow()
    db.add(RueckstellungApproval(request_id=req.id, action="cancelled",
                                  actor=actor, comment=payload.get("comment", "")))
    db.commit(); db.refresh(req)
    return req


@app.delete("/rueckstellung/requests/{request_id}", status_code=204, tags=["Rückstellung"])
def delete_rueckstellung(request_id: int, db: Session = Depends(get_db)):
    req = db.get(RueckstellungRequest, request_id)
    if not req:
        raise HTTPException(404, "Antrag nicht gefunden")
    if req.status == "gebucht":
        raise HTTPException(400, "Gebuchte Anträge können nicht gelöscht werden")
    db.delete(req); db.commit()


# ── Kontenplan-Endpoints ─────────────────────────────────────────────────────

@app.get("/rueckstellung/kontenplan", response_model=List[KontenplanOut], tags=["Rückstellung"])
def list_kontenplan(db: Session = Depends(get_db)):
    return db.query(RueckstellungKontenplan).order_by(
        RueckstellungKontenplan.sort_order, RueckstellungKontenplan.id).all()


@app.post("/rueckstellung/kontenplan", response_model=KontenplanOut, status_code=201, tags=["Rückstellung"])
def create_kontenplan(payload: KontenplanIn, db: Session = Depends(get_db)):
    entry = RueckstellungKontenplan(**payload.model_dump())
    db.add(entry); db.commit(); db.refresh(entry)
    return entry


@app.put("/rueckstellung/kontenplan/{entry_id}", response_model=KontenplanOut, tags=["Rückstellung"])
def update_kontenplan(entry_id: int, payload: KontenplanIn, db: Session = Depends(get_db)):
    entry = db.get(RueckstellungKontenplan, entry_id)
    if not entry:
        raise HTTPException(404, "Eintrag nicht gefunden")
    for k, v in payload.model_dump().items():
        setattr(entry, k, v)
    db.commit(); db.refresh(entry)
    return entry


@app.delete("/rueckstellung/kontenplan/{entry_id}", status_code=204, tags=["Rückstellung"])
def delete_kontenplan(entry_id: int, db: Session = Depends(get_db)):
    entry = db.get(RueckstellungKontenplan, entry_id)
    if not entry:
        raise HTTPException(404, "Eintrag nicht gefunden")
    db.delete(entry); db.commit()


# ── Template-Endpoints ────────────────────────────────────────────────────────

@app.get("/rueckstellung/templates", response_model=List[RueckstellungTemplateOut], tags=["Rückstellung"])
def list_templates(bukrs: str = "", kategorie: str = "", db: Session = Depends(get_db)):
    q = db.query(RueckstellungTemplate)
    if bukrs:
        q = q.filter(RueckstellungTemplate.bukrs == bukrs)
    if kategorie:
        q = q.filter(RueckstellungTemplate.kategorie == kategorie)
    return q.order_by(RueckstellungTemplate.id).all()


@app.post("/rueckstellung/templates", response_model=RueckstellungTemplateOut, status_code=201, tags=["Rückstellung"])
def create_template(payload: RueckstellungTemplateIn, db: Session = Depends(get_db)):
    now = datetime.utcnow()
    tpl = RueckstellungTemplate(**payload.model_dump(), created_at=now, updated_at=now)
    db.add(tpl); db.commit(); db.refresh(tpl)
    return tpl


@app.put("/rueckstellung/templates/{tpl_id}", response_model=RueckstellungTemplateOut, tags=["Rückstellung"])
def update_template(tpl_id: int, payload: RueckstellungTemplateIn, db: Session = Depends(get_db)):
    tpl = db.get(RueckstellungTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "Vorlage nicht gefunden")
    for k, v in payload.model_dump().items():
        setattr(tpl, k, v)
    tpl.updated_at = datetime.utcnow()
    db.commit(); db.refresh(tpl)
    return tpl


@app.delete("/rueckstellung/templates/{tpl_id}", status_code=204, tags=["Rückstellung"])
def delete_template(tpl_id: int, db: Session = Depends(get_db)):
    tpl = db.get(RueckstellungTemplate, tpl_id)
    if not tpl:
        raise HTTPException(404, "Vorlage nicht gefunden")
    db.delete(tpl); db.commit()


# ── AfA-Lauf Endpoints ────────────────────────────────────────────────────────
@app.get("/afa/runs", response_model=List[AfaRunOut], tags=["AfA-Lauf"])
def list_afa_runs(sap_system: str = "", status: str = "",
                  db: Session = Depends(get_db)):
    q = db.query(AfaRun)
    if sap_system:
        q = q.filter(AfaRun.sap_system == sap_system)
    if status:
        q = q.filter(AfaRun.status == status)
    return q.order_by(AfaRun.fire_at.desc()).limit(200).all()


@app.post("/afa/runs", response_model=AfaRunOut, status_code=201, tags=["AfA-Lauf"])
def create_afa_run(payload: AfaRunIn, db: Session = Depends(get_db)):
    from datetime import timezone
    # fire_at als UTC parsen (naive oder mit TZ)
    try:
        fire_at = datetime.fromisoformat(payload.fire_at.replace("Z", "+00:00"))
        if fire_at.tzinfo is not None:
            fire_at = fire_at.astimezone(timezone.utc).replace(tzinfo=None)
    except ValueError:
        raise HTTPException(400, f"Ungültiges fire_at-Format: {payload.fire_at!r}")
    ar = AfaRun(
        fire_at      = fire_at,
        bukrs        = payload.bukrs.strip(),
        fiscal_year  = payload.fiscal_year.strip(),
        period       = payload.period.strip().zfill(2),
        reason       = payload.reason,
        testrun      = 1 if payload.testrun else 0,
        stop_on_warn = 1 if payload.stop_on_warn else 0,
        posting_date = payload.posting_date.strip(),
        sap_system   = payload.sap_system,
        created_by   = payload.created_by,
        status       = "pending",
    )
    db.add(ar); db.commit(); db.refresh(ar)
    # Im Scheduler eintragen (sofort)
    if fire_at > datetime.utcnow():
        scheduler.add_job(trigger_afa_run,
                          trigger=DateTrigger(run_date=fire_at),
                          args=[ar.id], id=f"afa-run-{ar.id}",
                          replace_existing=True)
        log.info("AfA-Lauf %d für %s eingeplant.", ar.id, fire_at)
    return ar


@app.get("/afa/runs/{run_id}", response_model=AfaRunOut, tags=["AfA-Lauf"])
def get_afa_run(run_id: int, db: Session = Depends(get_db)):
    ar = db.get(AfaRun, run_id)
    if not ar:
        raise HTTPException(404, "AfA-Lauf nicht gefunden")
    return ar


@app.patch("/afa/runs/{run_id}/cancel", response_model=AfaRunOut, tags=["AfA-Lauf"])
def cancel_afa_run(run_id: int, db: Session = Depends(get_db)):
    ar = db.get(AfaRun, run_id)
    if not ar:
        raise HTTPException(404, "AfA-Lauf nicht gefunden")
    if ar.status not in ("pending",):
        raise HTTPException(409, f"Nur pending-Laeufe koennen storniert werden (aktuell: {ar.status})")
    ar.status = "cancelled"
    db.commit(); db.refresh(ar)
    try:
        scheduler.remove_job(f"afa-run-{run_id}")
    except Exception:
        pass
    return ar


@app.delete("/afa/runs/{run_id}", status_code=204, tags=["AfA-Lauf"])
def delete_afa_run(run_id: int, db: Session = Depends(get_db)):
    ar = db.get(AfaRun, run_id)
    if not ar:
        raise HTTPException(404, "AfA-Lauf nicht gefunden")
    if ar.status == "running":
        raise HTTPException(409, "Laufender AfA-Job kann nicht geloescht werden")
    try:
        scheduler.remove_job(f"afa-run-{run_id}")
    except Exception:
        pass
    db.delete(ar); db.commit()


@app.post("/afa/runs/{run_id}/execute", response_model=AfaRunOut, tags=["AfA-Lauf"])
def execute_afa_run_now(run_id: int, db: Session = Depends(get_db)):
    """Fuehrt einen pending AfA-Lauf sofort aus (unabhaengig vom fire_at)."""
    ar = db.get(AfaRun, run_id)
    if not ar:
        raise HTTPException(404, "AfA-Lauf nicht gefunden")
    if ar.status != "pending":
        raise HTTPException(409, f"Nur pending-Laeufe koennen sofort ausgefuehrt werden (aktuell: {ar.status})")
    scheduler.add_job(trigger_afa_run, args=[run_id],
                      id=f"afa-run-{run_id}-now", replace_existing=True)
    return ar


# ── Buchungskreis-Registry Endpoint ───────────────────────────────────────────
@app.get("/bukrs", tags=["System"])
def get_bukrs_list():
    """Gibt alle konfigurierten Buchungskreise (SEP & SEQ) zurück."""
    return BUKRS_LIST



# ── AP-Aging / FBL1N-Import ───────────────────────────────────────────────────

class ApAgingRun(Base):
    """Gespeicherter AP-Aging-Import aus SAP (mit vollständigen Einzelposten)."""
    __tablename__ = "ap_aging_runs"
    id:           Mapped[int]      = mapped_column(primary_key=True)
    stichtag:     Mapped[str]      = mapped_column(String(10))   # 'YYYY-MM-DD'
    stichtag_fmt: Mapped[str]      = mapped_column(String(10), default="")  # 'DD.MM.YYYY'
    bukrs_list:   Mapped[str]      = mapped_column(String(200))  # 'VV9,0334,...'
    normal_items: Mapped[bool]     = mapped_column(Boolean, default=True)
    special_gl:   Mapped[bool]     = mapped_column(Boolean, default=True)
    months_back:  Mapped[int]      = mapped_column(Integer, default=36)
    item_count:   Mapped[int]      = mapped_column(Integer, default=0)
    sap_system:   Mapped[str]      = mapped_column(String(16), default="")
    created_by:   Mapped[str]      = mapped_column(String(128), default="")
    created_at:   Mapped[datetime] = mapped_column(DateTime, default=datetime.utcnow)
    items_json:   Mapped[str]      = mapped_column(Text, default="[]")  # JSON-Array


# Migration: Tabelle anlegen falls nicht vorhanden
with engine.connect() as _conn:
    _conn.execute(text(
        "CREATE TABLE IF NOT EXISTS ap_aging_runs ("
        "id           INTEGER PRIMARY KEY AUTOINCREMENT,"
        "stichtag     TEXT    NOT NULL,"
        "stichtag_fmt TEXT    NOT NULL DEFAULT '',"
        "bukrs_list   TEXT    NOT NULL DEFAULT '',"
        "normal_items INTEGER NOT NULL DEFAULT 1,"
        "special_gl   INTEGER NOT NULL DEFAULT 1,"
        "months_back  INTEGER NOT NULL DEFAULT 36,"
        "item_count   INTEGER NOT NULL DEFAULT 0,"
        "sap_system   TEXT    NOT NULL DEFAULT '',"
        "created_by   TEXT    NOT NULL DEFAULT '',"
        "created_at   TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,"
        "items_json   TEXT    NOT NULL DEFAULT '[]'"
        ")"
    ))
    _conn.commit()


class ApAgingImportIn(BaseModel):
    bukrs:        str  = ""
    key_date:     str  = ""
    normal_items: bool = True
    special_gl:   bool = True
    months_back:  int  = 36


class ApAgingSaveIn(BaseModel):
    stichtag:     str
    stichtag_fmt: str  = ""
    bukrs_list:   str  = ""
    normal_items: bool = True
    special_gl:   bool = True
    months_back:  int  = 36
    sap_system:   str  = ""
    created_by:   str  = ""
    items:        list = []


@app.post("/ap-aging/import", tags=["Reporting"])
def ap_aging_import(body: ApAgingImportIn):
    payload = {
        "bukrs":        body.bukrs,
        "key_date":     body.key_date,
        "normal_items": body.normal_items,
        "special_gl":   body.special_gl,
        "months_back":  body.months_back,
    }
    result = _bridge_post("/ap-aging/fbl1n-import", payload, timeout=300)
    if result.get("status") == "error":
        raise HTTPException(502, f"SAP-Import fehlgeschlagen: {result.get('message','?')}")
    return result


@app.post("/ap-aging/imports", tags=["Reporting"])
def ap_aging_save(body: ApAgingSaveIn):
    import json as _json
    with SessionLocal() as db:
        run = ApAgingRun(
            stichtag     = body.stichtag,
            stichtag_fmt = body.stichtag_fmt or body.stichtag,
            bukrs_list   = body.bukrs_list,
            normal_items = body.normal_items,
            special_gl   = body.special_gl,
            months_back  = body.months_back,
            item_count   = len(body.items),
            sap_system   = body.sap_system,
            created_by   = body.created_by,
            items_json   = _json.dumps(body.items, ensure_ascii=False),
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        return {"id": run.id, "item_count": run.item_count, "created_at": str(run.created_at)}


@app.get("/ap-aging/imports", tags=["Reporting"])
def ap_aging_list():
    with SessionLocal() as db:
        runs = db.execute(
            text("SELECT id, stichtag, stichtag_fmt, bukrs_list, normal_items, special_gl, "
                 "months_back, item_count, sap_system, created_by, created_at "
                 "FROM ap_aging_runs ORDER BY created_at DESC")
        ).fetchall()
        return [dict(r._mapping) for r in runs]


@app.get("/ap-aging/imports/{run_id}", tags=["Reporting"])
def ap_aging_get(run_id: int):
    import json as _json
    with SessionLocal() as db:
        row = db.execute(
            text("SELECT * FROM ap_aging_runs WHERE id = :id"),
            {"id": run_id}
        ).fetchone()
        if not row:
            raise HTTPException(404, "Import nicht gefunden")
        d = dict(row._mapping)
        d["items"] = _json.loads(d.pop("items_json", "[]"))
        return d


@app.delete("/ap-aging/imports/{run_id}", tags=["Reporting"])
def ap_aging_delete(run_id: int):
    with SessionLocal() as db:
        result = db.execute(
            text("DELETE FROM ap_aging_runs WHERE id = :id"), {"id": run_id}
        )
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(404, "Import nicht gefunden")
        return {"deleted": run_id}


@app.post("/ap-aging/export-excel", tags=["Reporting"])
def ap_aging_export_excel(payload: dict):
    """Erzeugt eine farbige Excel-Datei (openpyxl) mit AP-Aging-Daten."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    raw_data  = payload.get("rawData",  [])
    gl_data   = payload.get("glData",   [])
    sup_data  = payload.get("supData",  [])
    stichtag  = payload.get("stichtag", "30.06.2026")
    bukrs_val = payload.get("bukrs",    "(Mehrere Elemente)")
    dt_str    = stichtag.replace(".", "")  # "30062026"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # leeres Standard-Sheet entfernen

    # ── Stilhilfen ──────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    TITLE_FONT = Font(bold=True, size=11)
    TOTAL_FONT = Font(bold=True, size=10)
    TOTAL_FILL = PatternFill("solid", fgColor="D9D9D9")
    CAT_FILL  = PatternFill("solid", fgColor="DBEAFE")
    CAT_FONT  = Font(color="1D4ED8", size=10)
    SCH_FILL  = PatternFill("solid", fgColor="FFF7ED")
    SCH_FONT  = Font(color="C2410C", size=10)
    DEF_FONT  = Font(size=10)
    THIN = Side(style="thin", color="B0B0B0")
    CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    NUM_FMT = '#,##0.00'

    def _set_hdr(ws, row, cols):
        for ci, val in enumerate(cols, 1):
            c = ws.cell(row=row, column=ci, value=val)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = CELL_BORDER

    def _set_num(c, val):
        c.number_format = NUM_FMT
        c.alignment = Alignment(horizontal="right")
        c.border = CELL_BORDER
        try: c.value = float(val) if val not in (None, "") else 0.0
        except: c.value = 0.0

    def _meta_rows(ws, title):
        ws.append([None, title])
        ws.cell(ws.max_row, 2).font = TITLE_FONT
        ws.append([None, "Catensys Group"])
        ws.append([None, "Company Code", bukrs_val])
        ws.append([None, "remark", "(Leer)", "(ohne Catensys-Intern und Schaeffler)"])
        ws.append([])
        ws.append([None, None, None, "Werte"])

    # ── Sheet 1: GL-Abstimmung ───────────────────────────────────────────────
    ws1 = wb.create_sheet("Aging_AP_GL-Abstimmung")
    _meta_rows(ws1, f"Aging -AP-GL Abstimmung  per {stichtag}")
    hdr1 = ["G/L Account", "Local Currency 2", "Fällig <30 days",
            "Fällig >=30 days", "Fällig >60 days", "Fällig >90 days"]
    HDR_ROW1 = ws1.max_row + 1
    _set_hdr(ws1, HDR_ROW1, hdr1)
    totals1 = [0.0, 0.0, 0.0, 0.0]
    for r in gl_data:
        dr = ws1.max_row + 1
        ws1.cell(dr, 1, r.get("gl","")).border = CELL_BORDER
        ws1.cell(dr, 2, r.get("cur","EUR")).border = CELL_BORDER
        for ci, k in enumerate(["d30","d60","d90","d90p"], 3):
            _set_num(ws1.cell(dr, ci), r.get(k, 0))
            totals1[ci-3] += float(r.get(k, 0) or 0)
    tr = ws1.max_row + 1
    ws1.cell(tr, 1, "Gesamtergebnis").font = TOTAL_FONT
    for ci, v in enumerate(totals1, 3):
        c = ws1.cell(tr, ci, v); c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        c.number_format = NUM_FMT; c.alignment = Alignment(horizontal="right")
    for col in range(1, 7): ws1.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 2: Kreditoren Aging ────────────────────────────────────────────
    sh2_name = f"Aging_AP_{dt_str[:6]}"
    ws2 = wb.create_sheet(sh2_name)
    _meta_rows(ws2, f"Aging -Kreditoren per {stichtag}")
    hdr2 = ["Supplier name", "Local Currency 2", "Fällig <30 days",
            "Fällig >=30 days", "Fällig >60 days", "Fällig >90 days"]
    HDR_ROW2 = ws2.max_row + 1
    _set_hdr(ws2, HDR_ROW2, hdr2)
    totals2 = [0.0, 0.0, 0.0, 0.0]
    for r in sup_data:
        dr = ws2.max_row + 1
        ws2.cell(dr, 1, r.get("sup","")).border = CELL_BORDER
        ws2.cell(dr, 2, r.get("cur","EUR")).border = CELL_BORDER
        for ci, k in enumerate(["d30","d60","d90","d90p"], 3):
            _set_num(ws2.cell(dr, ci), r.get(k, 0))
            totals2[ci-3] += float(r.get(k, 0) or 0)
    tr2 = ws2.max_row + 1
    ws2.cell(tr2, 1, "Gesamtergebnis").font = TOTAL_FONT
    for ci, v in enumerate(totals2, 3):
        c = ws2.cell(tr2, ci, v); c.fill = TOTAL_FILL; c.font = TOTAL_FONT
        c.number_format = NUM_FMT; c.alignment = Alignment(horizontal="right")
    ws2.column_dimensions["A"].width = 40
    for col in range(2, 7): ws2.column_dimensions[get_column_letter(col)].width = 20

    # ── Sheet 3: AP-Kreditoren Einzelposten ─────────────────────────────────
    sh3_name = f"AP-Kreditoren-{dt_str[:6]}"
    ws3 = wb.create_sheet(sh3_name)
    HDR3 = ["BuKr","G/L","Konto","Lieferant","Belegnr.","Art",
            "Belegdatum","Fällig Datum","Tage",
            "<30 Tage",">=30 Tage",">60 Tage",">90 Tage","Bemerkung"]
    _set_hdr(ws3, 1, HDR3)
    BEM_COL = 14  # Spalte N = "Bemerkung"
    for r in raw_data:
        dr = ws3.max_row + 1
        bem = r.get("bemerkung") or ""
        is_cat = "catensys" in bem.lower()
        is_sch = "schaeffler" in bem.lower()
        vals = [
            r.get("bukrs",""), r.get("gl",""), r.get("account",""), r.get("sup",""),
            r.get("docnum",""), r.get("dtype",""), r.get("docdate",""), r.get("duedate",""),
            r.get("days",""), r.get("d30",0), r.get("d60",0), r.get("d90",0), r.get("d90p",0),
            bem
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws3.cell(dr, ci, val)
            cell.border = CELL_BORDER
            cell.font = DEF_FONT
            if ci in (10, 11, 12, 13):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            if ci == BEM_COL:
                if is_cat:
                    cell.fill = CAT_FILL; cell.font = CAT_FONT
                    cell.alignment = Alignment(horizontal="center")
                elif is_sch:
                    cell.fill = SCH_FILL; cell.font = SCH_FONT
                    cell.alignment = Alignment(horizontal="center")
        if is_cat or is_sch:
            row_fill = CAT_FILL if is_cat else SCH_FILL
            for ci in range(1, BEM_COL):
                ws3.cell(dr, ci).fill = row_fill
    col_widths3 = [6,8,14,40,14,6,12,12,7,12,12,12,12,18]
    for ci, w in enumerate(col_widths3, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w
    ws3.freeze_panes = "A2"

    # ── Datei in Speicher schreiben ──────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"Aging_AP_{stichtag.replace('.', '-')}_YCONN.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'}
    )


# ─────────────────────────────────────────────────────────────────────────────
# GLOBALE BUCHUNGSHISTORIE
# ─────────────────────────────────────────────────────────────────────────────

class GlobalBookingEntry(Base):
    __tablename__ = "global_booking_history"
    id          : Mapped[int]      = mapped_column(primary_key=True)
    module      : Mapped[str]      = mapped_column(String(32),  default="")
    bukrs       : Mapped[str]      = mapped_column(String(8),   default="")
    sap_system  : Mapped[str]      = mapped_column(String(16),  default="")
    doc_date    : Mapped[str]      = mapped_column(String(10),  default="")
    yr          : Mapped[str]      = mapped_column(String(4),   default="")
    period      : Mapped[str]      = mapped_column(String(2),   default="")
    belnr       : Mapped[str]      = mapped_column(String(18),  default="")
    obj_key     : Mapped[str]      = mapped_column(String(20),  default="")
    ref_doc_no  : Mapped[str]      = mapped_column(String(64),  default="")
    header_txt  : Mapped[str]      = mapped_column(String(128), default="")
    amount      : Mapped[float]    = mapped_column(Float,       default=0.0)
    currency    : Mapped[str]      = mapped_column(String(4),   default="EUR")
    description : Mapped[str]      = mapped_column(String(256), default="")
    booked_by   : Mapped[str]      = mapped_column(String(128), default="")
    booked_at   : Mapped[str]      = mapped_column(String(32),  default="")
    status      : Mapped[str]      = mapped_column(String(16),  default="booked")
    source      : Mapped[str]      = mapped_column(String(16),  default="app")

# Migration
with engine.connect() as _conn:
    _conn.execute(text(
        "CREATE TABLE IF NOT EXISTS global_booking_history ("
        "id          INTEGER PRIMARY KEY AUTOINCREMENT,"
        "module      TEXT NOT NULL DEFAULT '',"
        "bukrs       TEXT NOT NULL DEFAULT '',"
        "sap_system  TEXT NOT NULL DEFAULT '',"
        "doc_date    TEXT NOT NULL DEFAULT '',"
        "yr          TEXT NOT NULL DEFAULT '',"
        "period      TEXT NOT NULL DEFAULT '',"
        "belnr       TEXT NOT NULL DEFAULT '',"
        "obj_key     TEXT NOT NULL DEFAULT '',"
        "ref_doc_no  TEXT NOT NULL DEFAULT '',"
        "header_txt  TEXT NOT NULL DEFAULT '',"
        "amount      REAL NOT NULL DEFAULT 0.0,"
        "currency    TEXT NOT NULL DEFAULT 'EUR',"
        "description TEXT NOT NULL DEFAULT '',"
        "booked_by   TEXT NOT NULL DEFAULT '',"
        "booked_at   TEXT NOT NULL DEFAULT '',"
        "status      TEXT NOT NULL DEFAULT 'booked',"
        "source      TEXT NOT NULL DEFAULT 'app'"
        ")"
    ))
    _conn.commit()

class GlobalBookingIn(BaseModel):
    module      : str   = ""
    bukrs       : str   = ""
    sap_system  : str   = ""
    doc_date    : str   = ""
    yr          : str   = ""
    period      : str   = ""
    belnr       : str   = ""
    obj_key     : str   = ""
    ref_doc_no  : str   = ""
    header_txt  : str   = ""
    amount      : float = 0.0
    currency    : str   = "EUR"
    description : str   = ""
    booked_by   : str   = ""
    booked_at   : str   = ""
    status      : str   = "booked"
    source      : str   = "app"

@app.post("/booking-history", tags=["Global"])
def booking_history_save(body: GlobalBookingIn):
    with SessionLocal() as db:
        row = GlobalBookingEntry(
            module=body.module, bukrs=body.bukrs, sap_system=body.sap_system,
            doc_date=body.doc_date, yr=body.yr, period=body.period,
            belnr=body.belnr, obj_key=body.obj_key,
            ref_doc_no=body.ref_doc_no, header_txt=body.header_txt,
            amount=body.amount, currency=body.currency,
            description=body.description, booked_by=body.booked_by,
            booked_at=body.booked_at or datetime.utcnow().isoformat(),
            status=body.status, source=body.source,
        )
        db.add(row); db.commit(); db.refresh(row)
        return {"id": row.id, "status": "saved"}


@app.get("/booking-history", tags=["Global"])
def booking_history_list(
    module: str = "", bukrs: str = "", limit: int = 500,
    db: Session = Depends(get_db)
):
    q = db.query(GlobalBookingEntry).order_by(GlobalBookingEntry.id.desc())
    if module:
        q = q.filter(GlobalBookingEntry.module == module)
    if bukrs:
        q = q.filter(GlobalBookingEntry.bukrs == bukrs)
    rows = q.limit(limit).all()
    return [
        {
            "id": r.id, "module": r.module, "bukrs": r.bukrs,
            "sap_system": r.sap_system, "doc_date": r.doc_date,
            "yr": r.yr, "period": r.period, "belnr": r.belnr,
            "obj_key": r.obj_key, "ref_doc_no": r.ref_doc_no,
            "header_txt": r.header_txt, "amount": r.amount,
            "currency": r.currency, "description": r.description,
            "booked_by": r.booked_by, "booked_at": r.booked_at,
            "status": r.status, "source": r.source,
        }
        for r in rows
    ]
# ═══════════════════════════════════════════════════════════════════════════════
# USt-Firmenwagen
# ═══════════════════════════════════════════════════════════════════════════════

class UstFirmenwagenRun(Base):
    __tablename__ = "ust_firmenwagen_runs"
    id           : Mapped[int]  = mapped_column(Integer, primary_key=True, autoincrement=True)
    periode      : Mapped[str]  = mapped_column(String(20),  default="")
    post_date    : Mapped[str]  = mapped_column(String(10),  default="")
    bukrs        : Mapped[str]  = mapped_column(String(8),   default="0435")
    fis_period   : Mapped[str]  = mapped_column(String(4),   default="")
    fisc_year    : Mapped[str]  = mapped_column(String(4),   default="")
    ref_doc      : Mapped[str]  = mapped_column(String(64),  default="")
    header_txt   : Mapped[str]  = mapped_column(String(128), default="")
    sap_system   : Mapped[str]  = mapped_column(String(16),  default="SEP")
    total_amount : Mapped[float]= mapped_column(Float,       default=0.0)
    row_count    : Mapped[int]  = mapped_column(Integer,     default=0)
    status       : Mapped[str]  = mapped_column(String(16),  default="saved")
    created_by   : Mapped[str]  = mapped_column(String(128), default="")
    created_at   : Mapped[str]  = mapped_column(String(32),  default="")
    lines_json   : Mapped[str]  = mapped_column(Text,        default="[]")

# Migration
with engine.connect() as _conn:
    _conn.execute(text(
        "CREATE TABLE IF NOT EXISTS ust_firmenwagen_runs ("
        "id           INTEGER PRIMARY KEY AUTOINCREMENT,"
        "periode      TEXT NOT NULL DEFAULT '',"
        "post_date    TEXT NOT NULL DEFAULT '',"
        "bukrs        TEXT NOT NULL DEFAULT '0435',"
        "fis_period   TEXT NOT NULL DEFAULT '',"
        "fisc_year    TEXT NOT NULL DEFAULT '',"
        "ref_doc      TEXT NOT NULL DEFAULT '',"
        "header_txt   TEXT NOT NULL DEFAULT '',"
        "sap_system   TEXT NOT NULL DEFAULT 'SEP',"
        "total_amount REAL NOT NULL DEFAULT 0.0,"
        "row_count    INTEGER NOT NULL DEFAULT 0,"
        "status       TEXT NOT NULL DEFAULT 'saved',"
        "created_by   TEXT NOT NULL DEFAULT '',"
        "created_at   TEXT NOT NULL DEFAULT '',"
        "lines_json   TEXT NOT NULL DEFAULT '[]'"
        ")"
    ))
    _conn.commit()

class UwLineIn(BaseModel):
    persnr       : str   = ""
    nachname     : str   = ""
    kostenstelle : str   = ""
    gvw          : float = 0.0
    fahrten      : float = 0.0
    amount       : float = 0.0
    status       : str   = "pending"
    belnr        : str   = ""

class UwRunIn(BaseModel):
    periode    : str          = ""
    post_date  : str          = ""
    bukrs      : str          = "0435"
    fis_period : str          = ""
    fisc_year  : str          = ""
    ref_doc    : str          = ""
    header_txt : str          = ""
    sap_system : str          = "SEP"
    created_by : str          = ""
    lines      : list[UwLineIn] = []

def _uw_status(lines: list) -> str:
    if not lines:
        return "saved"
    booked = sum(1 for l in lines if l.get("status") == "booked")
    if booked == len(lines):
        return "booked"
    if booked > 0:
        return "partial"
    return "saved"

@app.post("/ust-firmenwagen", tags=["USt-Firmenwagen"])
def uw_create(body: UwRunIn):
    import json as _json
    lines_data = [l.model_dump() for l in body.lines]
    total = round(sum(l["amount"] for l in lines_data), 2)
    with SessionLocal() as db:
        run = UstFirmenwagenRun(
            periode=body.periode, post_date=body.post_date, bukrs=body.bukrs,
            fis_period=body.fis_period, fisc_year=body.fisc_year,
            ref_doc=body.ref_doc, header_txt=body.header_txt,
            sap_system=body.sap_system, total_amount=total,
            row_count=len(lines_data), status=_uw_status(lines_data),
            created_by=body.created_by,
            created_at=datetime.utcnow().isoformat(),
            lines_json=_json.dumps(lines_data, ensure_ascii=False),
        )
        db.add(run)
        db.commit()
        db.refresh(run)
        return {"id": run.id, "total_amount": run.total_amount, "row_count": run.row_count}

@app.patch("/ust-firmenwagen/{run_id}", tags=["USt-Firmenwagen"])
def uw_update(run_id: int, body: UwRunIn):
    import json as _json
    lines_data = [l.model_dump() for l in body.lines]
    total = round(sum(l["amount"] for l in lines_data), 2)
    with SessionLocal() as db:
        run = db.get(UstFirmenwagenRun, run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        run.periode      = body.periode or run.periode
        run.post_date    = body.post_date or run.post_date
        run.bukrs        = body.bukrs or run.bukrs
        run.fis_period   = body.fis_period or run.fis_period
        run.fisc_year    = body.fisc_year or run.fisc_year
        run.ref_doc      = body.ref_doc or run.ref_doc
        run.header_txt   = body.header_txt or run.header_txt
        run.sap_system   = body.sap_system or run.sap_system
        run.total_amount = total
        run.row_count    = len(lines_data)
        run.status       = _uw_status(lines_data)
        run.lines_json   = _json.dumps(lines_data, ensure_ascii=False)
        db.commit()
        db.refresh(run)
        return {"id": run.id, "total_amount": run.total_amount,
                "row_count": run.row_count, "status": run.status}

@app.get("/ust-firmenwagen", tags=["USt-Firmenwagen"])
def uw_list(limit: int = 50):
    with SessionLocal() as db:
        runs = db.execute(
            text("SELECT * FROM ust_firmenwagen_runs ORDER BY id DESC LIMIT :l"),
            {"l": limit}
        ).mappings().all()
        return [dict(r) for r in runs]

@app.get("/ust-firmenwagen/{run_id}", tags=["USt-Firmenwagen"])
def uw_get(run_id: int):
    with SessionLocal() as db:
        run = db.get(UstFirmenwagenRun, run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        import json as _json
        result = {
            "id": run.id, "periode": run.periode, "post_date": run.post_date,
            "bukrs": run.bukrs, "fis_period": run.fis_period, "fisc_year": run.fisc_year,
            "ref_doc": run.ref_doc, "header_txt": run.header_txt, "sap_system": run.sap_system,
            "total_amount": run.total_amount, "row_count": run.row_count,
            "status": run.status, "created_by": run.created_by, "created_at": run.created_at,
            "lines": _json.loads(run.lines_json or "[]"),
        }
        return result

@app.delete("/ust-firmenwagen/{run_id}", tags=["USt-Firmenwagen"])
def uw_delete(run_id: int):
    with SessionLocal() as db:
        run = db.get(UstFirmenwagenRun, run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run not found")
        db.delete(run)
        db.commit()
    return Response(status_code=204)


# ═══════════════════════════════════════════════════════════════════════════════
# Management-Service Eingangsrechnungen
# ═══════════════════════════════════════════════════════════════════════════════

class MgmtInvoiceRun(Base):
    __tablename__ = "mgmt_invoice_runs"
    id          : Mapped[int] = mapped_column(Integer, primary_key=True, autoincrement=True)
    periode     : Mapped[str] = mapped_column(String(8),   default="")   # YYYY-MM
    doc_date    : Mapped[str] = mapped_column(String(10),  default="")
    pstng_date  : Mapped[str] = mapped_column(String(10),  default="")
    bukrs       : Mapped[str] = mapped_column(String(8),   default="")   # Buchungskreis
    sap_transaction : Mapped[str] = mapped_column(String(10),  default="")   # FB01 / MIRO
    ref_doc_no  : Mapped[str] = mapped_column(String(64),  default="")
    header_txt  : Mapped[str] = mapped_column(String(128), default="")
    sap_system  : Mapped[str] = mapped_column(String(16),  default="SEQ")
    vendor      : Mapped[str] = mapped_column(String(20),  default="L372961")
    tax_code    : Mapped[str] = mapped_column(String(4),   default="")
    business_place: Mapped[str] = mapped_column(String(10), default="")
    gross_amount: Mapped[float] = mapped_column(Float,     default=0.0)
    status      : Mapped[str] = mapped_column(String(20),  default="open")   # open/booked/error
    belnr       : Mapped[str] = mapped_column(String(20),  default="")       # SAP Belegnummer
    sap_msg     : Mapped[str] = mapped_column(Text,        default="")
    lines_json  : Mapped[str] = mapped_column(Text,        default="[]")
    created_by  : Mapped[str] = mapped_column(String(128), default="")
    created_at  : Mapped[str] = mapped_column(String(32),  default="")
    booked_at   : Mapped[str] = mapped_column(String(32),  default="")

with engine.connect() as _conn:
    _conn.execute(text("""
        CREATE TABLE IF NOT EXISTS mgmt_invoice_runs (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            periode       TEXT NOT NULL DEFAULT '',
            doc_date      TEXT NOT NULL DEFAULT '',
            pstng_date    TEXT NOT NULL DEFAULT '',
            bukrs         TEXT NOT NULL DEFAULT '',
            sap_transaction TEXT NOT NULL DEFAULT '',
            ref_doc_no    TEXT NOT NULL DEFAULT '',
            header_txt    TEXT NOT NULL DEFAULT '',
            sap_system    TEXT NOT NULL DEFAULT 'SEQ',
            vendor        TEXT NOT NULL DEFAULT 'L372961',
            tax_code      TEXT NOT NULL DEFAULT '',
            business_place TEXT NOT NULL DEFAULT '',
            gross_amount  REAL NOT NULL DEFAULT 0.0,
            status        TEXT NOT NULL DEFAULT 'open',
            belnr         TEXT NOT NULL DEFAULT '',
            sap_msg       TEXT NOT NULL DEFAULT '',
            lines_json    TEXT NOT NULL DEFAULT '[]',
            created_by    TEXT NOT NULL DEFAULT '',
            created_at    TEXT NOT NULL DEFAULT '',
            booked_at     TEXT NOT NULL DEFAULT ''
        )
    """))
    _conn.commit()

class MgmtInvoiceIn(BaseModel):
    periode      : str   = ""
    doc_date     : str   = ""
    pstng_date   : str   = ""
    bukrs        : str   = ""
    sap_transaction : str   = "FB01"
    ref_doc_no   : str   = ""
    header_txt   : str   = ""
    sap_system   : str   = "SEQ"
    vendor       : str   = "L372961"
    tax_code     : str   = ""
    business_place: str  = ""
    gross_amount : float = 0.0
    status       : str   = "open"
    belnr        : str   = ""
    sap_msg      : str   = ""
    lines        : list  = []
    created_by   : str   = ""

@app.get("/mgmt-invoice", tags=["Management Invoice"])
def mi_list(periode: str = "", limit: int = 100):
    with SessionLocal() as db:
        q = "SELECT * FROM mgmt_invoice_runs"
        params: dict = {}
        if periode:
            q += " WHERE periode = :p"
            params["p"] = periode
        q += " ORDER BY id DESC LIMIT :l"
        params["l"] = limit
        rows = db.execute(text(q), params).mappings().all()
        return [dict(r) for r in rows]

@app.post("/mgmt-invoice", tags=["Management Invoice"])
def mi_create(body: MgmtInvoiceIn):
    import json as _json
    now = datetime.utcnow().isoformat()
    with SessionLocal() as db:
        run = MgmtInvoiceRun(
            periode=body.periode, doc_date=body.doc_date, pstng_date=body.pstng_date,
            bukrs=body.bukrs, sap_transaction=body.sap_transaction, ref_doc_no=body.ref_doc_no,
            header_txt=body.header_txt, sap_system=body.sap_system, vendor=body.vendor,
            tax_code=body.tax_code, business_place=body.business_place,
            gross_amount=body.gross_amount, status=body.status,
            belnr=body.belnr, sap_msg=body.sap_msg,
            lines_json=_json.dumps(body.lines, ensure_ascii=False),
            created_by=body.created_by, created_at=now,
        )
        db.add(run); db.commit(); db.refresh(run)
        return {"id": run.id, "status": run.status}

@app.patch("/mgmt-invoice/{run_id}", tags=["Management Invoice"])
def mi_update(run_id: int, body: MgmtInvoiceIn):
    import json as _json
    now = datetime.utcnow().isoformat()
    with SessionLocal() as db:
        run = db.get(MgmtInvoiceRun, run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run nicht gefunden")
        for field in ("periode", "doc_date", "pstng_date", "bukrs", "sap_transaction",
                      "ref_doc_no", "header_txt", "sap_system", "vendor", "tax_code",
                      "business_place", "gross_amount", "status", "belnr", "sap_msg"):
            val = getattr(body, field, None)
            if val is not None and val != "":
                setattr(run, field, val)
        if body.lines:
            run.lines_json = _json.dumps(body.lines, ensure_ascii=False)
        if body.belnr and not run.booked_at:
            run.booked_at = now
        db.commit(); db.refresh(run)
        return {"id": run.id, "status": run.status, "belnr": run.belnr, "sap_msg": run.sap_msg}

@app.delete("/mgmt-invoice/{run_id}", tags=["Management Invoice"])
def mi_delete(run_id: int):
    with SessionLocal() as db:
        run = db.get(MgmtInvoiceRun, run_id)
        if not run:
            raise HTTPException(status_code=404, detail="Run nicht gefunden")
        db.delete(run); db.commit()
        return {"deleted": run_id}


# ══════════════════════════════════════════════════════════════════════════
# SALES REPORT – CO-PA / SD-Billing
# ══════════════════════════════════════════════════════════════════════════

class SalesRun(Base):
    __tablename__ = "sales_runs"
    id           : Mapped[int]   = mapped_column(Integer, primary_key=True, autoincrement=True)
    label        : Mapped[str]   = mapped_column(String(128), default="")
    date_from    : Mapped[str]   = mapped_column(String(10),  default="")
    date_to      : Mapped[str]   = mapped_column(String(10),  default="")
    comp_codes   : Mapped[str]   = mapped_column(String(256), default="")
    source       : Mapped[str]   = mapped_column(String(16),  default="vbrk")
    row_count    : Mapped[int]   = mapped_column(Integer,     default=0)
    total_revenue: Mapped[float] = mapped_column(Float,       default=0.0)
    currency     : Mapped[str]   = mapped_column(String(8),   default="EUR")
    sap_system   : Mapped[str]   = mapped_column(String(16),  default="")
    created_by   : Mapped[str]   = mapped_column(String(128), default="")
    created_at   : Mapped[str]   = mapped_column(String(32),  default="")
    rows_json    : Mapped[str]   = mapped_column(Text,        default="[]")
    meta_json    : Mapped[str]   = mapped_column(Text,        default="{}")


with engine.connect() as _conn:
    _conn.execute(text(
        "CREATE TABLE IF NOT EXISTS sales_runs ("
        "id            INTEGER PRIMARY KEY AUTOINCREMENT,"
        "label         TEXT NOT NULL DEFAULT '',"
        "date_from     TEXT NOT NULL DEFAULT '',"
        "date_to       TEXT NOT NULL DEFAULT '',"
        "comp_codes    TEXT NOT NULL DEFAULT '',"
        "source        TEXT NOT NULL DEFAULT 'vbrk',"
        "row_count     INTEGER NOT NULL DEFAULT 0,"
        "total_revenue REAL    NOT NULL DEFAULT 0.0,"
        "currency      TEXT NOT NULL DEFAULT 'EUR',"
        "sap_system    TEXT NOT NULL DEFAULT '',"
        "created_by    TEXT NOT NULL DEFAULT '',"
        "created_at    TEXT NOT NULL DEFAULT '',"
        "rows_json     TEXT NOT NULL DEFAULT '[]',"
        "meta_json     TEXT NOT NULL DEFAULT '{}'"
        ")"
    ))
    _conn.commit()


class SalesImportIn(BaseModel):
    comp_codes        : str         = ""
    date_from         : str         = ""
    date_to           : str         = ""
    source            : str         = "vbrk"
    operating_concern : str         = "0001"
    customer_filter   : str         = ""
    material_filter   : str         = ""
    maxrows           : int         = 5000
    sap_auth          : dict | None = None


class SalesSaveIn(BaseModel):
    label      : str  = ""
    date_from  : str  = ""
    date_to    : str  = ""
    comp_codes : str  = ""
    source     : str  = "vbrk"
    sap_system : str  = ""
    created_by : str  = ""
    currency   : str  = "EUR"
    rows       : list = []
    meta       : dict = {}


@app.post("/sales/import", tags=["Sales"])
def sales_import(body: SalesImportIn):
    """Startet SAP-Import (VBRK/VBRP oder CO-PA CE4xxxx) via Bridge."""
    payload = {
        "comp_codes":        body.comp_codes,
        "date_from":         body.date_from,
        "date_to":           body.date_to,
        "source":            body.source,
        "operating_concern": body.operating_concern,
        "customer_filter":   body.customer_filter,
        "material_filter":   body.material_filter,
        "maxrows":           body.maxrows,
        "_sap_auth":         body.sap_auth,
    }
    result = _bridge_post("/sales/copa-import", payload, timeout=300)
    if result.get("status") == "error":
        raise HTTPException(502, f"SAP-Import fehlgeschlagen: {result.get('message','?')}")
    return result


@app.post("/sales/runs", tags=["Sales"])
def sales_save(body: SalesSaveIn):
    import json as _json
    now   = datetime.utcnow().isoformat(timespec="seconds")
    total = sum(float(r.get("revenue", 0) or 0) for r in body.rows)
    with SessionLocal() as db:
        run = SalesRun(
            label         = body.label or f"Sales {body.date_from}–{body.date_to}",
            date_from     = body.date_from,
            date_to       = body.date_to,
            comp_codes    = body.comp_codes,
            source        = body.source,
            row_count     = len(body.rows),
            total_revenue = round(total, 2),
            currency      = body.currency or "EUR",
            sap_system    = body.sap_system,
            created_by    = body.created_by,
            created_at    = now,
            rows_json     = _json.dumps(body.rows, ensure_ascii=False),
            meta_json     = _json.dumps(body.meta, ensure_ascii=False),
        )
        db.add(run); db.commit(); db.refresh(run)
        return {"id": run.id, "row_count": run.row_count,
                "total_revenue": run.total_revenue, "created_at": now}


@app.get("/sales/runs", tags=["Sales"])
def sales_list():
    with SessionLocal() as db:
        rows = db.execute(text(
            "SELECT id,label,date_from,date_to,comp_codes,source,"
            "row_count,total_revenue,currency,sap_system,created_by,created_at "
            "FROM sales_runs ORDER BY created_at DESC"
        )).fetchall()
        return [dict(r._mapping) for r in rows]


@app.get("/sales/runs/{run_id}", tags=["Sales"])
def sales_get(run_id: int):
    import json as _json
    with SessionLocal() as db:
        row = db.execute(
            text("SELECT * FROM sales_runs WHERE id=:id"), {"id": run_id}
        ).fetchone()
        if not row:
            raise HTTPException(404, "Sales-Run nicht gefunden")
        d = dict(row._mapping)
        d["rows"] = _json.loads(d.pop("rows_json", "[]"))
        d["meta"] = _json.loads(d.pop("meta_json", "{}"))
        return d


@app.get("/sales/runs/{run_id}/report", tags=["Sales"])
def sales_report_filtered(
    run_id: int,
    bukrs:  str = "",
    kunnr:  str = "",
    matnr:  str = "",
    period: str = "",
):
    import json as _json
    with SessionLocal() as db:
        row = db.execute(
            text("SELECT rows_json,meta_json FROM sales_runs WHERE id=:id"),
            {"id": run_id}
        ).fetchone()
        if not row:
            raise HTTPException(404, "Sales-Run nicht gefunden")
        rows = _json.loads(row.rows_json or "[]")
        meta = _json.loads(row.meta_json or "{}")
    if bukrs:
        rows = [r for r in rows if r.get("bukrs","").startswith(bukrs)]
    if kunnr:
        rows = [r for r in rows if r.get("kunnr","").startswith(kunnr)]
    if matnr:
        rows = [r for r in rows if r.get("matnr","").startswith(matnr)]
    if period:
        rows = [r for r in rows if str(r.get("period",""))==period
                                or str(r.get("gjahr",""))==period]
    total_rev = sum(float(r.get("revenue",0) or 0) for r in rows)
    return {"rows": rows, "meta": meta,
            "filtered_count": len(rows), "filtered_total": round(total_rev,2)}


@app.delete("/sales/runs/{run_id}", tags=["Sales"])
def sales_delete(run_id: int):
    with SessionLocal() as db:
        result = db.execute(
            text("DELETE FROM sales_runs WHERE id=:id"), {"id": run_id}
        )
        db.commit()
        if result.rowcount == 0:
            raise HTTPException(404, "Sales-Run nicht gefunden")
        return {"deleted": run_id}
