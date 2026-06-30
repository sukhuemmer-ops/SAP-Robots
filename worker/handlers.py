"""
Konkrete SAP-Handler.
=====================
Pro SAP-Methode/-Transaktion ein Handler. Bekommt ``(task, payload)`` und gibt
einen Log-String zurueck. Verbindungs- und Passwortdaten kommen aus dem
Secret-Resolver (siehe ``secrets.py``).

Outputs (Excel-Reports etc.) landen unter dem Pfad aus ``REPORT_OUT_DIR``
(Default: ``./reports`` neben dem Worker). Pfad ist konfigurierbar per
Umgebungsvariable.
"""
from __future__ import annotations

import logging
import os
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Dict, Tuple

from sap_secrets import build_sap_rfc_params, build_sap_gui_params
import sap_db

log = logging.getLogger("worker.handlers")

REPORT_OUT_DIR = Path(os.getenv("REPORT_OUT_DIR", Path(__file__).resolve().parent / "reports"))
REPORT_OUT_DIR.mkdir(parents=True, exist_ok=True)


# ---------------------------------------------------------------------------
# Datums-Parser - tolerant gegen mehrere Formate und Encoding-Artefakte
# ---------------------------------------------------------------------------
def parse_date(value, default_today: bool = True):
    """
    Akzeptiert YYYYMMDD, YYYY-MM-DD, DD.MM.YYYY, DD/MM/YYYY, YYYY/MM/DD,
    'today'/'heute'/leer (-> heute), 'eom' (-> Monatsende).
    Entfernt URL-Encoding-Reste und Whitespace.
    """
    from datetime import date as _date, timedelta

    if value is None or value == "":
        if default_today:
            return datetime.now().date()
        raise ValueError("Datum fehlt.")

    s = str(value).strip()
    # Spezialwerte
    if s.lower() in ("today", "heute"):
        return datetime.now().date()
    if s.lower() in ("eom", "monatsende"):
        today = datetime.now().date()
        next_month = today.replace(day=28) + timedelta(days=4)
        return next_month - timedelta(days=next_month.day)

    # Encoding-Artefakte raeumen: %-Zeichen entfernen, dann nur Ziffern und Trenner behalten
    cleaned = s.replace("%", "")
    for fmt in ("%Y%m%d", "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(cleaned, fmt).date()
        except ValueError:
            continue

    # Letzte Notbremse: nur Ziffern uebrig lassen und als YYYYMMDD versuchen
    digits = "".join(c for c in cleaned if c.isdigit())
    if len(digits) == 8:
        try:
            return datetime.strptime(digits, "%Y%m%d").date()
        except ValueError:
            pass

    raise ValueError(
        f"Datum '{value}' nicht erkannt. Akzeptiert werden z. B. "
        "20260520, 2026-05-20, 20.05.2026, 'today' oder 'eom'."
    )


# ---------------------------------------------------------------------------
# Verbindungs-Helfer
# ---------------------------------------------------------------------------
def _rfc_connection():
    from pyrfc import Connection  # type: ignore  # pip install pyrfc
    params = build_sap_rfc_params("SAP")
    log.info("Oeffne RFC-Verbindung zu %s / Mandant %s als %s",
             params.get("ashost") or params.get("mshost"),
             params.get("client"), params.get("user"))
    return Connection(**params)


def _get_sapgui_from_rot():
    """
    Liest das SAP-GUI-Scripting-Objekt direkt aus der Windows Running Object Table (ROT).

    Hintergrund: Python's win32com.client.GetObject("SAPGUI") scheitert mit
    MK_E_SYNTAX (-2147221020), weil MkParseDisplayName() "SAPGUI" nicht als
    gueltigen OLE-Moniker erkennt. VBA's GetObject sucht hingegen direkt in der ROT.
    Dieses Verhalten bildet die Funktion nach.

    Voraussetzung: SAP GUI laeuft und Scripting ist aktiviert
    (SAP GUI Options → Accessibility & Scripting → Scripting).
    """
    import pythoncom           # type: ignore
    import win32com.client as wc  # type: ignore

    # COM muss pro Thread initialisiert werden (Flask-Worker-Threads rufen dies
    # aus einem anderen Thread auf als der Haupt-Thread).
    # CoInitialize() ist idempotent falls schon initialisiert (gibt S_FALSE zurueck).
    try:
        pythoncom.CoInitialize()
    except Exception:
        pass  # Bereits initialisiert oder anderer Apartment-Typ - ignorieren

    ctx    = pythoncom.CreateBindCtx(0)
    rot    = pythoncom.GetRunningObjectTable()
    enum_mk = rot.EnumRunning()

    while True:
        monikers = enum_mk.Next(1)
        if not monikers:
            break
        mk = monikers[0]
        try:
            name = mk.GetDisplayName(ctx, None)
            if name == "SAPGUI":
                obj = rot.GetObject(mk)
                return wc.Dispatch(obj.QueryInterface(pythoncom.IID_IDispatch))
        except Exception:
            continue

    raise RuntimeError(
        "SAP GUI nicht in der ROT gefunden. Bitte sicherstellen:\n"
        "  1. SAP GUI ist geoeffnet und eingeloggt.\n"
        "  2. SAP GUI Scripting ist aktiviert: Customizing → Accessibility & Scripting → Scripting → Enable.\n"
        "  3. Der Bridge-Prozess laeuft auf demselben Windows-Desktop wie SAP GUI."
    )


def _gui_session(creds=None):
    """
    Oeffnet eine NEUE SAP-GUI-Scripting-Session via SAP Logon (OpenConnection).
    Benoetigt einen gueltigen SAP_GUI_CONNECTION-Eintragsnamen in .env.

    Wenn creds 'user' UND 'password' enthaelt, wird der Vault-Resolver komplett
    uebersprungen - so kann SAP_USER/SAP_PASSWORD in .env leer bleiben.
    """
    import win32com.client as wc  # type: ignore  # pip install pywin32
    if creds and creds.get("user") and creds.get("password"):
        # Vollstaendige Override - kein Vault-Lookup noetig
        cfg = {
            "connection": creds.get("connection") or os.getenv("SAP_GUI_CONNECTION", "") or os.getenv("SAP_GUI_CONNECTION_PRD", ""),
            "user":       creds["user"],
            "password":   creds["password"],
            "client":     creds.get("client") or os.getenv("SAP_CLIENT", ""),
            "lang":       creds.get("lang") or os.getenv("SAP_LANG", "DE"),
        }
        log.info("GUI-Session mit Cockpit-User '%s' (Mandant %s)", cfg["user"], cfg["client"])
    else:
        # Fallback: aus .env / Vault
        cfg = build_sap_gui_params()
        if creds:
            # Partial-Override (nur Mandant/Sprache)
            if creds.get("client"): cfg["client"] = creds["client"]
            if creds.get("lang"):   cfg["lang"]   = creds["lang"]
    sap_gui_auto = _get_sapgui_from_rot()
    application  = sap_gui_auto.GetScriptingEngine
    connection   = application.OpenConnection(cfg["connection"], True)
    session = connection.Children(0)
    if session.findById("wnd[0]/usr/txtRSYST-MANDT", False):
        session.findById("wnd[0]/usr/txtRSYST-MANDT").text = cfg["client"]
        session.findById("wnd[0]/usr/txtRSYST-BNAME").text = cfg["user"]
        session.findById("wnd[0]/usr/pwdRSYST-BCODE").text = cfg["password"]
        session.findById("wnd[0]/usr/txtRSYST-LANGU").text = cfg["lang"]
        session.findById("wnd[0]").sendVKey(0)
    return connection, session


def _get_gui_session(creds=None, target_ashost: str = ""):
    """
    Haengt sich an eine LAUFENDE SAP-GUI-Session (ROT-Ansatz, kein SAP Logon).
    Respektiert G-5: bevorzugt Session auf ``target_ashost``, faellt auf eine
    beliebige freie Session zurueck wenn keine IP-/Hostname-Uebereinstimmung.

    Rueckgabe: (connection, session, should_close=False)
    should_close ist immer False – wir oeffnen KEINE neue Verbindung.

    WICHTIG: SAP GUI muss manuell geoeffnet und eingeloggt sein.
    OpenConnection (SAP Logon) wird bewusst nicht verwendet – der Verbindungsname
    in .env stimmt moeglicherweise nicht mit SAP Logon ueberein.
    """
    sap_gui = _get_sapgui_from_rot()
    app     = sap_gui.GetScriptingEngine

    def _host_matches(sess, target: str) -> bool:
        """Prueft ob die Session zum gewuenschten SAP-Server gehoert."""
        if not target:
            return True
        try:
            app_server = str(sess.info.applicationServer or "")
            # Vergleich: IP in Hostname, Hostname in IP, oder direkte Gleichheit
            return (target in app_server) or (app_server in target)
        except Exception:
            return True   # Info nicht lesbar → nicht ausschliessen

    # Alle verfuegbaren Sessions einsammeln
    all_sessions: list[tuple] = []     # (conn, sess, server_str)
    busy_sessions: list[str]  = []     # Diagnose: belegte Sessions
    try:
        n_conn = app.Children.Count
        log.info("GUI: %d SAP-Verbindung(en) gefunden", n_conn)
        for ci in range(n_conn):
            conn   = app.Children(ci)
            n_sess = conn.Children.Count
            for si in range(n_sess):
                sess = conn.Children(si)
                is_busy = getattr(sess, "Busy", False)   # default False → nicht vorschnell ausschliessen
                try:
                    server = str(sess.info.applicationServer or "?")
                    tx     = str(sess.info.transaction or "?")
                except Exception:
                    server, tx = "?", "?"
                if is_busy:
                    busy_sessions.append(f"conn[{ci}] sess[{si}] server={server} tx={tx} BUSY")
                    log.warning("GUI: Session conn[%d]/sess[%d] ist BUSY (tx=%s, server=%s) – wird übersprungen", ci, si, tx, server)
                    continue
                log.info("GUI: Freie Session gefunden – conn[%d]/sess[%d] server=%s tx=%s", ci, si, server, tx)
                all_sessions.append((conn, sess, server))
    except Exception as e:
        raise RuntimeError(
            f"GUI: Session-Auflistung fehlgeschlagen ({e}). "
            "SAP GUI ggf. nicht geoeffnet oder Scripting nicht aktiviert."
        )

    if not all_sessions:
        detail = ""
        if busy_sessions:
            detail = f" Gefundene belegte Sessions: {'; '.join(busy_sessions)}."
        raise RuntimeError(
            "Keine freie SAP GUI Session gefunden."
            f"{detail} Bitte sicherstellen:\n"
            f"  • SAP GUI ist geoeffnet und eingeloggt{' auf ' + target_ashost if target_ashost else ''}.\n"
            "  • Kein modaler Dialog / Popup ist offen (z.B. Mehrfachanmeldung, Hinweis).\n"
            "  • Die Session befindet sich im Easy Access Menu oder einer nicht-modalen Transaktion."
        )

    # Pass 1: Session auf dem richtigen System (G-5-konform)
    for conn, sess, server in all_sessions:
        if _host_matches(sess, target_ashost):
            log.info(
                "GUI: Session gefunden – Server=%s (Ziel-ashost=%s)",
                server, target_ashost or "(beliebig)",
            )
            return conn, sess, False

    # Pass 2: Kein Host-Match → erste verfuegbare Session verwenden (Warnung)
    conn, sess, server = all_sessions[0]
    log.warning(
        "GUI: Kein Session-Match fuer ashost='%s'. "
        "Verwende erste freie Session auf '%s' (G-5 eingeschraenkt erfuellt). "
        "Bitte sicherstellen, dass SAP GUI auf dem richtigen System eingeloggt ist.",
        target_ashost, server,
    )
    return conn, sess, False


# ---------------------------------------------------------------------------
# GUI Session mit Cockpit-Credentials (neue Session, automatischer Login)
# ---------------------------------------------------------------------------

# ashost → SAP-Logon-Verbindungsname (aus .env)
_ASHOST_TO_GUI_CONN: dict[str, str] = {}

def _init_ashost_map() -> None:
    """Füllt _ASHOST_TO_GUI_CONN aus Umgebungsvariablen."""
    global _ASHOST_TO_GUI_CONN
    _ASHOST_TO_GUI_CONN = {
        "172.28.189.8":  os.getenv("SAP_GUI_CONNECTION_SEP") or os.getenv("SAP_GUI_CONNECTION", ""),
        "172.28.189.11": os.getenv("SAP_GUI_CONNECTION_SEQ") or os.getenv("SAP_GUI_CONNECTION", ""),
    }

_init_ashost_map()


def _gui_session_with_sap_auth(sap_auth: dict):
    """
    Öffnet eine NEUE SAP GUI Session und loggt sich mit den Cockpit-Credentials ein.
    Gibt (connection, session) zurück; Aufrufer MUSS Session anschließend schließen.

    Benötigt in .env:
      SAP_GUI_CONNECTION_SEP  → SAP Logon Eintragsname für SEP (172.28.189.8)
      SAP_GUI_CONNECTION_SEQ  → SAP Logon Eintragsname für SEQ (172.28.189.11)
    Falls nicht gesetzt: Fallback auf SAP_GUI_CONNECTION.
    """
    ashost  = sap_auth.get("ashost", "")
    user    = sap_auth.get("user", "") or sap_auth.get("sap_username", "")
    passwd  = sap_auth.get("passwd", "") or sap_auth.get("password", "")
    client  = sap_auth.get("client", "") or os.getenv("SAP_CLIENT", "600")
    lang    = sap_auth.get("lang", "DE")

    conn_name = _ASHOST_TO_GUI_CONN.get(ashost) or os.getenv("SAP_GUI_CONNECTION", "")
    if not conn_name:
        raise RuntimeError(
            f"Kein SAP-Logon-Eintragsname für ashost={ashost!r} konfiguriert. "
            "Bitte SAP_GUI_CONNECTION_SEP / SAP_GUI_CONNECTION_SEQ in worker/.env eintragen."
        )

    creds = {
        "connection": conn_name,
        "user":       user,
        "password":   passwd,
        "client":     client,
        "lang":       lang,
    }
    log.info("Neue SAP GUI Session: user=%s conn=%s client=%s", user, conn_name, client)
    connection, session = _gui_session(creds)
    import time as _t; _t.sleep(1.5)   # kurz warten bis Logon abgeschlossen
    return connection, session


def _hide_session_window(session) -> None:
    """
    Versteckt das SAP GUI Fenster für die Hintergrundausführung.
    Strategie:
      1. session.findById("wnd[0]").Handle  → Win32-HWND → SW_HIDE (komplett unsichtbar)
      2. Fallback: iconify() → minimiert in die Taskleiste
    Das Fenster bleibt vollständig steuerbar via SAP GUI Scripting.
    """
    import time as _t
    _t.sleep(0.3)
    hidden = False

    # Versuch 1: HWND über SAP GUI Scripting Handle-Eigenschaft
    try:
        hwnd = int(session.findById("wnd[0]").Handle)
        import win32gui, win32con  # type: ignore
        win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
        log.info("SAP GUI Fenster versteckt (HWND=%d)", hwnd)
        hidden = True
    except Exception as e:
        log.debug("Win32 Handle-Hide fehlgeschlagen: %s", e)

    # Versuch 2: Alle SAP-Fenster via EnumWindows ausblenden
    if not hidden:
        try:
            import win32gui, win32con  # type: ignore
            _t.sleep(0.3)
            found = []
            def _cb(hwnd, _):
                if win32gui.IsWindowVisible(hwnd):
                    cls = win32gui.GetClassName(hwnd)
                    if "SAP" in cls.upper() or "sap" in cls.lower():
                        found.append(hwnd)
            win32gui.EnumWindows(_cb, None)
            for hwnd in found:
                win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
            if found:
                log.info("SAP GUI Fenster versteckt via EnumWindows (%d gefunden)", len(found))
                hidden = True
        except Exception as e:
            log.debug("EnumWindows Hide fehlgeschlagen: %s", e)

    # Fallback 3: Minimieren in Taskleiste
    if not hidden:
        try:
            session.findById("wnd[0]").iconify()
            log.info("SAP GUI Fenster minimiert (iconify)")
        except Exception as e:
            log.debug("iconify fehlgeschlagen: %s", e)


# ---------------------------------------------------------------------------
# AP / BAPI -- FI-Beleg buchen
# ---------------------------------------------------------------------------
def _rfc_connection_with_auth(sap_auth: dict | None = None, config=None):
    """
    Baut RFC-Verbindung auf. Wenn ``sap_auth`` aus dem Payload kommt (Cockpit-Login),
    werden diese Werte bevorzugt – andernfalls Fallback auf .env / Vault.
    Optionaler ``config``-Parameter wird direkt an pyrfc.Connection weitergereicht
    (z.B. config={'bcd': 'str'} als Workaround fuer den pyrfc 3.3 BCD-Bug).
    """
    from pyrfc import Connection  # type: ignore
    if sap_auth and sap_auth.get("user") and sap_auth.get("passwd"):
        params = {
            "ashost": sap_auth.get("ashost") or os.getenv("SAP_ASHOST"),
            "sysnr":  sap_auth.get("sysnr")  or os.getenv("SAP_SYSNR", "00"),
            "client": sap_auth.get("client") or os.getenv("SAP_CLIENT"),
            "user":   sap_auth["user"],
            "passwd": sap_auth["passwd"],
            "lang":   sap_auth.get("lang")   or os.getenv("SAP_LANG", "DE"),
        }
        log.info("RFC-Verbindung mit Cockpit-Credentials: %s@%s Mdt.%s",
                 params["user"], params["ashost"], params["client"])
    else:
        params = build_sap_rfc_params("SAP")
        log.info("RFC-Verbindung mit .env-Credentials: %s@%s Mdt.%s",
                 params.get("user"), params.get("ashost"), params.get("client"))
    return Connection(config=config, **params) if config else Connection(**params)


def _rfc_connection_service_account(sap_auth: dict | None = None, config=None):
    """
    Fallback-Verbindung mit Service-Account aus .env, aber SAP-System-Parameter
    aus sap_auth (Host/Sysnr/Client/Lang). Wird verwendet wenn der eingeloggte
    Benutzer keine S_RFC-Berechtigung für RFCPING hat.
    """
    from pyrfc import Connection  # type: ignore
    env_params = build_sap_rfc_params("SAP")
    params = {
        # SAP-System aus sap_auth (kann SEQ sein), Credentials aus .env
        "ashost": (sap_auth or {}).get("ashost") or env_params.get("ashost") or os.getenv("SAP_ASHOST"),
        "sysnr":  (sap_auth or {}).get("sysnr")  or env_params.get("sysnr")  or os.getenv("SAP_SYSNR", "00"),
        "client": (sap_auth or {}).get("client") or env_params.get("client") or os.getenv("SAP_CLIENT"),
        "user":   env_params.get("user")   or os.getenv("SAP_USER", "RFC_COFACE"),
        "passwd": env_params.get("passwd") or os.getenv("SAP_PASSWORD", ""),
        "lang":   (sap_auth or {}).get("lang") or env_params.get("lang") or os.getenv("SAP_LANG", "DE"),
    }
    log.info("RFC-Verbindung Service-Account (RFCPING-Fallback): %s@%s Mdt.%s",
             params["user"], params["ashost"], params["client"])
    return Connection(config=config, **params) if config else Connection(**params)


def bapi_acc_document_post(task: dict, payload: dict) -> str:
    sap_auth = payload.get("_sap_auth") or payload.get("_sap_conn")

    # Geplante Buchungen (Scheduler): use_service_account=True -> direkt Service-Account,
    # SAP-System-Parameter (ashost/sysnr/client) kommen trotzdem aus sap_auth.
    use_service_acct = bool((sap_auth or {}).get("use_service_account"))

    if use_service_acct:
        log.info("Scheduler-Buchung: verwende Service-Account aus .env, System aus sap_auth.")
        conn = _rfc_connection_service_account(sap_auth)
    else:
        # Interaktive Buchung: erst mit Benutzer-Credentials versuchen,
        # bei RFCPING-Sperre Fallback auf Service-Account (RFC_COFACE aus .env).
        try:
            conn = _rfc_connection_with_auth(sap_auth)
        except Exception as _conn_exc:
            _exc_str = str(_conn_exc)
            if "RFC_NO_AUTHORITY" in _exc_str and "RFCPING" in _exc_str:
                log.warning(
                    "Benutzer '%s' hat keine S_RFC-Berechtigung für RFCPING – "
                    "Fallback auf Service-Account für RFC-Verbindung.",
                    (sap_auth or {}).get("user", "?")
                )
                conn = _rfc_connection_service_account(sap_auth)
            else:
                raise
    try:
        # Payload-Schluessel tolerant: akzeptiert Gross- und Kleinschreibung
        def _get(key):
            return payload.get(key) or payload.get(key.lower()) or payload.get(key.upper()) or {}

        doc_header   = _get("DOCUMENTHEADER") or _get("header")
        account_gl   = _get("ACCOUNTGL")      or _get("accountgl")      or []
        account_ap   = _get("ACCOUNTPAYABLE") or _get("accountpayable") or []
        account_ar   = _get("ACCOUNTRECEIVABLE") or _get("accountreceivable") or []
        account_tax  = _get("ACCOUNTTAX")     or _get("accounttax")     or []
        curr_amount  = _get("CURRENCYAMOUNT") or _get("currencyamount") or []

        # Betragsfelder normalisieren -> pyrfc erwartet float, nicht int
        # JSON-Parser liefert 2270 als int, 2270.33 als float – beides auf float casten
        _amount_fields = ("AMT_DOCCUR", "AMT_BASE", "AMT_DOCCUR2", "AMT_DOCCUR3",
                          "AMT_DOCCUR4", "AMT_DOCCUR5")
        for item in (curr_amount if isinstance(curr_amount, list) else []):
            for f in _amount_fields:
                if f in item and item[f] is not None:
                    try:
                        item[f] = float(item[f])
                    except (TypeError, ValueError):
                        pass

        # ACCOUNTTAX vorbereiten (P-10: RESY Vorsteuer V1/V4)
        # Nur ITEMNO_ACC + TAX_CODE an BAPI übergeben – BAPITAX09 dieses Systems kennt
        # TAX_BASE_AMOUNT / TAX_AMOUNT / COUNTRY / TAX_RATE nicht (RFC_INVALID_PARAMETER).
        # Steuerbeträge laufen über CURRENCYAMOUNT (jeweils ein Eintrag pro ACCOUNTTAX-Item).
        _TAX_BAPI_FIELDS = {"ITEMNO_ACC", "TAX_CODE"}
        bapi_account_tax = []
        for item in (account_tax if isinstance(account_tax, list) else []):
            bapi_account_tax.append({k: v for k, v in item.items() if k in _TAX_BAPI_FIELDS})
        if bapi_account_tax:
            log.info("BAPI_ACC_DOCUMENT_POST: ACCOUNTTAX %d Eintraege (RESY P-10): %s",
                     len(bapi_account_tax),
                     ", ".join(f"{t.get('TAX_CODE','?')} Basis={item.get('TAX_BASE_AMOUNT',0):.2f} Steuer={item.get('TAX_AMOUNT',0):.2f}"
                               for t, item in zip(bapi_account_tax, account_tax)))

        # Datumsfelder normalisieren -> pyrfc erwartet YYYYMMDD (ohne Trennzeichen)
        for date_field in ("DOC_DATE", "PSTNG_DATE", "BASELINE_DATE"):
            raw = doc_header.get(date_field)
            if raw:
                try:
                    d = parse_date(raw, default_today=False)
                    doc_header[date_field] = d.strftime("%Y%m%d")
                except Exception:
                    pass  # unveraendert weitergeben - SAP gibt dann klare Fehlermeldung

        # Trading Partner (Intercompany) -- Feld TRADE_ID in BAPIACGL09 (je GL-Position).
        # Akzeptiert aus dem Payload: "TRADE_ID", "trading_partner", "TRADING_PARTNER".
        # TRADING_PART im DOCUMENTHEADER existiert auf diesem System nicht (RFC_INVALID_PARAMETER).
        trading_partner = (
            payload.get("TRADE_ID")
            or payload.get("trade_id")
            or payload.get("trading_partner")
            or payload.get("TRADING_PARTNER")
        )
        # Falls das Cockpit TRADE_ID direkt in den ACCOUNTGL-Items mitschickt,
        # vorhandene Werte beibehalten – nur fehlende ergaenzen.
        if trading_partner:
            trading_partner = str(trading_partner).strip()
            for item in (account_gl if isinstance(account_gl, list) else []):
                if not item.get("TRADE_ID"):
                    item["TRADE_ID"] = trading_partner
            log.info("BAPI_ACC_DOCUMENT_POST: TRADE_ID (Trading Partner) = %s", trading_partner)

        log.info("BAPI_ACC_DOCUMENT_POST: BUKRS=%s BLART=%s USER=%s DOC_DATE=%s GL-Pos=%d",
                 doc_header.get("COMP_CODE","?"),
                 doc_header.get("DOC_TYPE","?"),
                 doc_header.get("USERNAME","?"),
                 doc_header.get("DOC_DATE","?"),
                 len(account_gl) if isinstance(account_gl, list) else 0)

        call_kwargs = dict(
            DOCUMENTHEADER=doc_header,
            ACCOUNTGL=account_gl,
            ACCOUNTPAYABLE=account_ap,
            ACCOUNTRECEIVABLE=account_ar,
            CURRENCYAMOUNT=curr_amount,
        )
        if bapi_account_tax:
            call_kwargs["ACCOUNTTAX"] = bapi_account_tax

        result = conn.call("BAPI_ACC_DOCUMENT_POST", **call_kwargs)
        # Alle RETURN-Meldungen loggen für Diagnose
        for r in result.get("RETURN", []):
            lvl = (logging.ERROR if r["TYPE"] in ("E","A")
                   else logging.WARNING if r["TYPE"] == "W"
                   else logging.INFO)
            log.log(lvl, "BAPI RETURN [%s] ID=%s NR=%s: %s (Feld: %s)",
                    r["TYPE"], r.get("ID",""), r.get("NUMBER",""),
                    r.get("MESSAGE",""), r.get("FIELD",""))
        errs = [r for r in result.get("RETURN", []) if r["TYPE"] in ("E", "A")]
        if errs:
            detail = "; ".join(
                f"[{r.get('ID','?')}/{r.get('NUMBER','?')}] {r['MESSAGE']}" for r in errs
            )
            raise RuntimeError(f"BAPI-Fehler: {detail}")
        conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")
        obj_key = result.get("OBJ_KEY") or result.get("OBJECTKEY") or ""
        # OBJ_KEY Format (je nach SAP-Konfiguration):
        #   "2300000510VV9 2026"           → parts[0] = "2300000510VV9", parts[1] = "2026"
        #   "BKPFF 2300000498VV9 2026 ..." → parts[0] = "BKPFF",         parts[1] = "2300000498VV9"
        # BELNR = erste 10 Ziffern des ersten numerisch-startenden Tokens
        doc_nr = ""
        if obj_key:
            parts = obj_key.split()
            for token in parts:
                if token and token[0].isdigit():
                    doc_nr = token[:10]   # z.B. "2300000510"
                    break
            if not doc_nr:
                doc_nr = obj_key[:10]
        log.info("BAPI_ACC_DOCUMENT_POST OK: Beleg %s (OBJ_KEY=%s)", doc_nr or "(leer)", obj_key)
        return {"doc_number": doc_nr or obj_key or "?", "obj_key": obj_key}
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# FI -- Beleg stornieren (BAPI_ACC_DOCUMENT_REV)
# ---------------------------------------------------------------------------
def bapi_acc_document_rev(task: dict, payload: dict) -> str:
    """Storniert einen FI-Beleg via BAPI_ACC_DOCUMENT_REV."""
    sap_auth = payload.get("_sap_auth") or payload.get("_sap_conn")
    use_service_acct = bool((sap_auth or {}).get("use_service_account"))
    if use_service_acct:
        conn = _rfc_connection_service_account(sap_auth)
    else:
        try:
            conn = _rfc_connection_with_auth(sap_auth)
        except Exception as _conn_exc:
            _exc_str = str(_conn_exc)
            if "RFC_NO_AUTHORITY" in _exc_str and "RFCPING" in _exc_str:
                log.warning("RFCPING-Fallback auf Service-Account für Storno.")
                conn = _rfc_connection_service_account(sap_auth)
            else:
                raise
    try:
        belnr = str(payload.get("BELNR") or payload.get("doc_number", "")).strip().zfill(10)
        bukrs = str(payload.get("BUKRS") or payload.get("comp_code", "")).strip()
        gjahr = str(payload.get("GJAHR") or payload.get("fiscal_year", "")).strip()
        user  = str((sap_auth or {}).get("user", "")).strip()

        if not belnr.strip("0") or not bukrs or not gjahr:
            raise ValueError(f"Pflichtfelder fehlen: BELNR={belnr}, BUKRS={bukrs}, GJAHR={gjahr}")

        from datetime import date as _date
        today = _date.today().strftime("%Y%m%d")

        log.info("BAPI_ACC_DOCUMENT_REV: BELNR=%s BUKRS=%s GJAHR=%s USER=%s", belnr, bukrs, gjahr, user)

        # Versuch 1: BAPI_ACC_DOCUMENT_REV (Standard-BAPI)
        storno_doc = None
        try:
            result = conn.call(
                "BAPI_ACC_DOCUMENT_REV",
                DOCUMENTHEADER={
                    "USERNAME":   user,
                    "COMP_CODE":  bukrs,
                    "DOC_DATE":   today,
                    "PSTNG_DATE": today,
                },
                REVERSAL={
                    "BELNR": belnr,
                    "BUKRS": bukrs,
                    "GJAHR": gjahr,
                    "REASON": "01",
                },
            )
            for r in result.get("RETURN", []):
                lvl = logging.ERROR if r["TYPE"] in ("E", "A") else logging.INFO
                log.log(lvl, "BAPI_ACC_DOCUMENT_REV RETURN [%s] %s: %s",
                        r["TYPE"], r.get("ID", ""), r.get("MESSAGE", ""))
            errs = [r for r in result.get("RETURN", []) if r["TYPE"] in ("E", "A")]
            if errs:
                raise RuntimeError("Storno-Fehler: " + "; ".join(
                    f"[{r.get('ID','?')}/{r.get('NUMBER','?')}] {r['MESSAGE']}" for r in errs))
            conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")
            storno_doc = result.get("OBJ_KEY") or result.get("OBJECTKEY") or ""
            log.info("BAPI_ACC_DOCUMENT_REV OK: Stornobeleg %s", storno_doc)

        except Exception as _bapi_exc:
            _exc_str = str(_bapi_exc)
            # FU_NOT_FOUND = Funktionsbaustein existiert nicht im System → Fallback
            if "FU_NOT_FOUND" not in _exc_str and "FL 046" not in _exc_str and "Number:046" not in _exc_str:
                raise  # anderer Fehler → direkt weiterwerfen

            log.warning("BAPI_ACC_DOCUMENT_REV nicht verfügbar in diesem SAP-System (%s). "
                        "Versuche Fallback FI_DOCUMENT_REVERSE.", _exc_str[:80])

            # Versuch 2: FI_DOCUMENT_REVERSE (älteres FM, breiter verfügbar)
            try:
                result2 = conn.call(
                    "FI_DOCUMENT_REVERSE",
                    I_BUKRS=bukrs,
                    I_BELNR=belnr,
                    I_GJAHR=gjahr,
                    I_BUDAT=today,
                    I_BLDAT=today,
                    I_STGRD="01",
                )
                storno_doc = (result2.get("E_BELNR") or result2.get("BELNR_NEW") or "").strip()
                if not storno_doc:
                    storno_doc = ""
                conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")
                log.info("FI_DOCUMENT_REVERSE OK: Stornobeleg %s", storno_doc)

            except Exception as _fb_exc:
                _fb_str = str(_fb_exc)
                if "FU_NOT_FOUND" in _fb_str or "FL 046" in _fb_str or "Number:046" in _fb_str:
                    # Beide FMs fehlen → klare Handlungsanweisung
                    raise RuntimeError(
                        "SAP-Konfigurationsfehler: Weder BAPI_ACC_DOCUMENT_REV noch FI_DOCUMENT_REVERSE "
                        f"sind im System verfügbar (ashost={conn.get_connection_attributes().get('partnerHost','')}). "
                        "Bitte SAP Basis bitten, den Funktionsbaustein BAPI_ACC_DOCUMENT_REV zu aktivieren. "
                        f"Alternativ: Beleg {belnr}/{bukrs}/{gjahr} manuell in SAP über Transaktion FB08 stornieren."
                    ) from None
                raise  # anderer Fehler von FI_DOCUMENT_REVERSE

        storno_doc = storno_doc or "(kein Stornobeleg)"
        log.info("Storno abgeschlossen: Stornobeleg %s", storno_doc)
        return f"Storno erfolgreich: {storno_doc}"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# AP / BAPI -- Offene Posten Lieferanten (FBL1N)
# ---------------------------------------------------------------------------
def bapi_vendor_open_items(task: dict, payload: dict) -> str:
    conn = _rfc_connection()
    try:
        res = conn.call(
            "BAPI_AP_ACC_GETOPENITEMS",
            COMPANYCODE=payload.get("comp_code", "1000"),
            VENDOR=payload.get("vendor"),
            KEYDATE=payload.get("keydate", datetime.now().strftime("%Y%m%d")),
        )
        items = res.get("LINEITEMS", [])
        return f"OP-Liste Lieferanten gezogen: {len(items)} Posten."
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# AR / BAPI -- Offene Posten Kunden (FBL5N) + Excel-Report
# ---------------------------------------------------------------------------
def bapi_customer_open_items(task: dict, payload: dict) -> str:
    """
    Zieht offene Posten je Kunde via BAPI_AR_ACC_GETKEYDATEBALANCE bzw.
    BAPI_AR_ACC_GETOPENITEMS, baut Altersstruktur (0-30/31-60/61-90/>90)
    und schreibt einen Excel-Bericht ins REPORT_OUT_DIR.

    Erwartete Payload-Felder (mit Defaults):
        comp_code        Buchungskreis              "1000"
        customer_from    Kunde von                  ""
        customer_to      Kunde bis                  "9999999999"
        keydate          Stichtag YYYYMMDD          heute
        aging_buckets    Altersklassen in Tagen     "30,60,90"
        currency         Berichts-Waehrung          "EUR"
    """
    comp_code     = payload.get("comp_code",     "1000")
    customer_from = payload.get("customer_from", "")
    customer_to   = payload.get("customer_to",   "9999999999")
    keydate       = parse_date(payload.get("keydate"), default_today=True)
    keydate_str   = keydate.strftime("%Y%m%d")
    buckets       = [int(x) for x in str(payload.get("aging_buckets", "30,60,90")).split(",")]

    # ---- 1. Daten aus SAP ziehen ----
    items = []
    conn = _rfc_connection()
    try:
        # Strategie: zuerst alle Kunden im Range ermitteln (DDIC: KNB1),
        # dann pro Kunde BAPI_AR_ACC_GETOPENITEMS aufrufen.
        # Fuer Demos reicht ein einziger Range-Call ueber RFC_READ_TABLE.
        cust_resp = conn.call(
            "RFC_READ_TABLE",
            QUERY_TABLE="KNB1",
            DELIMITER="|",
            FIELDS=[{"FIELDNAME": "KUNNR"}],
            OPTIONS=[{"TEXT": f"BUKRS = '{comp_code}'"}],
            ROWCOUNT=5000,
        )
        customers = [row["WA"].strip("|") for row in cust_resp.get("DATA", [])]
        customers = [c for c in customers if (not customer_from or c >= customer_from)
                                          and (not customer_to   or c <= customer_to)]
        log.info("FBL5N: %d Kunden im Range %s..%s", len(customers), customer_from, customer_to)

        for kunnr in customers:
            try:
                op = conn.call(
                    "BAPI_AR_ACC_GETOPENITEMS",
                    COMPANYCODE=comp_code,
                    CUSTOMER=kunnr,
                    KEYDATE=keydate_str,
                )
                for it in op.get("LINEITEMS", []):
                    items.append({
                        "Kunde":         kunnr,
                        "Belegnummer":   it.get("DOC_NO", ""),
                        "Belegart":      it.get("DOC_TYPE", ""),
                        "Belegdatum":    it.get("DOC_DATE", ""),
                        "Faelligkeit":   it.get("NETDUEDATE", "") or it.get("DUE_DATE", ""),
                        "Betrag":        float(it.get("AMT_DOCCUR", 0) or 0),
                        "Waehrung":      it.get("CURRENCY", ""),
                        "Buchungstext":  it.get("ITEM_TEXT", ""),
                    })
            except Exception as exc:  # noqa: BLE001
                log.warning("Kunde %s: %s", kunnr, exc)
    finally:
        conn.close()

    if not items:
        return "Keine offenen Posten gefunden - kein Report erzeugt."

    # ---- 2. Altersstruktur berechnen ----
    def bucket(due: str) -> str:
        if not due:
            return ">faellig"
        try:
            d = datetime.strptime(due, "%Y%m%d").date()
        except ValueError:
            return "unbekannt"
        days_overdue = (keydate - d).days
        if days_overdue < 0:
            return "nicht-faellig"
        last = 0
        for b in buckets:
            if days_overdue <= b:
                return f"{last+1}-{b} Tage"
            last = b
        return f">{buckets[-1]} Tage"

    for it in items:
        it["Bucket"] = bucket(it["Faelligkeit"])

    # ---- 3. Excel-Report schreiben ----
    fname = f"FBL5N_OP_Kunden_{comp_code}_{keydate_str}_{datetime.now().strftime('%H%M%S')}.xlsx"
    fpath = REPORT_OUT_DIR / fname
    _write_open_items_xlsx(fpath, items, comp_code, keydate, buckets)

    total = sum(it["Betrag"] for it in items)
    return (f"FBL5N: {len(items)} OP, Gesamt {total:,.2f} EUR. "
            f"Report: {fpath}")


def _write_open_items_xlsx(path, items, comp_code, keydate, buckets) -> None:
    """Schreibt OP-Liste + Pivot nach Altersbucket in eine .xlsx."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl fehlt. Installiere: pip install openpyxl"
        ) from exc

    wb = Workbook()

    # --- Tab 1: Detail ---
    ws = wb.active
    ws.title = "Offene Posten"
    headers = ["Kunde", "Belegnummer", "Belegart", "Belegdatum", "Faelligkeit",
               "Bucket", "Betrag", "Waehrung", "Buchungstext"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0a66c2")
    for it in items:
        ws.append([it.get(h, "") for h in headers])
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64+i)].width = 16

    # --- Tab 2: Altersanalyse ---
    pivot = wb.create_sheet("Altersanalyse")
    bucket_labels = ["nicht-faellig"] + [f"1-{buckets[0]} Tage"]
    last = buckets[0]
    for b in buckets[1:]:
        bucket_labels.append(f"{last+1}-{b} Tage"); last = b
    bucket_labels += [f">{buckets[-1]} Tage", ">faellig", "unbekannt"]

    by_bucket = {b: 0.0 for b in bucket_labels}
    by_bucket_count = {b: 0 for b in bucket_labels}
    for it in items:
        b = it.get("Bucket", "unbekannt")
        if b not in by_bucket:
            by_bucket[b] = 0.0; by_bucket_count[b] = 0
        by_bucket[b] += it["Betrag"]
        by_bucket_count[b] += 1

    pivot.append(["Altersklasse", "Anzahl Posten", "Summe EUR"])
    for cell in pivot[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0a66c2")
    for b in bucket_labels:
        if by_bucket_count[b] > 0:
            pivot.append([b, by_bucket_count[b], round(by_bucket[b], 2)])
    pivot.append(["GESAMT", sum(by_bucket_count.values()), round(sum(by_bucket.values()), 2)])
    pivot[pivot.max_row][0].font = Font(bold=True)
    pivot.column_dimensions["A"].width = 22
    pivot.column_dimensions["B"].width = 16
    pivot.column_dimensions["C"].width = 16

    # --- Tab 3: Meta ---
    meta = wb.create_sheet("Info")
    meta["A1"] = "Bericht"; meta["B1"] = "Offene Posten Kunden (FBL5N)"
    meta["A2"] = "Buchungskreis"; meta["B2"] = comp_code
    meta["A3"] = "Stichtag"; meta["B3"] = keydate.strftime("%d.%m.%Y")
    meta["A4"] = "Erzeugt"; meta["B4"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    meta["A5"] = "Anzahl Posten"; meta["B5"] = len(items)
    meta["A6"] = "Altersklassen"; meta["B6"] = ", ".join(str(b) for b in buckets) + " Tage"
    for row in meta["A1":"A6"]:
        for cell in row:
            cell.font = Font(bold=True)
    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 40

    wb.save(path)


# ---------------------------------------------------------------------------
# AR / SQL -- Offene Posten Kunden direkt aus Sybase ASE (BSID-Tabelle)
# ---------------------------------------------------------------------------
def sql_customer_open_items(task: dict, payload: dict) -> str:
    """
    Wie ``bapi_customer_open_items``, aber statt BAPI direkter SQL-Zugriff auf
    SAP-Tabellen in Sybase ASE. Brauchst, wenn kein NW RFC SDK installiert ist.

    Liest aus:
        BSID  - offene Debitorenposten (open AR items)
        KNA1  - Kundenstamm (allgemein)

    Erwartete Payload-Felder (mit Defaults):
        comp_code        Buchungskreis              "1000"
        customer_from    Kunde von                  "0000000000"
        customer_to      Kunde bis                  "9999999999"
        keydate          Stichtag YYYYMMDD          heute
        aging_buckets    Altersklassen in Tagen     "30,60,90"

    Ausgabe: dreitabige Excel-Datei nach ``REPORT_OUT_DIR``.
    """
    comp_code     = payload.get("comp_code",     "1000")
    customer_from = payload.get("customer_from", "0000000000") or "0000000000"
    customer_to   = payload.get("customer_to",   "9999999999") or "9999999999"
    keydate       = parse_date(payload.get("keydate"), default_today=True)
    keydate_str   = keydate.strftime("%Y%m%d")
    buckets       = [int(x) for x in str(payload.get("aging_buckets", "30,60,90")).split(",")]
    mandt         = os.getenv("SAP_CLIENT", "600")
    # SAP-on-ASE Tabellen liegen in einem Schema (Owner) - default 'SAPSR3'.
    # Fuer voll qualifizierten Namen 'DB.SCHEMA' kannst du z. B. SAP_DB_SCHEMA=SEQ.SAPSR3 setzen.
    schema = os.getenv("SAP_DB_SCHEMA", "SAPSR3")
    pfx = f"{schema}." if schema else ""
    log.info("FBL5N (SQL): Schema=%s Mandant=%s BUKRS=%s Kunde %s..%s Stichtag=%s",
             schema or "(kein)", mandt, comp_code, customer_from, customer_to, keydate_str)

    sql = f"""
        SELECT
            b.KUNNR,
            k.NAME1,
            b.BELNR,
            b.BLART,
            b.BLDAT,
            b.BUDAT,
            b.ZFBDT,
            b.WAERS,
            b.DMBTR,
            b.WRBTR,
            b.SGTXT,
            b.SHKZG
        FROM {pfx}BSID b
        LEFT JOIN {pfx}KNA1 k
               ON k.MANDT = b.MANDT
              AND k.KUNNR = b.KUNNR
        WHERE b.MANDT = ?
          AND b.BUKRS = ?
          AND b.KUNNR BETWEEN ? AND ?
          AND b.BUDAT <= ?
        ORDER BY b.KUNNR, b.BLDAT
    """
    items = []
    conn = sap_db.get_connection()
    try:
        cur = conn.cursor()
        cur.execute(sql, mandt, comp_code, customer_from, customer_to, keydate_str)
        for row in cur.fetchall():
            sign = -1 if (row[11] or "").strip() == "H" else 1
            betrag_hw = float(row[8] or 0) * sign
            betrag_bw = float(row[9] or 0) * sign
            items.append({
                "Kunde":         (row[0] or "").strip(),
                "Name":          (row[1] or "").strip(),
                "Belegnummer":   (row[2] or "").strip(),
                "Belegart":      (row[3] or "").strip(),
                "Belegdatum":    str(row[4] or ""),
                "Buchungsdatum": str(row[5] or ""),
                "Faelligkeit":   str(row[6] or ""),
                "Waehrung":      (row[7] or "").strip(),
                "Betrag (HW)":   round(betrag_hw, 2),
                "Betrag (BW)":   round(betrag_bw, 2),
                "Buchungstext":  (row[10] or "").strip(),
            })
    finally:
        conn.close()

    if not items:
        return "Keine offenen Debitorenposten gefunden - kein Report erzeugt."

    # ---- Altersbucket berechnen ----
    def bucket(due) -> str:
        try:
            d = parse_date(due, default_today=False)
        except Exception:
            return "unbekannt"
        days_overdue = (keydate - d).days
        if days_overdue < 0:
            return "nicht-faellig"
        last = 0
        for b in buckets:
            if days_overdue <= b:
                return f"{last+1}-{b} Tage"
            last = b
        return f">{buckets[-1]} Tage"

    for it in items:
        # ZFBDT (Basisdatum Faelligkeit) bevorzugt, sonst BUDAT als Fallback
        due = it["Faelligkeit"] if it["Faelligkeit"] else it["Buchungsdatum"]
        it["Bucket"] = bucket(due)

    # ---- Excel-Report schreiben ----
    fname = f"FBL5N_SQL_OP_Kunden_{comp_code}_{keydate_str}_{datetime.now().strftime('%H%M%S')}.xlsx"
    fpath = REPORT_OUT_DIR / fname
    _write_sql_open_items_xlsx(fpath, items, comp_code, keydate, buckets, mandt)

    total = sum(it["Betrag (HW)"] for it in items)
    return (f"FBL5N (SQL): {len(items)} OP, Gesamt {total:,.2f}. "
            f"Report: {fpath.name}")


def _write_sql_open_items_xlsx(path, items, comp_code, keydate, buckets, mandt) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # Tab 1: Detail
    ws = wb.active
    ws.title = "Offene Posten"
    headers = ["Kunde", "Name", "Belegnummer", "Belegart", "Belegdatum", "Buchungsdatum",
               "Faelligkeit", "Bucket", "Waehrung", "Betrag (HW)", "Betrag (BW)", "Buchungstext"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0a66c2")
    for it in items:
        ws.append([it.get(h, "") for h in headers])
    for i in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+i)].width = 16

    # Tab 2: Altersanalyse
    pivot = wb.create_sheet("Altersanalyse")
    bucket_labels = ["nicht-faellig", f"1-{buckets[0]} Tage"]
    last = buckets[0]
    for b in buckets[1:]:
        bucket_labels.append(f"{last+1}-{b} Tage"); last = b
    bucket_labels += [f">{buckets[-1]} Tage", "unbekannt"]

    by_bucket = {b: 0.0 for b in bucket_labels}
    by_bucket_count = {b: 0 for b in bucket_labels}
    for it in items:
        b = it.get("Bucket", "unbekannt")
        if b not in by_bucket:
            by_bucket[b] = 0.0; by_bucket_count[b] = 0
        by_bucket[b] += it["Betrag (HW)"]
        by_bucket_count[b] += 1

    pivot.append(["Altersklasse", "Anzahl Posten", "Summe HW"])
    for cell in pivot[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="0a66c2")
    for b in bucket_labels:
        if by_bucket_count[b] > 0:
            pivot.append([b, by_bucket_count[b], round(by_bucket[b], 2)])
    pivot.append(["GESAMT", sum(by_bucket_count.values()), round(sum(by_bucket.values()), 2)])
    pivot[pivot.max_row][0].font = Font(bold=True)
    for col, width in zip("ABC", [22, 16, 16]):
        pivot.column_dimensions[col].width = width

    # Tab 3: Info
    meta = wb.create_sheet("Info")
    meta["A1"] = "Bericht"; meta["B1"] = "Offene Posten Kunden (SQL/Sybase ASE)"
    meta["A2"] = "Datenquelle"; meta["B2"] = "BSID + KNA1 ueber ODBC"
    meta["A3"] = "Mandant"; meta["B3"] = mandt
    meta["A4"] = "Buchungskreis"; meta["B4"] = comp_code
    meta["A5"] = "Stichtag"; meta["B5"] = keydate.strftime("%d.%m.%Y")
    meta["A6"] = "Erzeugt"; meta["B6"] = datetime.now().strftime("%d.%m.%Y %H:%M:%S")
    meta["A7"] = "Anzahl Posten"; meta["B7"] = len(items)
    meta["A8"] = "Altersklassen"; meta["B8"] = ", ".join(str(b) for b in buckets) + " Tage"
    for row in meta["A1":"A8"]:
        for cell in row:
            cell.font = Font(bold=True)
    meta.column_dimensions["A"].width = 22
    meta.column_dimensions["B"].width = 50

    wb.save(path)


# ---------------------------------------------------------------------------
# Diagnose: ASE-Ping
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# Web-Pivot fuer Offene Posten Kunden (HTML)
# ---------------------------------------------------------------------------
def _write_pivot_html(path, items, comp_code, keydate, buckets, mandt):
    """Erzeugt eine Pivot-HTML-Tabelle: Kunde x Bucket, Summe Betrag (HW)."""
    # Bucket-Labels in sinnvoller Reihenfolge
    labels = ["nicht-faellig", f"1-{buckets[0]} Tage"]
    last = buckets[0]
    for b in buckets[1:]:
        labels.append(f"{last+1}-{b} Tage")
        last = b
    labels.append(f">{buckets[-1]} Tage")
    # "unbekannt" nur anhaengen wenn vorhanden
    if any(it.get("Bucket") == "unbekannt" for it in items):
        labels.append("unbekannt")

    # Pivot aufbauen
    pivot = {}     # (Kunde, Name) -> {bucket: summe}
    col_tot = {b: 0.0 for b in labels}
    grand = 0.0
    for it in items:
        key = (it["Kunde"], it["Name"])
        if key not in pivot:
            pivot[key] = {b: 0.0 for b in labels}
        b = it.get("Bucket", "unbekannt")
        if b not in pivot[key]:
            pivot[key][b] = 0.0
            col_tot.setdefault(b, 0.0)
        v = float(it.get("Betrag (HW)", 0) or 0)
        pivot[key][b] += v
        col_tot[b] += v
        grand += v

    def fmt(x):
        if not x: return ""
        return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    rows = []
    for (kunde, name) in sorted(pivot.keys()):
        cells = "".join(f"<td class='num'>{fmt(pivot[(kunde,name)].get(b,0))}</td>" for b in labels)
        row_total = sum(pivot[(kunde,name)].values())
        rows.append(
            f"<tr><td>{kunde}</td><td>{name}</td>{cells}"
            f"<td class='num total'>{fmt(row_total)}</td></tr>"
        )

    total_row = (
        "<tr class='gtotal'><td colspan='2'>Gesamtergebnis</td>"
        + "".join(f"<td class='num'>{fmt(col_tot[b])}</td>" for b in labels)
        + f"<td class='num'>{fmt(grand)}</td></tr>"
    )

    header_buckets = "".join(f"<th>{b}</th>" for b in labels)
    html = f"""<!doctype html>
<html lang="de">
<head>
<meta charset="utf-8">
<title>Offene Posten Kunden - Pivot</title>
<style>
  body {{ margin:0; padding:20px; background:#f6f7f9; color:#1a1a1a;
    font-family:-apple-system,"Segoe UI",Roboto,Arial,sans-serif; font-size:13px; }}
  h1 {{ font-size:20px; margin:0 0 4px 0; }}
  .meta {{ color:#666; font-size:12px; margin-bottom:18px; }}
  table {{ border-collapse:collapse; background:#fff; box-shadow:0 1px 3px rgba(0,0,0,.06); }}
  th, td {{ border:1px solid #e3e5e8; padding:6px 10px; text-align:left; }}
  th {{ background:#f3f4f6; font-weight:600; font-size:12px; }}
  thead tr:first-child th {{ background:#0a66c2; color:#fff; }}
  td.num {{ text-align:right; font-variant-numeric:tabular-nums; }}
  td.total {{ font-weight:600; background:#f9fafb; }}
  tr.gtotal td {{ background:#0a66c2; color:#fff; font-weight:700; }}
  tr:hover td:not(.gtotal) {{ background:#f9fafb; }}
  .summary {{ display:flex; gap:18px; margin-bottom:14px; flex-wrap:wrap; }}
  .summary .box {{ background:#fff; border:1px solid #e3e5e8; padding:10px 16px; border-radius:8px; }}
  .summary .v {{ font-size:18px; font-weight:700; color:#0a66c2; }}
  .summary .l {{ font-size:11px; color:#666; }}
</style>
</head>
<body>
<h1>Offene Posten Kunden</h1>
<div class="meta">
  Buchungskreis <b>{comp_code}</b> &middot; Mandant <b>{mandt}</b> &middot;
  Stichtag <b>{keydate.strftime("%d.%m.%Y")}</b> &middot;
  Erzeugt {datetime.now().strftime("%d.%m.%Y %H:%M:%S")}
</div>
<div class="summary">
  <div class="box"><div class="v">{len(items)}</div><div class="l">Posten</div></div>
  <div class="box"><div class="v">{len(pivot)}</div><div class="l">Kunden</div></div>
  <div class="box"><div class="v">{fmt(grand)}</div><div class="l">Gesamt EUR (HW)</div></div>
</div>
<table>
  <thead>
    <tr>
      <th colspan="2" style="background:#0958a8">Summe von Betrag (HW)</th>
      <th colspan="{len(labels)}">Bucket</th>
      <th></th>
    </tr>
    <tr>
      <th>Kunde</th><th>Name</th>
      {header_buckets}
      <th>Gesamtergebnis</th>
    </tr>
  </thead>
  <tbody>
    {''.join(rows)}
    {total_row}
  </tbody>
</table>
</body>
</html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)


# ---------------------------------------------------------------------------
# SD / Queue -- SD-Rechnung anlegen (Antrag in Queue-Datei, ohne RFC SDK)
# ---------------------------------------------------------------------------
def queue_sd_invoice_create(task, payload):
    """
    Schreibt eine SD-Rechnungs-Anforderung als JSON in den Queue-Ordner.
    Wird benoetigt, solange kein NW RFC SDK installiert ist - der Antrag
    kann dann manuell via LSMW oder VF01-Massenupload importiert werden.

    Erwartete Payload-Felder:
        doc_date         Datum YYYYMMDD oder 'today'
        amount           Betrag (Dezimal)
        customer         Kundennummer (10 Stellen)
        plant            Werk
        po_reference     Auftragsbezeichnung / Bestellnummer Kunde
        service_type     Leistungsart / Material
        currency         Waehrung (EUR, USD, ...)
        payment_terms    Zahlungsbedingung (z. B. 0001)
        tax_code         Steuerkennzeichen (z. B. A1)
    """
    import json
    from pathlib import Path as _P

    keydate = parse_date(payload.get("doc_date"), default_today=True)
    record = {
        "_id":            datetime.now().strftime("%Y%m%d_%H%M%S"),
        "_created":       datetime.now().isoformat(timespec="seconds"),
        "_status":        "pending",
        "doc_date":       keydate.strftime("%Y%m%d"),
        "amount":         payload.get("amount", ""),
        "customer":       payload.get("customer", ""),
        "plant":          payload.get("plant", ""),
        "po_reference":   payload.get("po_reference", ""),
        "service_type":   payload.get("service_type", ""),
        "currency":       payload.get("currency", "EUR"),
        "payment_terms":  payload.get("payment_terms", ""),
        "tax_code":       payload.get("tax_code", ""),
    }

    # Pflichtfeld-Check
    missing = [k for k in ("amount", "customer") if not record[k]]
    if missing:
        raise RuntimeError(f"Pflichtfelder fehlen: {missing}")

    queue_dir = _P(os.getenv("QUEUE_OUT_DIR", REPORT_OUT_DIR.parent / "queue"))
    queue_dir.mkdir(parents=True, exist_ok=True)
    fname = f"sd_invoice_{record['_id']}.json"
    fpath = queue_dir / fname
    with open(fpath, "w", encoding="utf-8") as f:
        json.dump(record, f, indent=2, ensure_ascii=False)

    # CSV-Sammeltabelle fortschreiben (LSMW-tauglich)
    import csv
    csv_path = queue_dir / "sd_invoices_queue.csv"
    is_new = not csv_path.exists()
    with open(csv_path, "a", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(record.keys()))
        if is_new:
            w.writeheader()
        w.writerow(record)

    return (f"SD-Rechnungsantrag aufgenommen: Kunde {record['customer']}, "
            f"Betrag {record['amount']} {record['currency']}, "
            f"Datum {record['doc_date']}. Datei: {fname}")


# ---------------------------------------------------------------------------
# SD / BAPI -- SD-Rechnung via BAPI_SALESORDER_CREATEFROMDAT2 (braucht pyrfc)
# ---------------------------------------------------------------------------
def bapi_sd_invoice_create(task, payload):
    """
    Erstellt einen SD-Auftrag (= Vorstufe einer Rechnung) via BAPI.
    Benoetigt installierten NW RFC SDK und pyrfc.

    Erwartete Payload-Felder: siehe queue_sd_invoice_create.
    Zusaetzlich optional:
        doc_type         Auftragsart, default 'TA' (Standardauftrag)
                          'L2' = Lastschriftanforderung
                          'G2' = Gutschriftanforderung
        sales_org        Verkaufsorganisation, default '1000'
        distr_chan       Vertriebsweg, default '10'
        division         Sparte, default '00'
        item_category    Positionstyp, default 'TAD' (Service)
    """
    conn = _rfc_connection()
    try:
        keydate = parse_date(payload.get("doc_date"), default_today=True)
        order_header = {
            "DOC_TYPE":   payload.get("doc_type",   "TA"),
            "SALES_ORG":  payload.get("sales_org",  "1000"),
            "DISTR_CHAN": payload.get("distr_chan", "10"),
            "DIVISION":   payload.get("division",   "00"),
            "PURCH_NO_C": payload.get("po_reference", "")[:35],
            "DOC_DATE":   keydate.strftime("%Y%m%d"),
            "CURRENCY":   payload.get("currency", "EUR"),
            "PMNTTRMS":   payload.get("payment_terms", ""),
        }
        order_partners = [{"PARTN_ROLE": "AG", "PARTN_NUMB": payload.get("customer", "")}]
        order_items = [{
            "ITM_NUMBER":  "000010",
            "MATERIAL":    payload.get("service_type", ""),
            "PLANT":       payload.get("plant", ""),
            "TARGET_QTY":  1,
            "TARGET_QU":   payload.get("unit", "LE"),
            "ITEM_CATEG":  payload.get("item_category", "TAD"),
            "TAX_CLASS1":  payload.get("tax_code", ""),
        }]
        order_conditions = [{
            "ITM_NUMBER":  "000010",
            "COND_TYPE":   "PR00",
            "COND_VALUE":  float(payload.get("amount", 0)),
            "CURRENCY":    payload.get("currency", "EUR"),
        }]
        result = conn.call(
            "BAPI_SALESORDER_CREATEFROMDAT2",
            ORDER_HEADER_IN=order_header,
            ORDER_PARTNERS=order_partners,
            ORDER_ITEMS_IN=order_items,
            ORDER_CONDITIONS_IN=order_conditions,
        )
        errs = [r for r in result.get("RETURN", []) if r["TYPE"] in ("E", "A")]
        if errs:
            raise RuntimeError("BAPI-Fehler: " + "; ".join(r["MESSAGE"] for r in errs))
        conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")
        salesdoc = result.get("SALESDOCUMENT") or "(keine Nummer)"
        return f"SD-Auftrag erstellt: Belegnummer {salesdoc} - Folgeschritt: VF01 zur Rechnung."
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# SD / BAPI -- Mehrfach-Positionen: SD-Auftrag + Sofortverrechnung (VF01)
# ---------------------------------------------------------------------------
def bapi_sd_invoice_create_multi(task, payload):
    """
    Erstellt einen SD-Auftrag mit beliebig vielen Positionen und ruft
    anschliessend BAPI_BILLINGDOC_CREATEMULTIPLE auf.

    Payload-Felder:
        customer        Kundennummer (Sold-To)
        sales_org       Verkaufsorganisation  (default ZTA)
        doc_type        Auftragsart            (default C060)
        distr_chan      Vertriebsweg           (default 10)
        division        Sparte                 (default 10)
        sales_office    Verkaufsbüro           (default 1000)
        sales_group     Verkaufsgruppe         (default 1000)
        material        Materialnummer         (default 000000000000000013)
        plant           Werk                   (default 0435)
        currency        Waehrung               (default EUR)
        payment_terms   Zahlungsbedingung      (default '0000')
        tax_code        MWSt-Kennzeichen (MWSKZ; default Z4)
        doc_date        Belegdatum YYYYMMDD    (default heute)
        items           Liste von Positionen:
            item_nr      z. B. '000010', '000020', …
            buchungstext Kurztext (max 40 Zeichen)
            betrag       Wert in Waehrung
    """
    # BCD-Bug-Workaround: _fill_bcd() belegt alle numerischen Felder mit 0 vor,
    # damit pyrfc keine leeren BCD-Strings (Decimal("")) bekommt. Kein config-Parameter nötig.
    sap_auth = payload.get("_sap_auth") or payload.get("_sap_conn")
    try:
        conn = _rfc_connection_with_auth(sap_auth)
    except Exception as _conn_exc:
        if "RFC_NO_AUTHORITY" in str(_conn_exc) and "RFCPING" in str(_conn_exc):
            log.warning(
                "Benutzer '%s' hat keine S_RFC-Berechtigung fuer RFCPING – "
                "Fallback auf Service-Account fuer RFC-Verbindung.",
                (sap_auth or {}).get("user", "?")
            )
            conn = _rfc_connection_service_account(sap_auth)
        else:
            raise
    try:
        from pyrfc import Decimal  # pyrfc-eigener Decimal-Typ (= Python decimal.Decimal)

        D0 = Decimal("0")

        # RFC-Feldtypen die KEINE numerischen Standardwerte brauchen:
        # CHAR=0, DATE=1, TIME=3, BYTE=4, TABLE=5, NULL=14,
        # ABAPOBJECT=16, STRUCTURE=17, STRING=29, XSTRING=30
        _NON_NUMERIC = frozenset([0, 1, 3, 4, 5, 14, 16, 17, 29, 30])

        def _fill_bcd(conn, rows, func_name, param_name):
            """Setzt alle numerischen Felder auf "0" (pyrfc-Bug-Workaround).
            "0" wird von pyrfc für BCD, NUM, FLOAT, INT, DECF akzeptiert.
            Wirft Exception wenn get_function_description fehlschlägt."""
            desc  = conn.get_function_description(func_name)
            # pyrfc 3.3: parameters ist eine Liste von dicts (nicht Objekte)
            param = next((p for p in desc.parameters if p['name'] == param_name), None)
            if not param:
                log.warning("_fill_bcd: Parameter %s nicht in %s gefunden", param_name, func_name)
                return rows
            # Diagnose: alle Keys im Parameter-Dict loggen
            if isinstance(param, dict):
                log.info("_fill_bcd %s.%s param-keys: %s", func_name, param_name, list(param.keys()))
            type_desc = param.get('type_description') if isinstance(param, dict) else getattr(param, 'type_description', None)
            if not type_desc:
                log.warning("_fill_bcd: %s.%s – keine type_description (param-keys: %s)",
                            func_name, param_name,
                            list(param.keys()) if isinstance(param, dict) else type(param))
                return rows
            try:
                fields_iter = list(type_desc.fields)
            except Exception as e:
                log.error("_fill_bcd: Fehler bei type_desc.fields für %s.%s: %s", func_name, param_name, e)
                return rows
            log.info("_fill_bcd %s.%s: type_desc hat %d Felder", func_name, param_name, len(fields_iter))
            num_fields = []
            for fld in fields_iter:
                ft = fld.get('field_type') if isinstance(fld, dict) else getattr(fld, 'field_type', None)
                try:
                    ft_int = int(ft)
                except (TypeError, ValueError):
                    ft_int = -1
                fname = fld['name'] if isinstance(fld, dict) else fld.name
                # Alles was NICHT ein bekannter nicht-numerischer Typ ist → "0"
                if ft_int not in _NON_NUMERIC and ft_int >= 0:
                    num_fields.append(fname)
                elif ft_int < 0:
                    # Unbekannter Typ: String-Fallback
                    ft_str = str(ft).upper()
                    if any(k in ft_str for k in ('BCD', 'NUM', 'FLOAT', 'INT', 'DECF')):
                        num_fields.append(fname)
            log.info("_fill_bcd %s.%s: %d numerische Felder → '%s'…",
                     func_name, param_name, len(num_fields),
                     ", ".join(num_fields[:5]))
            for row in rows:
                for fn in num_fields:
                    if fn not in row or row[fn] == "" or row[fn] is None:
                        row[fn] = "0"
            return rows

        _BAPI = "BAPI_SALESORDER_CREATEFROMDAT2"

        # Datum: posting_date (DD.MM.YYYY) oder doc_date (YYYYMMDD)
        raw_date = payload.get("posting_date") or payload.get("doc_date")
        keydate = parse_date(raw_date, default_today=True)
        if raw_date and "." in str(raw_date):
            parts = str(raw_date).split(".")
            keydate_str = f"{parts[2]}{parts[1]}{parts[0]}" if len(parts) == 3 else keydate.strftime("%Y%m%d")
        else:
            keydate_str = keydate.strftime("%Y%m%d")

        currency = payload.get("currency", "EUR")
        material = payload.get("material", "000000000000000013")
        plant    = payload.get("plant",    "C060")
        items_in = payload.get("items", [])

        order_header = {
            "DOC_TYPE":   payload.get("doc_type",     "ZTA"),
            "SALES_ORG":  payload.get("sales_org",    "C060"),
            "DISTR_CHAN": payload.get("distr_chan",   "10"),
            "DIVISION":   payload.get("division",     "10"),
            "SALES_OFF":  payload.get("sales_office", "1000"),
            "SALES_GRP":  payload.get("sales_group",  "100"),
            "DOC_DATE":   keydate_str,
            "CURRENCY":   currency,
            "PMNTTRMS":   payload.get("payment_terms", ""),
            "INCOTERMS1": payload.get("inco1", "FCA"),
            "INCOTERMS2": payload.get("inco2", "Germany"),
            "ORD_REASON": payload.get("augru", "Z06"),
            "PURCH_NO_C": payload.get("bstkd", "Service Level Agreement")[:35],
            "PRICE_DATE": keydate_str,
        }
        order_partners = [
            {"PARTN_ROLE": "AG", "PARTN_NUMB": str(payload.get("customer", ""))},
            {"PARTN_ROLE": "WE", "PARTN_NUMB": str(payload.get("customer", ""))},
        ]

        order_items      = []
        order_conditions = []
        order_text       = []

        for i, pos in enumerate(items_in):
            item_nr = str((i + 1) * 10).zfill(6)
            betrag  = Decimal(str(pos.get("betrag", 0) or 0))
            text    = str(pos.get("buchungstext", ""))[:40]
            taxm1   = str(pos.get("taxm1", payload.get("taxm1", "0")))
            ktgrm   = str(pos.get("ktgrm", payload.get("ktgrm", "")))

            item_row = {
                "ITM_NUMBER": item_nr,
                "MATERIAL":   material,
                "PLANT":      plant,
                "TARGET_QTY": Decimal("1"),  # BCD-Feld: Decimal
                "TARGET_QU":  "PC",
                "ITEM_CATEG": payload.get("item_category", "ZTAD"),
                "SHORT_TEXT": text,
                "CUST_MAT22": payload.get("kdmat", ""),
            }
            if taxm1:
                item_row["TAX_CLASS1"] = taxm1
            if ktgrm:
                item_row["MATL_GRP_4"] = ktgrm
            order_items.append(item_row)

            order_conditions.append({
                "ITM_NUMBER": item_nr,
                "COND_TYPE":  "PR00",
                "COND_VALUE": betrag,          # BCD-Feld: Decimal
                "CURRENCY":   currency,
                "COND_P_UNT": Decimal("1"),    # BCD-Feld: Decimal
                "COND_UNIT":  "PC",
            })
            order_text.append({
                "ITM_NUMBER": item_nr,
                "TEXT_ID":    "ARKTX",
                "LANGU":      "DE",
                "FORMAT_COL": "*",
                "TEXT_LINE":  text,
            })

        # BCD-Felder auffüllen: zuerst explizite Decimal-Werte (oben gesetzt),
        # dann _fill_bcd für alle weiteren unbekannten BCD-Felder in der Struktur
        _hdr_list        = _fill_bcd(conn, [order_header],    _BAPI, "ORDER_HEADER_IN")
        order_header     = _hdr_list[0] if _hdr_list else order_header
        order_items      = _fill_bcd(conn, order_items,      _BAPI, "ORDER_ITEMS_IN")
        order_conditions = _fill_bcd(conn, order_conditions, _BAPI, "ORDER_CONDITIONS_IN")
        order_partners   = _fill_bcd(conn, order_partners,   _BAPI, "ORDER_PARTNERS")

        # Diagnose: alle Felder mit "" oder None loggen (potenzielle BCD-Fehlerquelle)
        def _log_empty(name, rows_or_struct):
            lst = rows_or_struct if isinstance(rows_or_struct, list) else [rows_or_struct]
            for i, row in enumerate(lst):
                empties = [k for k, v in row.items() if v == "" or v is None]
                if empties:
                    log.warning("BAPI-Diagnose %s[%d] leere Felder: %s", name, i, empties)
        _log_empty("ORDER_HEADER_IN", order_header)
        _log_empty("ORDER_ITEMS_IN", order_items)
        _log_empty("ORDER_CONDITIONS_IN", order_conditions)
        _log_empty("ORDER_PARTNERS", order_partners)

        result = conn.call(
            _BAPI,
            ORDER_HEADER_IN=order_header,
            ORDER_PARTNERS=order_partners,
            ORDER_ITEMS_IN=order_items,
            ORDER_CONDITIONS_IN=order_conditions,
            ORDER_TEXT=order_text,
        )

        errs = [r for r in result.get("RETURN", []) if r["TYPE"] in ("E", "A")]
        if errs:
            conn.call("BAPI_TRANSACTION_ROLLBACK")
            raise RuntimeError("BAPI Auftrag-Fehler: " + "; ".join(r["MESSAGE"] for r in errs))

        conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")
        salesdoc = result.get("SALESDOCUMENT") or ""
        if not salesdoc:
            raise RuntimeError("BAPI hat keine Auftragsnummer zurückgegeben.")

        # Sofort verrechnen: Faktura anlegen
        billing_result = conn.call(
            "BAPI_BILLINGDOC_CREATEMULTIPLE",
            BILLINGDATAIN=[{
                "REF_DOC":      salesdoc,
                "REF_DOC_CA":   "C",
                "BILL_DATE":    keydate.strftime("%Y%m%d"),
            }],
        )
        conn.call("BAPI_TRANSACTION_COMMIT", WAIT="X")

        bill_docs = billing_result.get("BILLINGDOCUMENTITEM", [])
        invoice_nr = bill_docs[0].get("BILL_DOC", "") if bill_docs else ""

        all_msgs = result.get("RETURN", []) + billing_result.get("RETURN", [])
        return {
            "status":       "ok",
            "sales_order":  salesdoc,
            "invoice_nr":   invoice_nr,
            "result":       {"RETURN": all_msgs},
        }
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# SD / GUI -- SD-Rechnung via SAP GUI Scripting (VA01/VF01) - VBA-portierbar
# ---------------------------------------------------------------------------
def gui_sd_invoice_create(task, payload):
    """
    Erstellt eine SD-Rechnung via SAP GUI Scripting (Single-Item, Legacy).
    Haengt sich bevorzugt an eine laufende SAP-GUI-Session (VBA-Ansatz).
    """
    creds = {
        "user":     payload.get("sap_user", ""),
        "password": payload.get("sap_password", ""),
        "client":   payload.get("sap_client", ""),
    }
    use_creds = creds if any(creds.values()) else None
    safe_payload = {k: v for k, v in payload.items() if not k.startswith("sap_")}
    connection, session, should_close = _get_gui_session(creds=use_creds)
    try:
        from gui_scripts.sd_invoice import record_invoice  # lazy import
        return record_invoice(session, safe_payload)
    finally:
        if should_close:
            try:
                connection.CloseSession(session.Id)
            except Exception:
                pass


# ---------------------------------------------------------------------------
# SD / GUI -- Multi-Position SD-Rechnung (VBA-Portierung: VA01 + Faktura)
# ---------------------------------------------------------------------------
def gui_sd_invoice_create_multi(task, payload):
    """
    Erstellt eine SD-Rechnung mit beliebig vielen Positionen via SAP GUI Scripting.

    Ablauf:
        1. GUI-Session öffnen  → VA01 (Auftrag anlegen, multi-item)
        2. RFC Service-Account → BAPI_BILLINGDOC_CREATEMULTIPLE (Faktura)

    Umgeht den pyrfc 3.3 BCD-Bug bei BAPI_SALESORDER_CREATEFROMDAT2.
    Credentials kommen aus payload['_sap_auth'] (Cockpit-Anmeldefenster).
    """
    sap_auth = payload.get("_sap_auth") or payload.get("_sap_conn") or {}

    # GUI-Session: password-Key muss 'password' sein (nicht 'passwd' wie in RFC)
    creds = {
        "user":     sap_auth.get("user", ""),
        "password": sap_auth.get("passwd", ""),    # rechnung.html sendet 'passwd'
        "client":   sap_auth.get("client", ""),
        "lang":     sap_auth.get("lang", "DE"),
    }
    use_creds = creds if (creds["user"] and creds["password"]) else None

    # _sap_auth aus safe_payload entfernen (Logging-Schutz)
    safe_payload = {k: v for k, v in payload.items() if k != "_sap_auth"}

    # Schutz gegen vertauschte sales_org / doc_type aus altem localStorage-Cache
    # (alter Bug: Auftragsart-Feld enthielt "C060" und VKOrg enthielt "ZTA")
    _sorg = safe_payload.get("sales_org", "")
    _dtype = safe_payload.get("doc_type", "")
    if _sorg in ("ZTA", "ZG2") and _dtype not in ("ZTA", "ZG2"):
        log.warning(
            "Payload-Korrektur: sales_org='%s' und doc_type='%s' waren vertauscht "
            "(veralteter localStorage-Cache). Tausche → sales_org='%s', doc_type='%s'.",
            _sorg, _dtype, _dtype or "C060", _sorg,
        )
        safe_payload = dict(safe_payload)
        safe_payload["sales_org"] = _dtype or "C060"
        safe_payload["doc_type"]  = _sorg
    # Werk-Korrektur: 0435 ist falsch für C060-Aufträge (VBA-Referenz: C060)
    if safe_payload.get("plant") == "0435":
        log.warning("Payload-Korrektur: plant='0435' → 'C060' (veralteter Cache).")
        safe_payload = dict(safe_payload)
        safe_payload["plant"] = "C060"

    # ── Schritt 1: GUI → VA01 anlegen ────────────────────────────────────────
    # G-5: Session auf dem System des angemeldeten Benutzers (ashost aus _sap_auth)
    target_ashost = sap_auth.get("ashost", "")
    debug_mode    = bool(payload.get("debug", False))
    background    = bool(payload.get("background", True))   # Standard: Hintergrund

    if sap_auth.get("user") and sap_auth.get("passwd"):
        connection, session = _gui_session_with_sap_auth(sap_auth)
        should_close = True
        if background:
            _hide_session_window(session)
    else:
        connection, session, should_close = _get_gui_session(
            creds=use_creds, target_ashost=target_ashost
        )
    gui_result = {}
    gui_error  = None
    try:
        from gui_scripts.sd_invoice import record_invoice_multi, _StepTracer  # lazy import
        tracer     = _StepTracer(session)
        gui_result = record_invoice_multi(session, safe_payload, tracer=tracer)
    except Exception as exc:
        gui_error = str(exc)
        # Schritte aus dem Tracer holen (falls schon befüllt)
        try:
            steps = tracer.steps
        except NameError:
            steps = []
        # Im Debug-Modus: Schrittprotokoll zurückgeben, Fehler nicht weiterwerfen
        if debug_mode:
            return {
                "status":  "debug_error",
                "error":   gui_error,
                "steps":   steps,
            }
        raise
    finally:
        if should_close:
            try:
                connection.CloseSession(session.Id)
            except Exception:
                pass

    steps       = gui_result.get("steps", [])
    sales_order = gui_result.get("sales_order", "")
    invoice_nr  = gui_result.get("invoice_nr", "")

    # Im Debug-Modus: Schrittprotokoll zurückgeben, ohne Faktura anzulegen
    if debug_mode:
        return {
            "status":      "debug_ok",
            "sales_order": sales_order,
            "steps":       steps,
        }

    if not sales_order:
        raise RuntimeError(
            f"GUI-Schritt fehlgeschlagen – keine Auftragsnummer. Detail: {gui_result}"
        )

    # GUI hat Faktura bereits angelegt → BAPI-Fallback überspringen
    if invoice_nr:
        log.info("Faktura: SO %s → Rechnung %s (GUI)", sales_order, invoice_nr)
        return {
            "status":      "ok",
            "sales_order": sales_order,
            "invoice_nr":  invoice_nr,
            "result":      {"RETURN": []},
            "steps":       steps,
        }

    # ── Kein BAPI-Fallback für Faktura-Erstellung ────────────────────────────
    # BAPI_BILLINGDOC_CREATEMULTIPLE läuft unter Service-Account RFC_COFACE,
    # nicht unter dem angemeldeten SAP-GUI-Benutzer (HUEMMKMA).
    # Faktura-Erstellung MUSS über SAP GUI (VF01) erfolgen.
    # Wenn GUI-Billing leer → Fehlermeldung zurückgeben, kein BAPI-Fallback.
    log.warning(
        "Faktura: SO %s → GUI-Billing ohne Ergebnis (Faktura-Nr leer). "
        "BAPI-Fallback deaktiviert (würde falschen Benutzer RFC_COFACE verwenden).",
        sales_order,
    )
    return {
        "status":      "error",
        "sales_order": sales_order,
        "invoice_nr":  "",
        "result":      {"RETURN": [{"TYPE": "E", "MESSAGE": "GUI-Faktura-Erstellung fehlgeschlagen – bitte SAP-Log prüfen"}]},
        "steps":       steps,
    }


def gui_sd_invoice_from_excel(task, payload):
    """
    Batch-Rechnungserstellung aus Excel-Control-Sheet.

    Akzeptiert entweder:
      a) 'filepath' → liest Excel-Datei serverseitig via parse_control_sheet()
      b) 'rows'     → nimmt vorgelesene Zeilen (client-seitig geparst via SheetJS)

    Format der Zeilen in 'rows' (aus rechnung.html / SheetJS):
        customer     str   Kunden-Nr (Spalte A)
        name         str   Kundenname (Spalte B, info only)
        datum        str   MM/YYYY (Spalte C) → wird in letzten Tag konvertiert
        buchungstext str   Positionstext (Spalte D)
        betrag       float Betrag (Spalte E)
        leistungsart str   Leistungsart (Spalte F)
        currency     str   Währung (Spalte G, default EUR)
        zterm        str   Zahlungsbedingung (Spalte H, default 0001)
        mwst         int   MWSt-Code (Spalte I)
        invoice_nr   str   Vorhandene Nr. (Spalte J)
        status       str   Status (Spalte K) → 'ready created' → skip
        row          int   Zeilennummer (1-basiert, für Protokoll)

    Pflicht:
        _sap_auth    SAP-Credentials (wie bei GUI_SD_INVOICE_MULTI)

    Optional:
        debug    bool  → gibt nur ersten Nicht-Skip-Lauf zurück, bucht nicht

    Rückgabe:
        {
          'status':   'ok' | 'partial' | 'error'
          'total':    int
          'ok':       int
          'skipped':  int
          'errors':   int
          'results':  [{'row', 'customer', 'buchungstext', 'sales_order',
                         'error', 'skipped', 'steps'}]
        }
    """
    from gui_scripts.sd_invoice import (
        parse_control_sheet,
        record_invoice_from_control_sheet,
        _last_day_of_month,
        SKIP_STATUS,
    )

    sap_auth   = payload.get("_sap_auth") or payload.get("_sap_conn") or {}
    debug_mode = bool(payload.get("debug", False))

    # ── Zeilen beschaffen ────────────────────────────────────────────────────
    rows_raw = payload.get("rows")
    if rows_raw:
        # Vorgelesene Zeilen aus dem Browser (SheetJS) normalisieren
        rows: list[dict] = []
        for i, r in enumerate(rows_raw):
            customer = str(r.get("customer") or r.get("kunden_nr") or "").strip()
            if not customer:
                continue
            datum = str(r.get("datum") or "").strip()
            status_str = str(r.get("status") or "").strip().lower()
            mwst_raw = r.get("mwst") or r.get("mwst_code") or 0
            try:
                mwst = int(float(mwst_raw))
            except (ValueError, TypeError):
                mwst = 0
            try:
                betrag = float(r.get("betrag") or 0)
            except (ValueError, TypeError):
                betrag = 0.0

            row_dict = {
                "row":          r.get("row") or (i + 5),   # Standardmäßig ab Zeile 5
                "customer":     customer,
                "name":         str(r.get("name") or ""),
                "datum":        datum,
                "buchungstext": str(r.get("buchungstext") or "")[:40],
                "betrag":       betrag,
                "leistungsart": str(r.get("leistungsart") or "Management-Service"),
                "currency":     str(r.get("currency") or r.get("waehrung") or "EUR"),
                "zterm":        str(r.get("zterm") or r.get("payment_terms") or "0001"),
                "mwst":         mwst,
                "invoice_nr":   str(r.get("invoice_nr") or ""),
                "status":       str(r.get("status") or ""),
                "skip":         status_str in SKIP_STATUS,
                "posting_date": _last_day_of_month(datum) if datum else "",
            }
            rows.append(row_dict)
    elif payload.get("filepath"):
        rows = parse_control_sheet(payload["filepath"])
    else:
        raise ValueError("Entweder 'rows' (JSON-Array) oder 'filepath' muss angegeben werden.")

    if not rows:
        return {"status": "ok", "total": 0, "ok": 0, "skipped": 0, "errors": 0, "results": []}

    # ── GUI-Session aufbauen ────────────────────────────────────────────────
    creds = {
        "user":     sap_auth.get("user", ""),
        "password": sap_auth.get("passwd", ""),
        "client":   sap_auth.get("client", ""),
        "lang":     sap_auth.get("lang", "DE"),
    }
    use_creds     = creds if (creds["user"] and creds["password"]) else None
    target_ashost = sap_auth.get("ashost", "")

    connection, session, should_close = _get_gui_session(
        creds=use_creds, target_ashost=target_ashost
    )
    try:
        run_rows = rows
        if debug_mode:
            # Debug: nur erster Nicht-Skip-Lauf
            run_rows = [r for r in rows if not r.get("skip")][:1]
        results = record_invoice_from_control_sheet(session, run_rows)
    finally:
        if should_close:
            try:
                connection.CloseSession(session.Id)
            except Exception:
                pass

    ok_count   = sum(1 for r in results if r.get("sales_order") and not r.get("error"))
    skip_count = sum(1 for r in results if r.get("skipped"))
    err_count  = sum(1 for r in results if r.get("error"))

    if err_count == 0:
        status = "ok"
    elif ok_count > 0:
        status = "partial"
    else:
        status = "error"

    return {
        "status":  status,
        "total":   len(results),
        "ok":      ok_count,
        "skipped": skip_count,
        "errors":  err_count,
        "results": results,
    }


def sql_ping(task: dict, payload: dict) -> str:
    """Health-Check der ASE-DB-Verbindung."""
    return sap_db.ping()


# ---------------------------------------------------------------------------
# MM / SQL -- Materialbelegliste MB51 aus Sybase ASE (MKPF + MSEG)
# ---------------------------------------------------------------------------
def sql_mb51(task: dict, payload: dict) -> str:
    """
    Liest Materialbelege (Warenbewegungen) direkt aus MKPF + MSEG via Sybase ASE.
    Entspricht dem Inhalt der SAP-Transaktion MB51.

    Payload-Parameter (alle optional):
        date_from    YYYYMMDD  Buchungsdatum von  (default: 01.01. aktuelles Jahr)
        date_to      YYYYMMDD  Buchungsdatum bis  (default: heute)
        werks        Werk-Filter (leer = alle)
        matnr        Material-Filter, Teilstring moeglich (leer = alle)
        bwart        Bewegungsart-Filter (leer = alle)
        out_path     Ausgabepfad  (default: C:\\DEV\\Tariff-Database\\MB51_<ts>.xlsx)
        max_rows     Zeilengrenze (default: 200000)
    """
    today = datetime.today()
    date_from = payload.get("date_from") or f"{today.year}0101"
    date_to   = payload.get("date_to")   or today.strftime("%Y%m%d")
    werks     = (payload.get("werks") or "").strip()
    matnr     = (payload.get("matnr") or "").strip()
    bwart     = (payload.get("bwart") or "").strip()
    max_rows  = int(payload.get("max_rows", 200000))
    mandt     = os.getenv("SAP_CLIENT", "600")
    schema    = os.getenv("SAP_DB_SCHEMA", "SAPSR3")
    pfx       = f"{schema}." if schema else ""

    ts       = today.strftime("%Y%m%d_%H%M%S")
    out_path = payload.get("out_path") or os.path.join(
        r"C:\DEV\Tariff-Database", f"MB51_{ts}.xlsx"
    )

    log.info("MB51 SQL: %s–%s  Werk=%s  Material=%s  BewArt=%s  Schema=%s",
             date_from, date_to, werks or "*", matnr or "*", bwart or "*", schema)

    sql = f"""
        SELECT TOP {max_rows}
            h.MBLNR   AS Belegnummer,
            h.MJAHR   AS Jahr,
            h.BUDAT   AS Buchungsdatum,
            h.BLDAT   AS Belegdatum,
            h.CPUDT   AS Erfassungsdatum,
            h.USNAM   AS Benutzer,
            h.TCODE   AS Transaktion,
            h.XBLNR   AS Referenzbeleg,
            h.BKTXT   AS Belegkopftext,
            i.ZEILE   AS Position,
            i.BWART   AS Bewegungsart,
            i.MATNR   AS Material,
            i.WERKS   AS Werk,
            i.LGORT   AS Lagerort,
            i.CHARG   AS Charge,
            i.MENGE   AS Menge,
            i.MEINS   AS Mengeneinheit,
            i.DMBTR   AS Betrag_HW,
            i.WAERS   AS Waehrung,
            i.EBELN   AS Bestellnummer,
            i.EBELP   AS Bestellposition,
            i.LIFNR   AS Lieferant,
            i.KUNNR   AS Kunde,
            i.KOSTL   AS Kostenstelle,
            i.AUFNR   AS Auftrag,
            i.SGTXT   AS Positionstext
        FROM {pfx}MKPF h
        JOIN {pfx}MSEG i
          ON h.MANDT = i.MANDT
         AND h.MBLNR = i.MBLNR
         AND h.MJAHR = i.MJAHR
        WHERE h.MANDT = ?
          AND h.BUDAT BETWEEN ? AND ?
    """
    params: list = [mandt, date_from, date_to]
    if werks:
        sql += "  AND i.WERKS = ?\n"; params.append(werks)
    if matnr:
        sql += "  AND i.MATNR LIKE ?\n"; params.append(f"%{matnr}%")
    if bwart:
        sql += "  AND i.BWART = ?\n"; params.append(bwart)
    sql += "  ORDER BY h.BUDAT DESC, h.MBLNR, i.ZEILE"

    conn = sap_db.get_connection()
    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        columns = [desc[0] for desc in cur.description]
        rows = cur.fetchall()
    finally:
        conn.close()

    log.info("MB51: %d Zeilen gelesen.", len(rows))

    # --- Excel ---
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl fehlt: pip install openpyxl") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "MB51"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    date_cols = {"Buchungsdatum", "Belegdatum", "Erfassungsdatum"}
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            if columns[ci - 1] in date_cols and val:
                try:
                    val = datetime.strptime(str(val).strip(), "%Y%m%d").strftime("%d.%m.%Y")
                except (ValueError, TypeError):
                    pass
            ws.cell(row=ri, column=ci, value=val)

    for ci, col in enumerate(columns, 1):
        sample = [str(ws.cell(row=r, column=ci).value or "") for r in range(2, min(len(rows)+2, 52))]
        w = max(len(col), *(len(s) for s in sample), 8)
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 2, 40)

    ws2 = wb.create_sheet("Abfrage-Info")
    for label, val in [
        ("Datum von", date_from), ("Datum bis", date_to),
        ("Werk", werks or "(alle)"), ("Material", matnr or "(alle)"),
        ("Bewegungsart", bwart or "(alle)"), ("Mandant", mandt),
        ("Schema", schema), ("Zeilen gesamt", len(rows)),
        ("Erstellt am", today.strftime("%d.%m.%Y %H:%M:%S")),
    ]:
        ws2.append([label, val])

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    log.info("Gespeichert: %s", out_path)

    return (f"MB51-Export: {len(rows)} Warenbewegungen "
            f"({date_from}–{date_to}) gespeichert unter {out_path}")


# ---------------------------------------------------------------------------
# AP / GUI -- Zahllauf F110
# ---------------------------------------------------------------------------
def gui_f110_payment_run(task: dict, payload: dict) -> str:
    connection, session, should_close = _get_gui_session()
    try:
        session.findById("wnd[0]/tbar[0]/okcd").text = "/nF110"
        session.findById("wnd[0]").sendVKey(0)
        today = payload.get("run_date", datetime.now().strftime("%d.%m.%Y"))
        ident = payload.get("ident", datetime.now().strftime("R%H%M"))
        session.findById("wnd[0]/usr/ctxtREGUTL-LAUFD").text = today
        session.findById("wnd[0]/usr/ctxtREGUTL-LAUFI").text = ident
        session.findById("wnd[0]/usr/tabsTS_BANK/tabpPARA").select()
        para = session.findById("wnd[0]/usr/tabsTS_BANK/tabpPARA/ssubPAGE:SAPF110V:1102")
        para.findById("ctxtREGUV-ZAHLT").text = payload.get("pay_date", today)
        session.findById("wnd[0]/tbar[1]/btn[15]").press()    # Vorschlag
        import time as _time
        _time.sleep(int(payload.get("proposal_wait_sec", 30)))
        session.findById("wnd[0]/tbar[1]/btn[17]").press()    # Durchfuehrung
        return f"F110 angestossen (Lauf {today}/{ident}). Status in SM37 verifizieren."
    finally:
        if should_close:
            try:
                connection.CloseSession(session.Id)
            except Exception:
                pass


# ---------------------------------------------------------------------------
# AA / Batch -- Abschreibung AFAB
# ---------------------------------------------------------------------------
def batch_afab_depreciation(task: dict, payload: dict) -> str:
    log.info("AFAB Lauf wuerde geplant: %s", payload)
    return "AFAB-Job eingeplant (Skelett - echte Variante in Produktion ergaenzen)."


# ---------------------------------------------------------------------------
# Diagnose: RFC_PING
# ---------------------------------------------------------------------------
def rfc_ping(task: dict, payload: dict) -> str:
    conn = _rfc_connection()
    try:
        conn.call("RFC_PING")
        return "RFC_PING OK"
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Stammdaten – Debitoren / Kreditoren / GL-Konten / Kostenstellen
# ---------------------------------------------------------------------------
def _parse_rfc_table(result: dict) -> list:
    """RFC_READ_TABLE-Ergebnis anhand der FIELDS-Offsets parsen."""
    fields = result.get("FIELDS", [])
    records = []
    for row in result.get("DATA", []):
        wa = row.get("WA", "")
        rec = {}
        for f in fields:
            name = f.get("FIELDNAME", "")
            try:
                offset = int(f.get("OFFSET", 0))
                length = int(f.get("LENGTH", 0))
                rec[name] = wa[offset:offset + length].strip() if length > 0 else ""
            except (ValueError, TypeError):
                rec[name] = ""
        records.append(rec)
    return records


def _rfc_read(conn, table: str, fields: list, options: list = None, maxrows: int = 1000) -> list:
    """Kurz-Wrapper um RFC_READ_TABLE."""
    kwargs = dict(
        QUERY_TABLE=table,
        FIELDS=[{"FIELDNAME": f} for f in fields],
        ROWCOUNT=maxrows,
    )
    if options:
        kwargs["OPTIONS"] = options
    result = conn.call("RFC_READ_TABLE", **kwargs)
    return _parse_rfc_table(result)


def bapi_stammdaten(task: dict, payload: dict) -> dict:
    """
    Liest SAP-Stammdaten via RFC_READ_TABLE.
    payload.entity: 'debitoren' | 'kreditoren' | 'konten' | 'kostenstellen'
    payload.comp_code: Buchungskreis (default '0435')
    payload.kokrs: Kostenrechnungskreis für KST (default = comp_code)
    payload.maxrows: max. Zeilen je Tabelle (default 1000)
    """
    sap_auth  = payload.get("_sap_auth")
    entity    = payload.get("entity", "debitoren")
    comp_code = payload.get("comp_code", "0435")
    kokrs     = payload.get("kokrs") or comp_code
    maxrows   = int(payload.get("maxrows", 1000))
    conn = _rfc_connection_with_auth(sap_auth)

    try:
        if entity == "debitoren":
            # Buchungskreisdaten
            knb1 = _rfc_read(conn, "KNB1",
                ["KUNNR","BUKRS","AKONT","ZTERM"],
                [{"TEXT": f"BUKRS EQ '{comp_code}'"}], maxrows)
            kunnr_set = {r["KUNNR"] for r in knb1 if r.get("KUNNR")}
            # Allgemeine Adressdaten
            kna1_raw = _rfc_read(conn, "KNA1",
                ["KUNNR","NAME1","ORT01","LAND1","TELF1"], [], maxrows * 2)
            kna1 = {r["KUNNR"]: r for r in kna1_raw if r.get("KUNNR") in kunnr_set}
            records = []
            for r in knb1:
                n = kna1.get(r["KUNNR"], {})
                records.append({
                    "nr": r["KUNNR"], "name": n.get("NAME1",""),
                    "city": n.get("ORT01",""), "land": n.get("LAND1",""),
                    "tel": n.get("TELF1",""),
                    "akont": r.get("AKONT",""), "zterm": r.get("ZTERM",""),
                })

        elif entity == "kreditoren":
            lfb1 = _rfc_read(conn, "LFB1",
                ["LIFNR","BUKRS","AKONT","ZTERM"],
                [{"TEXT": f"BUKRS EQ '{comp_code}'"}], maxrows)
            lifnr_set = {r["LIFNR"] for r in lfb1 if r.get("LIFNR")}
            lfa1_raw = _rfc_read(conn, "LFA1",
                ["LIFNR","NAME1","ORT01","LAND1","TELF1"], [], maxrows * 2)
            lfa1 = {r["LIFNR"]: r for r in lfa1_raw if r.get("LIFNR") in lifnr_set}
            records = []
            for r in lfb1:
                n = lfa1.get(r["LIFNR"], {})
                records.append({
                    "nr": r["LIFNR"], "name": n.get("NAME1",""),
                    "city": n.get("ORT01",""), "land": n.get("LAND1",""),
                    "tel": n.get("TELF1",""),
                    "akont": r.get("AKONT",""), "zterm": r.get("ZTERM",""),
                })

        elif entity == "konten":
            skb1 = _rfc_read(conn, "SKB1",
                ["SAKNR","BUKRS","XSPEA","XLOEV","WAERS"],
                [{"TEXT": f"BUKRS EQ '{comp_code}'"}], maxrows)
            saknr_set = {r["SAKNR"] for r in skb1 if r.get("SAKNR")}
            ska1_raw = _rfc_read(conn, "SKA1",
                ["SAKNR","KTOPL","TXT20","TXT50","KTOKS","XBILK"], [],                ["SAKNR","KTOPL","TXT20","TXT50","KTOKS","XBILK"], [],
                maxrows * 2)
            ska1 = {r["SAKNR"]: r for r in ska1_raw if r.get("SAKNR") in saknr_set}
            records = []
            for r in skb1:
                n = ska1.get(r["SAKNR"], {})
                records.append({
                    "nr":    r["SAKNR"],
                    "name":  n.get("TXT20", "") or n.get("TXT50", ""),
                    "name50": n.get("TXT50", ""),
                    "ktoks": n.get("KTOKS", ""),
                    "xbilk": n.get("XBILK", ""),
                    "bukrs": r.get("BUKRS", ""),
                    "xspea": r.get("XSPEA", ""),
                    "xloev": r.get("XLOEV", ""),
                    "waers": r.get("WAERS", ""),
                })

        elif entity == "kostenstellen":
            csks = _rfc_read(conn, "CSKS",
                ["KOSTL","KOKRS","DATBI","DATAB","VERAK","ABTEI","MCCOA"],
                [{"TEXT": f"KOKRS EQ '{kokrs}'"}], maxrows)
            cskt_raw = _rfc_read(conn, "CSKT",
                ["KOSTL","KOKRS","SPRAS","KTEXT"],
                [{"TEXT": f"KOKRS EQ '{kokrs}' AND SPRAS EQ 'D'"}], maxrows * 2)
            cskt = {r["KOSTL"]: r for r in cskt_raw if r.get("KOSTL")}
            records = []
            for r in csks:
                t = cskt.get(r["KOSTL"], {})
                records.append({
                    "nr":    r["KOSTL"],
                    "name":  t.get("KTEXT", ""),
                    "kokrs": r.get("KOKRS", ""),
                    "verak": r.get("VERAK", ""),
                    "abtei": r.get("ABTEI", ""),
                    "datbi": r.get("DATBI", ""),
                    "datab": r.get("DATAB", ""),
                })
        else:
            records = []

        return {"status": "ok", "entity": entity, "records": records, "count": len(records)}

    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Handler-Registry  (method, tcode) -> function
# ---------------------------------------------------------------------------
def gui_va42_zterm_change(task: dict, payload: dict) -> dict:
    """
    Ändert ZTERM in Kundenverträgen via VA42 (SAP GUI Scripting).
    payload: { contracts: [vbeln, ...], new_zterm: "X004", _sap_auth: {...} }
    Nutzt laufende SAP-GUI-Session (G-5).
    """
    import time
    sap_auth   = payload.get("_sap_auth") or {}
    contracts  = payload.get("contracts", [])
    new_zterm  = str(payload.get("new_zterm", "")).strip()
    background = bool(payload.get("background", True))   # Standard: Hintergrund

    if not contracts:
        return {"status": "error", "message": "Keine Vertragsnummern übergeben"}
    if not new_zterm:
        return {"status": "error", "message": "new_zterm fehlt"}

    target_ashost = sap_auth.get("ashost", "")

    # Neue Session mit Cockpit-Credentials (bevorzugt), sonst laufende Session
    if sap_auth.get("user") and sap_auth.get("passwd"):
        connection, session = _gui_session_with_sap_auth(sap_auth)
        should_close = True
        log.info("VA42: Neue SAP GUI Session geöffnet für User=%s (background=%s)",
                 sap_auth.get("user"), background)
        if background:
            _hide_session_window(session)
    else:
        connection, session, should_close = _get_gui_session(target_ashost=target_ashost)
        log.info("VA42: Laufende SAP GUI Session verwendet")

    results = []
    try:
        for vbeln in contracts:
            vbeln = str(vbeln).strip()
            if not vbeln:
                continue
            step = {"vbeln": vbeln, "status": "ok", "message": ""}
            try:
                # /nVA42 aufrufen
                session.findById("wnd[0]/tbar[0]/okcd").text = "/nVA42"
                session.findById("wnd[0]").sendVKey(0)
                time.sleep(1.5)

                # Vertragsnummer eingeben
                session.findById("wnd[0]/usr/ctxtKEAD-VBELN").text = vbeln
                session.findById("wnd[0]").sendVKey(0)
                time.sleep(2.0)

                # ZTERM-Feld suchen und setzen
                try:
                    zf = session.findAllByName("VBKD-ZTERM")
                    if zf and zf.count > 0:
                        zf.elementAt(0).text = new_zterm
                    else:
                        step["status"]  = "warn"
                        step["message"] = "VBKD-ZTERM Feld nicht gefunden"
                        results.append(step)
                        continue
                except Exception as e:
                    step["status"]  = "warn"
                    step["message"] = f"ZTERM setzen: {e}"
                    results.append(step)
                    continue

                # Sichern (F11)
                session.findById("wnd[0]").sendVKey(11)
                time.sleep(1.0)
                step["message"] = f"ZTERM={new_zterm} gesetzt und gesichert"
            except Exception as e:
                step["status"]  = "error"
                step["message"] = str(e)
            results.append(step)

        ok    = sum(1 for r in results if r["status"] == "ok")
        warn  = sum(1 for r in results if r["status"] == "warn")
        err   = sum(1 for r in results if r["status"] == "error")
        log.info("gui_va42_zterm_change: %d ok, %d warn, %d err", ok, warn, err)
        return {
            "status":   "ok" if err == 0 else "partial",
            "ok":       ok,
            "warn":     warn,
            "errors":   err,
            "results":  results,
        }
    finally:
        if should_close and connection:
            try:
                connection.CloseSession(session.Id)
            except Exception:
                pass


def gui_xd02_kzterm_change(task: dict, payload: dict) -> dict:
    """
    Ändert KNB1-ZTERM (Buchungskreisdaten → Reiter 'Payment Transactions') via XD02.
    payload: { customers: [kunnr, ...], bukrs: "0439", new_zterm: "X004", _sap_auth: {...} }
    """
    import time
    sap_auth   = payload.get("_sap_auth") or {}
    customers  = payload.get("customers", [])
    bukrs      = str(payload.get("bukrs", "0439")).strip()
    new_zterm  = str(payload.get("new_zterm", "")).strip()
    background = bool(payload.get("background", True))   # Standard: Hintergrund

    if not customers:
        return {"status": "error", "message": "Keine Kundennummern übergeben"}
    if not new_zterm:
        return {"status": "error", "message": "new_zterm fehlt"}

    target_ashost = sap_auth.get("ashost", "")

    # Neue Session mit Cockpit-Credentials (bevorzugt), sonst laufende Session
    if sap_auth.get("user") and sap_auth.get("passwd"):
        connection, session = _gui_session_with_sap_auth(sap_auth)
        should_close = True
        log.info("XD02: Neue SAP GUI Session geöffnet für User=%s (background=%s)",
                 sap_auth.get("user"), background)
        if background:
            _hide_session_window(session)
    else:
        connection, session, should_close = _get_gui_session(target_ashost=target_ashost)
        log.info("XD02: Laufende SAP GUI Session verwendet")

    results = []
    try:
        for kunnr in customers:
            kunnr = str(kunnr).strip()
            if not kunnr:
                continue
            step = {"kunnr": kunnr, "status": "ok", "message": ""}
            try:
                # XD02 aufrufen
                session.findById("wnd[0]/tbar[0]/okcd").text = "/nXD02"
                session.findById("wnd[0]").sendVKey(0)
                time.sleep(1.5)
                # XD02 Einstiegsmaske: Popup erscheint als wnd[1]
                import time as _t
                _t.sleep(0.5)

                # Kundennummer im Popup (wnd[1]) eintragen
                kunnr_entered = False
                for fid in (
                    "wnd[1]/usr/ctxtRF02D-KUNNR",
                    "wnd[0]/usr/ctxtRF02D-KUNNR",
                    "wnd[1]/usr/ctxtKUNA-KUNNR",
                    "wnd[0]/usr/ctxtKUNA-KUNNR",
                ):
                    try:
                        session.findById(fid).text = kunnr
                        kunnr_entered = True
                        break
                    except Exception:
                        pass
                if not kunnr_entered:
                    step["status"] = "error"
                    step["message"] = "Kunden-Feld nicht gefunden"
                    results.append(step)
                    continue

                # Buchungskreis im Popup eintragen
                for fid in (
                    "wnd[1]/usr/ctxtRF02D-BUKRS",
                    "wnd[0]/usr/ctxtRF02D-BUKRS",
                    "wnd[1]/usr/ctxtKUNA-BUKRS",
                    "wnd[0]/usr/ctxtKUNA-BUKRS",
                ):
                    try:
                        session.findById(fid).text = bukrs
                        break
                    except Exception:
                        pass

                # Sales Area Felder leeren (koennten vorbelegt sein aus letzter Session)
                # Wir aendern nur Buchungskreisdaten (KNB1), Sales Area muss leer bleiben
                for fid in (
                    "wnd[1]/usr/ctxtRF02D-VKORG",  # Sales Organisation
                    "wnd[0]/usr/ctxtRF02D-VKORG",
                    "wnd[1]/usr/ctxtRF02D-VTWEG",  # Vertriebsweg (Distribution Channel)
                    "wnd[0]/usr/ctxtRF02D-VTWEG",
                    "wnd[1]/usr/ctxtRF02D-SPART",  # Sparte (Division)
                    "wnd[0]/usr/ctxtRF02D-SPART",
                ):
                    try:
                        session.findById(fid).text = ""
                    except Exception:
                        pass

                # Popup mit Enter bestätigen (wnd[1] oder wnd[0])
                for wnd in ("wnd[1]", "wnd[0]"):
                    try:
                        session.findById(wnd).sendVKey(0)
                        break
                    except Exception:
                        pass
                _t.sleep(2.5)

                # Falls weiterer Dialog erscheint: auch bestätigen
                try:
                    session.findById("wnd[1]").sendVKey(0)
                    _t.sleep(1.0)
                except Exception:
                    pass

                # Explizit Tab "Payment Transactions" (2. Tab) anwaehlen
                for tab_id in (
                    "wnd[0]/usr/tabsTABSTRIP1/tabpT02",
                    "wnd[0]/usr/tabsTABSTRIP1/tabpZAHL",
                    "wnd[0]/usr/tabsTABSTRIP1/tabpT" + chr(92) + "02",
                ):
                    try:
                        session.findById(tab_id).select()
                        _t.sleep(1.0)
                        log.info("XD02: Tab gewaehlt: %s", tab_id)
                        break
                    except Exception:
                        pass

                # KNB1-ZTERM: alten Wert lesen, dann neuen setzen
                zterm_old = ""
                zterm_set = False

                # Container-Objekte fuer findAllByName
                containers = []
                for cpath in ("wnd[0]", "wnd[0]/usr"):
                    try:
                        containers.append(session.findById(cpath))
                    except Exception:
                        pass
                containers.insert(0, session)

                # Zuerst alten ZTERM-Wert lesen
                for container in containers:
                    if zterm_old:
                        break
                    for fname in ("KNB1-ZTERM", "ZTERM"):
                        if zterm_old:
                            break
                        for ftype in ("GuiCTextField", "GuiTextField"):
                            try:
                                zf = container.findAllByName(fname, ftype)
                                if zf and zf.count > 0:
                                    zterm_old = (zf.elementAt(0).text or "").strip()
                                    break
                            except Exception:
                                pass

                if not zterm_old:
                    for fp in (
                        "wnd[0]/usr/tabsTABSTRIP1/tabpT02/ssubSUBSCR_BODY:SAPMF02D:7215/ctxtKNB1-ZTERM",
                        "wnd[0]/usr/tabsTABSTRIP1/tabpT02/ssubSUBSCR_BODY:SAPMF02D:7215/txtKNB1-ZTERM",
                        "wnd[0]/usr/subSUBSCR_BODY:SAPMF02D:7215/ctxtKNB1-ZTERM",
                        "wnd[0]/usr/ctxtKNB1-ZTERM",
                        "wnd[0]/usr/txtKNB1-ZTERM",
                    ):
                        try:
                            zterm_old = (session.findById(fp).text or "").strip()
                            if zterm_old:
                                break
                        except Exception:
                            pass

                step["zterm_old"] = zterm_old
                log.info("XD02 ZTERM alt: '%s' -> neu: '%s'", zterm_old, new_zterm)

                # Neuen ZTERM setzen
                for container in containers:
                    if zterm_set:
                        break
                    for fname in ("KNB1-ZTERM", "ZTERM"):
                        if zterm_set:
                            break
                        for ftype in ("GuiCTextField", "GuiTextField"):
                            try:
                                zf = container.findAllByName(fname, ftype)
                                if zf and zf.count > 0:
                                    elem = zf.elementAt(0)
                                    elem.setFocus()
                                    elem.text = new_zterm
                                    zterm_set = True
                                    log.info("XD02 ZTERM gesetzt via findAllByName(%s,%s)", fname, ftype)
                                    break
                            except Exception:
                                pass

                # Fallback: findById mit allen bekannten Pfadvarianten
                if not zterm_set:
                    for fp in (
                        "wnd[0]/usr/tabsTABSTRIP1/tabpT02/ssubSUBSCR_BODY:SAPMF02D:7215/ctxtKNB1-ZTERM",
                        "wnd[0]/usr/tabsTABSTRIP1/tabpT02/ssubSUBSCR_BODY:SAPMF02D:7215/txtKNB1-ZTERM",
                        "wnd[0]/usr/subSUBSCR_BODY:SAPMF02D:7215/ctxtKNB1-ZTERM",
                        "wnd[0]/usr/subSUBSCR_BODY:SAPMF02D:7215/txtKNB1-ZTERM",
                        "wnd[0]/usr/ctxtKNB1-ZTERM",
                        "wnd[0]/usr/txtKNB1-ZTERM",
                    ):
                        try:
                            f = session.findById(fp)
                            f.setFocus()
                            f.text = new_zterm
                            zterm_set = True
                            log.info("XD02 ZTERM gesetzt via findById: %s", fp)
                            break
                        except Exception:
                            pass

                # Letzter Fallback: WScript.Shell.SendKeys (tippt in fokussiertes Feld)
                if not zterm_set:
                    try:
                        import win32com.client as _w32
                        shell = _w32.Dispatch("WScript.Shell")
                        # Fokus auf SAP-Fenster
                        session.findById("wnd[0]").setFocus()
                        _t.sleep(0.3)
                        # Feld via Tab-Navigation ansteuern: Ctrl+F -> Payment Transactions
                        # Einfacher: Feld-Inhalt loeschen und neu tippen
                        shell.SendKeys("^a")    # Alles markieren
                        _t.sleep(0.1)
                        shell.SendKeys("{DEL}") # Loeschen
                        _t.sleep(0.1)
                        for ch in new_zterm:
                            shell.SendKeys(ch)
                            _t.sleep(0.05)
                        zterm_set = True
                        log.info("XD02 ZTERM via SendKeys gesetzt: %s", new_zterm)
                    except Exception as sk_err:
                        log.warning("XD02 SendKeys fehlgeschlagen: %s", sk_err)

                if not zterm_set:
                    step["status"] = "warn"
                    step["message"] = "KNB1-ZTERM Feld nicht setzbar (alle Methoden fehlgeschlagen)"
                    results.append(step)
                    continue

                # Direkt sichern (kein Enter vorher – verhindert doppelte Verarbeitung)
                session.findById("wnd[0]").sendVKey(11)  # Ctrl+S / Save
                _t.sleep(1.5)

                # Evtl. Bestätigungsdialog schließen
                try:
                    session.findById("wnd[1]").sendVKey(0)
                    _t.sleep(0.5)
                except Exception:
                    pass

                # XD02 schließen → Startmaske
                session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
                session.findById("wnd[0]").sendVKey(0)
                _t.sleep(0.5)

                step["message"] = f"KNB1-ZTERM {zterm_old or '?'}→{new_zterm} BK={bukrs} gesetzt und gesichert"
            except Exception as e:
                step["status"] = "error"
                step["message"] = str(e)
            results.append(step)

        ok   = sum(1 for r in results if r["status"] == "ok")
        warn = sum(1 for r in results if r["status"] == "warn")
        err  = sum(1 for r in results if r["status"] == "error")
        log.info("gui_xd02_kzterm_change: %d ok, %d warn, %d err", ok, warn, err)
        return {
            "status":  "ok" if err == 0 else "partial",
            "ok":      ok,
            "warn":    warn,
            "errors":  err,
            "results": results,
        }
    finally:
        if should_close and connection:
            try:
                connection.CloseSession(session.Id)
            except Exception:
                pass


HANDLERS: dict[tuple[str, str], callable] = {
    ("BAPI", "BAPI_ACC_DOCUMENT_POST"):  bapi_acc_document_post,
    ("BAPI", "BAPI_ACC_DOCUMENT_REV"):   bapi_acc_document_rev,
    ("BAPI", "FBL1N"):                   bapi_vendor_open_items,
    ("BAPI", "FBL5N"):                   bapi_customer_open_items,
    ("BAPI", "RFC_PING"):                rfc_ping,
    ("SQL",  "FBL5N"):                   sql_customer_open_items,
    ("SQL",  "ASE_PING"):                sql_ping,
    ("SQL",  "MB51"):                    sql_mb51,
    ("*",    "MB51"):                    sql_mb51,
    ("Queue","SD_INVOICE"):              queue_sd_invoice_create,
    ("BAPI", "SD_INVOICE"):              bapi_sd_invoice_create,
    ("BAPI", "SD_INVOICE_MULTI"):        bapi_sd_invoice_create_multi,
    ("GUI",  "SD_INVOICE"):              gui_sd_invoice_create,
    ("GUI",  "SD_INVOICE_MULTI"):        gui_sd_invoice_create_multi,
    ("GUI",  "SD_INVOICE_FROM_EXCEL"):   gui_sd_invoice_from_excel,
    ("GUI",  "F110"):                    gui_f110_payment_run,
    ("GUI",  "VA42_ZTERM"):              gui_va42_zterm_change,
    ("GUI",  "XD02_KZTERM"):             gui_xd02_kzterm_change,
    ("Batch","AFAB"):                    batch_afab_depreciation,
    ("BAPI", "STAMMDATEN"):              bapi_stammdaten,
}
