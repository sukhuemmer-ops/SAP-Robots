"""
IV (Eingangsrechnungen) Standalone-Extrakt
==========================================
Zieht alle Eingangsrechnungen (RBKP + RSEG) direkt aus der SAP Sybase ASE-DB
und schreibt das Ergebnis als Excel-Datei.

Kein SAP NW RFC SDK noetig – nur pyodbc + ODBC-Treiber 'Adaptive Server Enterprise'.

Start (im worker-Verzeichnis):
    .venv\\Scripts\\python iv_extract.py
    .venv\\Scripts\\python iv_extract.py --date-from 20260101 --date-to 20260609
    .venv\\Scripts\\python iv_extract.py --lifnr 123456
    .venv\\Scripts\\python iv_extract.py --matnr TEILENR --out C:\\Eigene\\iv.xlsx

Voraussetzungen:
    pip install pyodbc openpyxl python-dotenv   (schon in requirements.txt)
    ODBC-Treiber 'Adaptive Server Enterprise' installiert (SAP ASE SDK)
"""
from __future__ import annotations

import argparse
import io
import logging
import os
import sys
from datetime import datetime
from pathlib import Path

# Windows-Konsole auf UTF-8 setzen
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "buffer"):
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger("iv_extract")

# .env laden
here = Path(__file__).resolve().parent
try:
    from dotenv import load_dotenv
    for candidate in (here / ".env", here / ".env.example"):
        if candidate.exists():
            load_dotenv(candidate)
            log.info("Konfiguration geladen: %s", candidate.name)
            break
except ImportError:
    log.warning("python-dotenv nicht installiert – .env wird nicht geladen.")


def main() -> None:
    today = datetime.today()

    parser = argparse.ArgumentParser(
        description="IV – Eingangsrechnungen aus SAP ASE exportieren (RBKP + RSEG)"
    )
    parser.add_argument("--date-from", default=f"{today.year}0101",
                        help="Buchungsdatum von YYYYMMDD (default: Jahresanfang)")
    parser.add_argument("--date-to",   default=today.strftime("%Y%m%d"),
                        help="Buchungsdatum bis YYYYMMDD (default: heute)")
    parser.add_argument("--lifnr",  default="", help="Lieferantennummer (leer = alle)")
    parser.add_argument("--matnr",  default="", help="Material-Teilstring (leer = alle)")
    parser.add_argument("--werks",  default="", help="Werk-Filter (leer = alle)")
    parser.add_argument("--max-rows", type=int, default=200_000,
                        help="Maximale Zeilen (default: 200000)")
    parser.add_argument("--out", default="",
                        help=r"Ausgabedatei (default: C:\DEV\Tariff-Database\IV_<ts>.xlsx)")
    args = parser.parse_args()

    mandt   = os.getenv("SAP_CLIENT",    "600")
    schema  = os.getenv("SAP_DB_SCHEMA", "SAPSR3")
    pfx     = f"{schema}." if schema else ""
    ts      = today.strftime("%Y%m%d_%H%M%S")
    out_path = args.out or str(Path(r"C:\DEV\Tariff-Database") / f"IV_{ts}.xlsx")

    log.info("Verbinde mit SAP ASE (Host=%s, DB=%s, User=%s) ...",
             os.getenv("SAP_DB_HOST"), os.getenv("SAP_DB_NAME"), os.getenv("SAP_DB_USER"))

    sys.path.insert(0, str(here))
    try:
        from sap_db import get_connection
    except ImportError as exc:
        log.error("sap_db.py nicht gefunden: %s", exc)
        sys.exit(1)

    # ── SQL: RBKP (Kopf) + RSEG (Positionen) ────────────────────────────────
    sql = f"""
        SELECT TOP {args.max_rows}
            h.BELNR   AS Rechnungsnummer,
            h.GJAHR   AS Jahr,
            h.BLDAT   AS Belegdatum,
            h.BUDAT   AS Buchungsdatum,
            h.LIFNR   AS Lieferant,
            h.XBLNR   AS Ext_Referenz,
            h.WAERS   AS Waehrung,
            h.KURSF   AS Wechselkurs,
            i.BUZEI   AS Position,
            i.MATNR   AS Material,
            i.WERKS   AS Werk,
            i.MENGE   AS Menge,
            i.MEINS   AS Mengeneinheit,
            i.WRBTR   AS Betrag_BelegWaehrung,
            i.EBELN   AS Bestellnummer,
            i.EBELP   AS Bestellposition,
            i.SGTXT   AS Positionstext
        FROM {pfx}RBKP h
        JOIN {pfx}RSEG i
          ON h.MANDT = i.MANDT
         AND h.BELNR = i.BELNR
         AND h.GJAHR = i.GJAHR
        WHERE h.MANDT = ?
          AND CONVERT(VARCHAR(8), h.BUDAT, 112) BETWEEN ? AND ?
    """

    def _norm_date(d: str) -> str:
        d = str(d).strip().replace("-", "").replace("/", "")
        return d[:8].zfill(8)

    params: list = [mandt, _norm_date(args.date_from), _norm_date(args.date_to)]
    if args.lifnr:
        sql += "  AND h.LIFNR = ?\n";  params.append(args.lifnr)
    if args.matnr:
        sql += "  AND i.MATNR LIKE ?\n"; params.append(f"%{args.matnr}%")
    if args.werks:
        sql += "  AND i.WERKS = ?\n";  params.append(args.werks)
    sql += "  ORDER BY h.BUDAT DESC, h.BELNR, i.BUZEI"

    log.info("Abfrage: %s bis %s | Lieferant=%s | Material=%s | Werk=%s",
             args.date_from, args.date_to,
             args.lifnr or "*", args.matnr or "*", args.werks or "*")

    try:
        conn = get_connection()
    except Exception as exc:
        log.error("DB-Verbindung fehlgeschlagen: %s", exc)
        log.error("Pruefen: ODBC-Treiber installiert? VPN aktiv? .env korrekt?")
        sys.exit(1)

    try:
        cur = conn.cursor()
        cur.execute(sql, params)
        columns = [desc[0] for desc in cur.description]
        rows    = cur.fetchall()
    except Exception as exc:
        log.error("SQL-Fehler: %s", exc)
        sys.exit(1)
    finally:
        try:
            conn.close()
        except Exception:
            pass

    log.info("%d Zeilen gelesen.", len(rows))
    if not rows:
        log.warning("Keine Daten gefunden – Excel wird trotzdem (leer) erzeugt.")

    # ── Excel erstellen ──────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.error("openpyxl fehlt. Installiere: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Eingangsrechnungen"

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    for ci, col in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    date_cols = {"Belegdatum", "Buchungsdatum"}
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            if columns[ci - 1] in date_cols and val:
                try:
                    val = datetime.strptime(str(val).strip(), "%Y%m%d").strftime("%d.%m.%Y")
                except (ValueError, TypeError):
                    pass
            ws.cell(row=ri, column=ci, value=val)

    for ci, col in enumerate(columns, 1):
        sample = [len(str(ws.cell(row=r, column=ci).value or ""))
                  for r in range(2, min(len(rows) + 2, 52))]
        width = max(len(col), *(sample or [0]), 8) + 2
        ws.column_dimensions[get_column_letter(ci)].width = min(width, 42)
    ws.auto_filter.ref = ws.dimensions

    # Info-Blatt
    info = wb.create_sheet("Abfrage-Info")
    info_rows = [
        ("Transaktion",   "Eingangsrechnungen (RBKP + RSEG)"),
        ("Datenquelle",   f"Sybase ASE – {pfx}RBKP + {pfx}RSEG"),
        ("Mandant",       mandt),
        ("Datum von",     args.date_from),
        ("Datum bis",     args.date_to),
        ("Lieferant",     args.lifnr  or "(alle)"),
        ("Material",      args.matnr  or "(alle)"),
        ("Werk",          args.werks  or "(alle)"),
        ("Zeilen gesamt", len(rows)),
        ("Erstellt am",   today.strftime("%d.%m.%Y %H:%M:%S")),
        ("Ausgabedatei",  out_path),
    ]
    bold = Font(bold=True)
    for label, val in info_rows:
        info.append([label, val])
    for row in info[f"A1:A{len(info_rows)}"]:
        for cell in row:
            cell.font = bold
    info.column_dimensions["A"].width = 20
    info.column_dimensions["B"].width = 60

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    log.info("Fertig! Excel gespeichert: %s", out_path)
    print(f"\nOK: {len(rows)} Eingangsrechnungen exportiert -> {out_path}")


if __name__ == "__main__":
    main()
